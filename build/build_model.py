from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%'
NUM = '#,##0;(#,##0);"-"'
BLUE = Font(name="Arial", size=10, color="0000FF")
BLK = Font(name="Arial", size=10)
GRN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
BOLDW = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=13, bold=True, color="1F3864")
SEC = Font(name="Arial", size=10, bold=True, color="1F3864")
NOTE = Font(name="Arial", size=9, italic=True, color="555555")
YEL = PatternFill("solid", fgColor="FFFF00")
HDR = PatternFill("solid", fgColor="1F3864")
LT = PatternFill("solid", fgColor="E8EDF5")
TOPLINE = Border(top=Side(style="thin", color="1F3864"))

wb = Workbook()

def sheet(name, widths):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def put(ws, cell, val, font=BLK, fmt=None, fill=None, bold=False, wrap=False):
    c = ws[cell]
    c.value = val
    c.font = BOLD if bold else font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if wrap: c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

# ================= README =================
rm = sheet("ReadMe", [110])
put(rm, "A1", "Saving North Middletown Elementary School: Financial Model", TITLE)
put(rm, "A2", "Companion workbook to the report 'Saving North Middletown Elementary School: A Close Look at Bourbon County Schools' (v5.0, August 2026)", NOTE)
rows = [
 "",
 "PURPOSE",
 "This workbook backs up every calculation in the PDF report: the district's three-year General Fund picture, the net-savings test of the",
 "closure proposal, the 'grow the Kings' nonresident-enrollment model, the alternatives menu, the bond schedule, and a reserve-runway",
 "projection. Every output is a live formula; change the blue inputs or yellow judgment cells and the model recalculates.",
 "",
 "HOW TO READ THE CELLS",
 "  Blue text  = a hardcoded input you can edit (audited figures, state data, or scenario levers; source noted alongside).",
 "  Black text = a formula. Do not overwrite.",
 "  Green text = a value pulled from another sheet in this workbook.",
 "  Yellow fill = a key judgment call or estimate the district should replace with actual data. Review these first.",
 "",
 "SOURCES (full citations in the PDF report)",
 "  Audited figures: Bourbon County School District audited financial statements, FY2023-24 and FY2024-25 (posted by KDE).",
 "  Per-pupil spending: Kentucky School Report Card school-level (ESSA) expenditure data, 2023-24.",
 "  SEEK base amounts: Kentucky 2024-2026 and 2026-2028 state budgets. Enrollment/capacity: NCES; 2021 KBE-approved facility plan.",
 "  Every published scenario default (the four facts, the exodus ladder, both calculator defaults, the plan floor/default/top, Eminence, the ledger breakevens) as live formulas: Defaults and Exodus_Model tabs.",
 "  Multi-year school scores and NMES enrollment history: School_Data tab (backs report Figures 1 and 2).",
 "  County demographics and the full 1989-2025 NMES enrollment series: Demographics tab (backs Section 10 and Figure 14).",
 "  Tax rates, fund split, delinquency check, and the recallable options: Tax_History tab (backs Section 10 and Figures 15 and 16).",
 "  Boundary rebalancing and fill-to-capacity scenario: Redistricting tab (backs the Section 9 worked example and Figure 15).",
 "  Bonding capacity components and what closure can and cannot change: Debt_Service tab (backs Section 8).",
 "  Student density, route-mile math, and busing cost scenarios: Transport_Geo tab (backs Appendix B).",
 "",
 "CAVEAT",
 "Prepared by a former NMES King and Bourbon County Colonel working alongside Claude, an AI research assistant from Anthropic. Built from public records and",
 "Open Records Requests only. Estimates are labeled; every figure should be re-verified against the cited primary sources before",
 "formal use. Nothing here alleges misconduct by any official.",
]
r = 3
for t in rows:
    if t in ("PURPOSE", "HOW TO READ THE CELLS", "SOURCES (full citations in the PDF report)", "CAVEAT"):
        put(rm, f"A{r}", t, SEC)
    elif t.startswith("  Blue"):
        put(rm, f"A{r}", t, BLUE)
    elif t.startswith("  Green"):
        put(rm, f"A{r}", t, GRN)
    elif t.startswith("  Yellow"):
        put(rm, f"A{r}", t, BLK, fill=YEL)
    else:
        put(rm, f"A{r}", t, BLK if not t.startswith("  ") else NOTE)
    r += 1

# ================= ASSUMPTIONS =================
a = sheet("Assumptions", [52, 16, 60])
put(a, "A1", "Assumptions & Inputs", TITLE)
put(a, "A2", "Edit blue cells. Yellow cells are judgment calls/estimates; the district should replace them with actuals.", NOTE)

def arow(rr, label, val, fmt=CUR, src="", font=BLUE, fill=None, formula=False):
    put(a, f"A{rr}", label)
    put(a, f"B{rr}", val, font=BLK if formula else font, fmt=fmt, fill=fill)
    if src: put(a, f"C{rr}", src, NOTE)

put(a, "A4", "STATE FUNDING (SEEK base guarantee per pupil)", SEC)
arow(5,  "SEEK base, FY2026", 4586, CUR, "2024-2026 KY budget (HB 6); KDE SEEK files")
arow(6,  "SEEK base, FY2027", 4636, CUR, "Enacted 2026-2028 KY budget; corrected in v5.0 from the House-version $4,626")
arow(7,  "SEEK base, FY2028", 4792, CUR, "2026-2028 KY budget (HB 500)")
arow(8,  "SEEK base, FY2029 (held flat at FY2028)", "=B7", CUR, "Assumption", formula=True)

put(a, "A10", "NORTH MIDDLETOWN ELEMENTARY", SEC)
arow(11, "Enrollment, 2024-25", 128, NUM, "NCES CCD official count. Supt. has said 'around 100'; a '118' figure could not be verified in any official record")
arow(12, "Rated capacity", 174, NUM, "2021 KBE-approved District Facility Plan. A policy output, not a physical constant: the same building held 261 students in 1988-89. See report Section 7")
arow(13, "Open seats", "=B12-B11", NUM, "", formula=True)
arow(14, "Per-pupil spending, total (2023-24)", 19348, CUR, "KY School Report Card, ESSA school-level data")
arow(15, "Per-pupil spending, state/local share", 14173, CUR, "KY School Report Card, ESSA school-level data")
arow(16, "Classroom teachers (FTE)", 9.41, '0.00', "NCES")

put(a, "A18", "DISTRICT GENERAL FUND (audited)", SEC)
arow(19, "Revenues before transfers, FY2023", 27668655, CUR, "FY2024 audit (comparative)")
arow(20, "Revenues before transfers, FY2024", 24952644, CUR, "FY2024 audit")
arow(21, "Revenues before transfers, FY2025", 26449318, CUR, "FY2025 audit")
arow(22, "Expenditures before transfers, FY2023", 27905775, CUR, "FY2024 audit (comparative)")
arow(23, "Expenditures before transfers, FY2024", 27487732, CUR, "FY2024 audit")
arow(24, "Expenditures before transfers, FY2025", 29097404, CUR, "FY2025 audit")
arow(25, "Net transfers & other sources, FY2024", 1469431, CUR, "FY2024 audit")
arow(26, "Net transfers & other sources, FY2025", 1422621, CUR, "FY2025 audit")
arow(27, "Ending fund balance, FY2023", 6582802, CUR, "FY2024 audit")
arow(28, "Ending fund balance, FY2024", 5516305, CUR, "FY2024 audit")
arow(29, "Ending fund balance, FY2025", 4290840, CUR, "FY2025 audit")
arow(30, "Unassigned fund balance, FY2025", 3925193, CUR, "FY2025 audit")
arow(31, "Required contingency (share of spending)", 0.02, PCT, "Kentucky 2% statutory minimum")

put(a, "A33", "LOCAL REVENUE", SEC)
arow(34, "Property assessment, FY2025", 1843569625, CUR, "FY2025 audit MD&A")
arow(35, "GF property tax collected, FY2025", 7829060, CUR, "FY2025 audit")
arow(36, "Annual levy adjustment (KRS 160.470)", 0.04, PCT, "Board lever; up to 4% without recall")
arow(37, "Uncollected (delinquent) property tax, FY2024", 387840, CUR, "Calc yield $10,556,809 vs actual $10,168,969; ~3.7% delinquency, not foregone levy authority")
arow(38, "Uncollected (delinquent) property tax, FY2025", 239126, CUR, "Calc yield $9,880,143 vs actual $9,641,017; ~2.4% delinquency, not foregone levy authority")

put(a, "A40", "COST STRUCTURE", SEC)
arow(41, "Loaded cost per certified position (est.)", 85000, CUR, "Estimate; replace with district payroll data", fill=YEL)
arow(42, "Transportation expense, FY2025", 2913654, CUR, "FY2025 audit")
arow(43, "Transport optimization, low", 0.05, PCT, "Estimate", fill=YEL)
arow(44, "Transport optimization, high", 0.10, PCT, "Estimate", fill=YEL)
arow(45, "District administration expense, FY2023", 999727, CUR, "FY2024 audit")
arow(46, "District administration expense, FY2025", 1447164, CUR, "FY2025 audit")
arow(47, "Administrative rollback share of 2-yr growth", 0.5, PCT, "Judgment call", fill=YEL)
arow(48, "Attrition positions, district-wide", 4, NUM, "Judgment call", fill=YEL)

put(a, "A50", "CLOSURE SCENARIO JUDGMENTS (LEGACY single-point inputs, superseded; the published closure model is the Closure_Model 972-scenario grid at its row 37)", SEC)
arow(51, "School administration avoided (principal, office)", 115397, CUR, "MEASURED v4.5: district MUNIS ledger, FY2026 actuals, org 090 functions 2410+2420 (build/munis_extract.py). Was $131,724 from the working budget (program view; the function view prints $132,744, School_Costs!G32) through the prior release.", fill=YEL)
arow(52, "Plant operations avoided (custodial + building costs)", 128866, CUR, "MEASURED v4.5: MUNIS ledger, org 090 function 2610 actuals: custodial pay and benefits $49,655 + building utilities, disposal, telecom, supplies and repairs $79,211. Was $96,107 from the working budget.", fill=YEL)
arow(53, "Teaching positions truly eliminated", 3, NUM, "Estimate; via attrition only", fill=YEL)
arow(54, "Added busing cost per year", 137500, CUR, "Estimate; midpoint of $75K-$200K", fill=YEL)
arow(55, "Students leaving the district on closure", 10, NUM, "Judgment call; see sensitivity table", fill=YEL)
arow(56, "One-time transition cost", 100000, CUR, "Estimate", fill=YEL)

put(a, "A58", "GROWTH SCENARIO (nonresident enrollment under HB 563 / KRS 157.350)", SEC)
arow(59, "Transfer students, Year 1 (FY2027)", 15, NUM, "Scenario lever")
arow(60, "Transfer students, Year 2 (FY2028)", 30, NUM, "Scenario lever")
arow(61, "Transfer students, Year 3 (FY2029)", 46, NUM, "Scenario lever; capped at open seats in model")
arow(62, "Variable cost per transfer student", 400, CUR, "Estimate; supplies/materials", fill=YEL)
arow(63, "Added teacher once transfers exceed", 30, NUM, "Judgment call", fill=YEL)

put(a, "A65", "DISTRICT-FAVORABLE CLOSURE CASE (red-team upper bound)", SEC)
arow(66, "Positions eliminated, favorable case", 5, NUM, "Upper bound tested in Closure_Model", fill=YEL)
arow(67, "Added busing, favorable (low) case", 75000, CUR, "Low end of the $75K-$200K range", fill=YEL)
arow(69, "GF-borne loaded cost per certified position (v3)", 60000, CUR, "Published salary schedule $41,718-$71,447; state pays TRS/KEHP on-behalf; GF keeps salary + ~5%. Range $50K-$75K", fill=YEL)


# ================= DEFAULTS (published scenario defaults, all live) =================
df = sheet("Defaults", [58, 15, 15, 15, 15, 60])
put(df, "A1", "Published Scenario Defaults: every headline number on the site, executive summary, and report, as live formulas", TITLE)
put(df, "A2", "One row per published default. Blue cells are sourced inputs; black cells compute. If a formula here disagrees with a published artifact, the artifact is wrong: say so and it will be corrected publicly.", NOTE)

put(df, "A4", "FACT ONE AND TWO: THE SCHOOL AND ITS COST (Sections 2 and 3)", SEC)
put(df, "A5", "NMES per-student cost, newest state file (2024-25)"); put(df, "B5", 17903, BLUE, CUR)
put(df, "A6", "Kentucky elementary average, same file"); put(df, "B6", 19299, BLUE, CUR)
put(df, "A7", "NMES below the state average"); put(df, "B7", "=1-B5/B6", BLK, PCT, bold=True); put(df, "F7", "Published: 7 percent below", NOTE)
put(df, "A8", "District's own cost-of-delivery table, May 21, 2026: NMES / state average"); put(df, "B8", 19080, BLUE, CUR); put(df, "C8", 19020, BLUE, CUR)
put(df, "A9", "Gap on the district's own table"); put(df, "B9", "=B8/C8-1", BLK, PCT); put(df, "F9", "Published: three tenths of one percent", NOTE)
put(df, "A10", "Share of middle and high schoolers who came through NMES (128 of the 1,040 elementary seats)"); put(df, "B10", "=Assumptions!B11/1040", BLK, PCT); put(df, "F10", "Published: about one in eight", NOTE)

put(df, "A12", "FACT THREE: THE LEDGER WALK AND THE $661,139 CLAIM (Section 4)", SEC)
put(df, "A13", "All-funds dollars coded to org 090, fiscal 2025 MUNIS Cost by ORG"); put(df, "B13", 1285310, BLUE, CUR)
put(df, "A14", "Non-General-Fund dollars in that total (grants, on-behalf and other funds, per the ledger's fund split)"); put(df, "B14", 182952, BLUE, CUR); put(df, "C14", 161880, BLUE, CUR); put(df, "D14", 6941, BLUE, CUR)
put(df, "A15", "General Fund actuals coded to the school"); put(df, "B15", "=B13-B14-C14-D14", BLK, CUR, bold=True); put(df, "F15", "Published: $933,537 (within 0.55 percent of the $938,690 budget view)", NOTE)
put(df, "A16", "Building-bound costs on the ledger (utilities, disposal, telecom, supplies, repairs)"); put(df, "B16", 79211, BLUE, CUR)
put(df, "A17", "District worksheet: avoidable expense lines / insurance"); put(df, "B17", 107039, BLUE, CUR); put(df, "C17", 20000, BLUE, CUR)
put(df, "A18", "Staff-retained saving (ledger) and building-sold ceiling (worksheet)"); put(df, "B18", "=B16", BLK, CUR, bold=True); put(df, "C18", "=B17+C17", BLK, CUR, bold=True)
put(df, "A19", "The district's $661,139 decomposition: staffing + avoidable lines + insurance + supplies"); put(df, "B19", 493407, BLUE, CUR); put(df, "C19", "=B17", BLK, CUR); put(df, "D19", "=C17", BLK, CUR); put(df, "E19", 40693, BLUE, CUR)
put(df, "A20", "Sum (the claim) and what remains in year one with staff retained and supplies moving"); put(df, "B20", "=B19+C19+D19+E19", BLK, CUR, bold=True); put(df, "C20", "=C18", BLK, CUR, bold=True)

put(df, "A22", "FACT FOUR, PRE-v5 CONVENTION (superseded by the Exodus_Model tab; kept for the version record)", SEC)
put(df, "A23", "Per-leaver funding: SEEK base + typical add-ons"); put(df, "B23", "=Assumptions!B6+500", BLK, CUR); put(df, "F23", "$5,136 at the FY2027 base", NOTE)
put(df, "A24", "Share who leave / students (of 128) / year one / per year at full effect / total by grade 12", bold=True)
for i, pct in enumerate([0.10, 0.20, 0.30, 0.50]):
    rr = 25 + i
    put(df, f"A{rr}", f"{int(pct*100)} percent of the school")
    put(df, f"B{rr}", f"=ROUND(128*{pct},0)", BLK, NUM)
    put(df, f"C{rr}", f"=B{rr}*$B$23", BLK, CUR)
    put(df, f"D{rr}", f"=ROUND(B{rr}*13/6,0)*$B$23", BLK, CUR)
    put(df, f"E{rr}", f"=128*{pct}*$B$23*23.5", BLK, CUR)
put(df, "F24", "Year one and full effect price whole students (the site chart's convention); the grade-12 total prices the exact share times the 23.5-year factor: (6+7+8+9+10+11+12+6x13)/6 years of escalating cohorts. Reproduces the published $67K/$144K/$1.5M through $328K/$713K/$7.7M rows.", NOTE, wrap=True)
put(df, "A30", "Carried funding of the current 128 students: grade counts (K-5) x years to grade 12"); 
for i, (cnt, yrs) in enumerate(zip([22, 22, 19, 22, 16, 27], [13, 12, 11, 10, 9, 8])):
    put(df, get_column_letter(2+i) + "30", cnt, BLUE, NUM)
put(df, "A31", "Student-years, and the carried band at the $4,636 base / $5,136 with add-ons")
put(df, "B31", "=SUMPRODUCT(B30:G30,{13,12,11,10,9,8})", BLK, NUM, bold=True)
put(df, "C31", "=B31*Assumptions!B6", BLK, CUR); put(df, "D31", "=B31*(Assumptions!B6+500)", BLK, CUR)
put(df, "F31", "Published: 1,339 student-years, $6.2 to $6.9 million", NOTE)

put(df, "A33", "THE TWO CALCULATOR DEFAULTS (the numbers on the site's cards)", SEC)
put(df, "A34", "Closure default: scaled savings granted (teachers included, half the fixed overhead) + the statistical median of missing students")
put(df, "B34", "=B17+C17+Closure_Model!C40+3*54479.4-63000-136*(Assumptions!B6+500-Assumptions!B62)", BLK, CUR, bold=True)
put(df, "F34", "= $127,039 kept, plus half the fixed overhead positions ($107,052, Closure_Model C40) and three teachers at $54,479.40, minus $63,000 of busing, minus the statistical median of 136 missing students at $5,136 less the $400 supplies credit: the published -$309,567, the 82nd percentile of the 972-scenario grid; the weighted median is a $428,627 loss (Closure_Model row 50)", NOTE, wrap=True)
put(df, "A35", "Growth default: 30 added students at the grid's low busing and supplies")
put(df, "B35", "=30*(Assumptions!B6+500-400)", BLK, CUR, bold=True)
put(df, "F35", "The published +$142,080, within $140 of the 19,683-scenario grid's weighted median of $142,220 (Growth_Model rows 17-19); at the central $500 busing and $700 supplies the same 30 students net about $118,000", NOTE, wrap=True)

put(df, "A37", "THE PLAN: FLOOR, DEFAULT, AND TOP (the site plan calculator's three published cases)", SEC)
put(df, "A38", "", ); put(df, "B38", "Floor", BOLD); put(df, "C38", "Default", BOLD); put(df, "D38", "Top", BOLD)
put(df, "A39", "Recovered leakage students (of the 550-student pool)"); put(df, "B39", 0, BLUE, NUM); put(df, "C39", 275, BLUE, NUM); put(df, "D39", 550, BLUE, NUM)
put(df, "A40", "Enrollment lever at $4,236 net of supplies"); 
for col in "BCD": put(df, f"{col}40", f"={col}39*(Assumptions!B6-Assumptions!B62)", BLK, CUR)
put(df, "A41", "Counted-once fixed-cost package"); put(df, "B41", 760000, BLUE, CUR); put(df, "C41", 760000, BLUE, CUR); put(df, "D41", 1300000, BLUE, CUR)
put(df, "A42", "Full 2018 rate restore (live from Tax_History)"); 
for col in "BCD": put(df, f"{col}42", "=Tax_History!D79", BLK, CUR)
put(df, "A43", "Trending fiscal 2026 gap (June 2026 ledger, before transfers)"); 
for col in "BCD": put(df, f"{col}43", 1738653, BLUE, CUR)
put(df, "A44", "Surplus after the gap", bold=True)
for col in "BCD": put(df, f"{col}44", f"={col}40+{col}41+{col}42-{col}43", BLK, CUR, bold=True)
put(df, "A45", "5 percent certified raise ($10,000,388 GF certified payroll x 5% x 1.0145)"); 
for col in "BCD": put(df, f"{col}45", "=10000388*0.05*1.0145", BLK, CUR)
put(df, "A46", "Left for new debt service (negative at the floor: within $7,000 of the raise)")
for col in "BCD": put(df, f"{col}46", f"={col}44-{col}45", BLK, CUR)
put(df, "A47", "New bonds at 4.5 percent, 20 years (factor below), floored at zero")
for col in "BCD": put(df, f"{col}47", f"=MAX({col}46,0)*$B$50", BLK, CUR)
put(df, "A48", "Building capacity with the advisor's $32 million", bold=True)
for col in "BCD": put(df, f"{col}48", f"={col}47+32000000", BLK, CUR, bold=True)
put(df, "F44", "Published: about $500,000 / $1,665,325 / $3.4 million to spare", NOTE)
put(df, "F48", "Published: the advisor's $32 million at the floor; about $47 million at the default; about $69 million at the top", NOTE)
put(df, "A50", "Annuity factor, 4.5 percent, 20 years"); put(df, "B50", "=(1-1.045^-20)/0.045", BLK, '0.000')

put(df, "A52", "GROWTH CONTEXT DEFAULTS (Section 10)", SEC)
put(df, "A53", "Eminence Independent, 2014 to 2024 enrollment, and its growth"); put(df, "B53", 733, BLUE, NUM); put(df, "C53", 991, BLUE, NUM); put(df, "D53", "=C53/B53-1", BLK, PCT)
put(df, "A54", "Bourbon County Schools, 2014 to 2023 enrollment, and its decline"); put(df, "B54", 2912, BLUE, NUM); put(df, "C54", 2616, BLUE, NUM); put(df, "D54", "=C54/B54-1", BLK, PCT)
put(df, "F53", "Published: grew 35 percent in the decade Bourbon Schools shrank 10", NOTE)
put(df, "A55", "Leakage pool pricing: see Redistricting rows 136-140 (documented floor 483; gross $2.1-$2.5M at the full base; net $1.9-$2.3M at $4,236)", NOTE, wrap=True)
put(df, "A56", "Rate menu: see Tax_History rows 70-89 ($166,189 per real cent; the four options and the sequencing)", NOTE, wrap=True)

put(df, "A58", "BREAKEVEN TABLE, LEDGER-CODED VIEW (Section 4; the all-in view is on School_Costs)", SEC)
put(df, "A59", "Revenue per member (all-funds spending of about $41.8M across about 2,615 members)"); put(df, "B59", 15983, BLUE, CUR)
put(df, "A60", "School / coded by the district / enrolled / breakeven / clears by", bold=True)
for i, (nm, cost, enr) in enumerate([("North Middletown", 1285310, 128), ("Bourbon Central", 4033689, 459),
                                      ("Cane Ridge", 4326733, 453), ("Bourbon County Middle", 3868106, 590),
                                      ("Bourbon County High", 5515105, 766)]):
    rr = 61 + i
    put(df, f"A{rr}", nm); put(df, f"B{rr}", cost, BLUE, CUR); put(df, f"C{rr}", enr, BLUE, NUM)
    put(df, f"D{rr}", f"=ROUND(B{rr}/$B$59,0)", BLK, NUM); put(df, f"E{rr}", f"=C{rr}-D{rr}", BLK, NUM)
put(df, "F60", "Every school clears on the ledger-coded definition; the all-in report-card definition fails every school. The swing is the definitions, which is why they should be chosen and published before any vote.", NOTE, wrap=True)


put(df, "A67", "PEER YARDSTICK: FAYETTE COUNTY, SAME FISCAL YEAR (Section 7)", SEC)
put(df, "A68", ""); put(df, "B68", "Bourbon", BOLD); put(df, "C68", "Fayette (audited)", BOLD); put(df, "D68", "Fayette (adjusted)", BOLD)
put(df, "A69", "General Fund expenditures, FY2025"); put(df, "B69", 29097404, BLUE, CUR); put(df, "C69", 685348803, BLUE, CUR); put(df, "D69", 690460223, BLUE, CUR)
put(df, "A70", "Operating gap before transfers"); put(df, "B70", 2648086, BLUE, CUR); put(df, "C70", 38907376, BLUE, CUR); put(df, "D70", "", BLK)
put(df, "A71", "Gap per dollar spent"); 
for col in "BC": put(df, f"{col}71", f"={col}70/{col}69", BLK, PCT)
put(df, "A72", "Fund balance, beginning of year"); put(df, "B72", 5516305, BLUE, CUR); put(df, "C72", 43291115, BLUE, CUR); put(df, "D72", "n/a", NOTE)
put(df, "A73", "Fund balance, end of year"); put(df, "B73", 4290840, BLUE, CUR); put(df, "C73", 28361786, BLUE, CUR); put(df, "D73", 6902403, BLUE, CUR)
put(df, "A74", "Reserve per dollar spent (the published cushion)", bold=True)
for col in "BCD": put(df, f"{col}74", f"={col}73/{col}69", BLK, PCT, bold=True)
put(df, "A75", "Drawdown during the year")
for col in "BC": put(df, f"{col}75", f"={col}72-{col}73", BLK, CUR)
put(df, "A76", "Drawdown per dollar spent (the burn rate)", bold=True)
for col in "BC": put(df, f"{col}76", f"={col}75/{col}69", BLK, PCT, bold=True)
put(df, "A77", "Years of cushion on the net-change basis (flattered by transfers in)", bold=True)
for col in "BC": put(df, f"{col}77", f"={col}73/{col}75", BLK, '0.0', bold=True)
put(df, "A78", "Difference between the audited ending balance and Weaver's post-adjustment figure for the same date (not reconciled in Weaver's deck; the two rest on different bases)"); put(df, "C78", "=C73-D73", BLK, CUR, bold=True)
put(df, "A81", "Net other financing in (the gap less the net change): the transfers that flattered the burn"); 
for col in "BC": put(df, f"{col}81", f"={col}70-{col}75", BLK, CUR)
put(df, "A82", "OF WHICH the capital-to-General-Fund sweep (Bourbon; the plan ends it)"); put(df, "B82", 1320939, BLUE, CUR)
put(df, "A83", "Years of cushion on OPERATIONS ALONE (reserve / gap before transfers): the planning number", bold=True)
for col in "BC": put(df, f"{col}83", f"={col}73/{col}70", BLK, '0.0', bold=True)
put(df, "F83", "Bourbon about 1.6 years, Fayette about 0.7 on their audited books. This is the basis the published comparison leads with, because the plan ends the sweep and the sweep is what holds the net-change burn down to 4.2 cents. Weaver's unaudited view puts Fayette's cushion nearer 1 percent of spending.", NOTE, wrap=True)
put(df, "A79", "KRS 160.470(6)(a) 2 percent minimum reserve, on Fayette's adjusted spending"); put(df, "D79", "=D69*0.02", BLK, CUR)
put(df, "A80", "Fayette's own 6 percent administrative threshold"); put(df, "D80", "=D69*0.06", BLK, CUR)
put(df, "F69", "Bourbon and Fayette FY2025 audits, both archived under build/. The Fayette 'adjusted' column is Weaver, L.L.P.'s Audit of Budget Processes and Expenditures, presented to the Fayette board August 3, 2026 (build/fcps_weaver_audit_2026_08.pdf, Finding A.4): after unaudited corrections the district made to its own FY2025 ledger in June 2026, the ending balance was about $6,902,403, roughly 1 percent of $690,460,223. Weaver labels those figures unaudited and subject to change; they are carried here the same way.", NOTE, wrap=True)
put(df, "F74", "Published: Bourbon 14.7 cents of reserve per dollar, Fayette 4.1 on its audit and about 1.0 after its own adjustments", NOTE)
put(df, "F77", "On the net-change basis Bourbon shows about three and a half years and Fayette under two, but both are flattered by transfers in; row 83 gives the operations-alone runway the artifacts publish. Drawdown and runway are computed on audited figures only; Weaver's waterfall runs from the planned contingency rather than the audited opening balance, so its closing figure is not differenced against an audited opening here.", NOTE, wrap=True)
put(df, "F79", "Weaver: the adjusted balance sits below both the statutory minimum every Kentucky district must budget and Fayette's own 6 percent policy. Weaver's 70-plus recommendations and 10 priority actions concern budget controls, forecasting, reconciliation, reporting and reserve replenishment; none is a school closure.", NOTE, wrap=True)

# ================= GF_SUMMARY =================
g = sheet("GF_Summary", [46, 15, 15, 15])
put(g, "A1", "General Fund, Three-Year Summary (audited)", TITLE)
for col, yr in zip("BCD", ["FY2023", "FY2024", "FY2025"]):
    put(g, f"{col}3", yr, BOLDW, fill=HDR)
put(g, "A3", "", BOLDW, fill=HDR)
lines = [
 ("Revenues before transfers", ["=Assumptions!B19", "=Assumptions!B20", "=Assumptions!B21"], GRN),
 ("Expenditures before transfers", ["=Assumptions!B22", "=Assumptions!B23", "=Assumptions!B24"], GRN),
 ("Operating result before transfers", ["=B4-B5", "=C4-C5", "=D4-D5"], BLK),
 ("Net transfers & other sources", [None, "=Assumptions!B25", "=Assumptions!B26"], GRN),
 ("Change in fund balance", [None, "=C6+C7", "=D6+D7"], BLK),
 ("Ending fund balance", ["=Assumptions!B27", "=Assumptions!B28", "=Assumptions!B29"], GRN),
 ("Check vs audited balances (small residual = other audited items)", [None, "=C9-B9-C8", "=D9-C9-D8"], BLK),
]
r = 4
for label, vals, f in lines:
    put(g, f"A{r}", label)
    for col, v in zip("BCD", vals):
        if v is not None:
            put(g, f"{col}{r}", v, font=f, fmt=CUR)
    r += 1
put(g, "A12", "METRICS", SEC)
put(g, "A13", "Fund balance as % of expenditures, FY2025"); put(g, "D13", "=D9/D5", BLK, PCT)
put(g, "A14", "2% contingency floor (FY2025 spending)"); put(g, "D14", "=Assumptions!B31*D5", BLK, CUR)
put(g, "A15", "Unassigned balance above the floor"); put(g, "D15", "=Assumptions!B30-D14", BLK, CUR)
put(g, "A16", "Average annual drawdown (FY2024-25)"); put(g, "D16", "=-AVERAGE(C8:D8)", BLK, CUR)
put(g, "A17", "Years of runway at current pace"); put(g, "D17", "=IFERROR(D15/D16,0)", BLK, '0.0')
put(g, "A19", "Source: FY2023-24 and FY2024-25 audited financial statements. FY2023 shown as reported in the FY2024 audit's comparative statement.", NOTE)
put(g, "A20", "What the transfers are: moves between the district's own funds (indirect cost recoveries from grants and self-supporting operations, fund closeouts, and similar interfund items detailed in the audits' fund statements). They cushion the General Fund's bottom line but are not new district revenue, which is why the operating result BEFORE transfers is the honest measure of the structural deficit.", NOTE, wrap=True)

# ================= CLOSURE_MODEL =================
c = sheet("Closure_Model", [54, 16, 46])
put(c, "A1", "Closure of NMES: Net-Savings Test", TITLE)
put(c, "A2", "Gross site cost is not net saving; students, teachers, and their SEEK funding move to receiving schools.", NOTE)
put(c, "A4", "THE CLAIM AND THE OFFICIAL DATA", SEC)
put(c, "A5", "Superintendent's stated gross cost (public statement, not yet documented)")
put(c, "B5", 1000000, BLUE, CUR); put(c, "C5", "WKYT, July 2026; treated as a claim to verify", NOTE)
put(c, "A6", "Total site spending (state ESSA basis)"); put(c, "B6", "=Assumptions!B14*Assumptions!B11", GRN, CUR)
put(c, "A7", "State/local share of site spending"); put(c, "B7", "=Assumptions!B15*Assumptions!B11", GRN, CUR)
put(c, "A9", "RECURRING SAVINGS (costs that truly disappear)", SEC)
put(c, "A10", "Principal & office"); put(c, "B10", "=Assumptions!B51", GRN, CUR)
put(c, "A11", "Plant, utilities, insurance (net of carrying cost)"); put(c, "B11", "=Assumptions!B52", GRN, CUR)
put(c, "A12", "Teaching positions eliminated (via attrition, GF-borne cost)"); put(c, "B12", "=Assumptions!B53*Assumptions!B69", GRN, CUR)
put(c, "C12", "v3: GF-borne $60K per position (state pays TRS/KEHP on-behalf); prior versions used the $85K all-in figure here", NOTE, wrap=True)
put(c, "A13", "Gross avoidable cost", bold=True); put(c, "B13", "=SUM(B10:B12)", BLK, CUR, bold=True)
put(c, "A15", "RECURRING OFFSETS (new costs and lost revenue)", SEC)
put(c, "A16", "Added busing"); put(c, "B16", "=Assumptions!B54", GRN, CUR)
put(c, "A17", "SEEK revenue lost to departing students (FY2027 base)"); put(c, "B17", "=Assumptions!B55*Assumptions!B6", GRN, CUR)
put(c, "A18", "Total offsets", bold=True); put(c, "B18", "=SUM(B16:B17)", BLK, CUR, bold=True)
put(c, "A20", "NET RECURRING GENERAL FUND SAVING (LEGACY single-point scenario on superseded inputs; the published model is the 972-scenario grid at row 37, median a $428,627 LOSS)", bold=True)
nc = put(c, "B20", "=B13-B18", BLK, CUR, bold=True); nc.border = TOPLINE
put(c, "A21", "Share of the structural deficit ($2.65M) | of the reserve drawdown ($1.15M)")
put(c, "B21", "=B20/(Assumptions!B24-Assumptions!B21)", BLK, PCT)
put(c, "C21", "=B20/GF_Summary!D16", BLK, PCT)
put(c, "A22", "One-time transition cost (year one)"); put(c, "B22", "=Assumptions!B56", GRN, CUR)
put(c, "A24", "SENSITIVITY: STUDENTS LEAVING THE DISTRICT", SEC)
put(c, "A25", "Students leaving", bold=True); put(c, "B25", "Net recurring saving", bold=True)
for i, n in enumerate([0, 10, 20, 30]):
    rr = 26 + i
    put(c, f"A{rr}", n, BLUE, NUM)
    put(c, f"B{rr}", f"=$B$13-Assumptions!$B$54-A{rr}*Assumptions!$B$6", BLK, CUR)
put(c, "A31", "Each departing student removes at least the SEEK base guarantee, every year, permanently.", NOTE)
put(c, "A33", "V3 CORRECTION: WHAT A POSITION ACTUALLY COSTS THE GENERAL FUND", SEC)
put(c, "A34", "All-in cost per position (salary + state-paid on-behalf; filing basis)"); put(c, "B34", "=Assumptions!B41", GRN, CUR)
put(c, "C34", "Correct for KDE per-pupil comparisons; the district books $6.94M of on-behalf in FY2026", NOTE, wrap=True)
put(c, "A35", "GF-borne cost per position (salary + ~5%)"); put(c, "B35", "=Assumptions!B69", GRN, CUR)
put(c, "C35", "Published schedule: Rank III $41,718 (yr 0) to Rank I $71,447 (yr 29-30). The state pays TRS and KEHP on behalf of districts; eliminating a GF position saves the GF only $50K-$75K", NOTE, wrap=True)
put(c, "A37", "V5.0 TWO-TAILED SENSITIVITY: SIX LEVERS, 972 COMBINATIONS (backs Figure 5; THIS GRID, not the legacy single-point rows above, is the published closure model)", SEC)
put(c, "A38", "Lever (low / central / high)", BOLDW, fill=HDR); put(c, "B38", "Low", BOLDW, fill=HDR); put(c, "C38", "Central", BOLDW, fill=HDR); put(c, "D38", "High", BOLDW, fill=HDR); put(c, "E38", "Source", BOLDW, fill=HDR)
v3levers = [
 ("Non-salary capture (their worksheet, + insurance at the full stop)", 53519, 80279, 127039, "District Response Appendix A: $107,039 of building-bound lines (utilities, telecom, maintenance, custodial supplies) captured at 50/75/100 percent, plus its ~$20,000 insurance figure at the full stop. The worksheet's other $40,693 (supplies, books, field trips, printing = $318/student vs our measured $331) travels with the students."),
 ("Fixed positions cut over time (school admin + custodial + library)", 0, 107052.2, 214104.4, "MUNIS FY2026 actuals: school administration $115,397 + custodial $49,655 + library $49,052 = $214,104 (the district's own A.1 prices the same four roles at $209,700, within 2 percent, and states all current staff are retained in year one; the full-cut leg is an attrition end state)"),
 ("Teachers cut (grid legs 0/1/2/3) x $54,479.40 each", 0, 2, 3, "Priced at the district's OWN fully loaded 0-years-experience figure, $54,479.40 (Response Appendix A.1: 'Elementary Teachers: 2, $108,958.80'). The top leg credits 3 because Appendix B's own classroom count eliminates three homerooms net."),
 ("Added busing", 20000, 63000, 95000, "Derived bottom-up with uncertainty: 2-4 zone buses terminating in Paris (~9-11 road miles farther one-way), 2 loaded + 0-2 deadhead legs daily, 170 to 175 days, $3.25-$4.75/mile (KDE/NAPT band). The bottom-up maximum with a route split reached $190,000; v5.0 caps the high leg at half that, $95,000. The July 2026 records response produced the current routes but answered N/A for any routing study or ride-time analysis."),
 ("Students missing from the rolls at steady state (grid legs 117/136/154)", 117, 136, 154, "From the August 2026 school-choice survey (anonymized in build/) and build/exodus_model.py: the response-bias-corrected posterior 25th/50th/75th percentiles of the leave share, applied to the whole feeder stream (entering class of 21.5, the recent 19-24 SAAR midpoint, x 12.62 effective years of the district's own grade-to-grade survival), triangular with the median central. The 31 signed households and their 70 children are the hard evidence behind the share, not a priced scenario. Exits free and funded under HB 563. The state's SAAR files corroborate: kindergarten 12 in 2025-26 against a ten-year average of 22; end-of-year 141/128/115 across 2023-24 to 2025-26."),
 ("SEEK add-ons lost per leaver", 0, 500, 1000, "At-risk weight (15% of base on a ~72% FRL school), exceptional-child weights, transportation component, $100 capital outlay"),
]
for i, (lbl, lo, ce, hi, src) in enumerate(v3levers):
    rr = 39 + i
    put(c, f"A{rr}", lbl); put(c, f"B{rr}", lo, BLUE, CUR if lo > 100 else NUM); put(c, f"C{rr}", ce, BLUE, CUR if ce > 100 else NUM); put(c, f"D{rr}", hi, BLUE, CUR if hi > 100 else NUM); put(c, f"E{rr}", src, NOTE)
put(c, "A47", "Central case: net yearly effect")
put(c, "B47", "=C39+C40+C41*54479.4-C42-C43*(Assumptions!B6+C44-Assumptions!B62)", BLK, CUR, bold=True)
put(c, "C47", "-$410,806: the central case itself loses money, about 16 percent of the structural deficit added, not removed", NOTE)
put(c, "A48", "Unfavorable tail (all levers adverse)")
put(c, "B48", "=B39+B40+B41*54479.4-D42-D43*(Assumptions!B6+D44-Assumptions!B62)", BLK, CUR)
put(c, "C48", "-$847,825 a year: the closure loses money", NOTE)
put(c, "A49", "Favorable tail (all levers favorable)")
put(c, "B49", "=D39+D40+D41*54479.4-B42-B43*(Assumptions!B6+B44-Assumptions!B62)", BLK, CUR)
put(c, "C49", "-$11,030 a year: even the grid's best case (every lever at its closure-friendliest at once) still loses money, roughly $800,000 to $1 million short of the plan's requirement", NOTE)
put(c, "A50", "Distribution of all 972 combinations enumerated by build/closure_grid.py: capture, fixed-position, add-ons, busing and missing-students levers take three values each, teachers four (0/1/2/3): 3^5 x 4 = 972. Each missing student is priced at the enacted FY2027 SEEK base of $4,636 plus the add-ons lever, minus the $400 of supplies that stop being spent (Assumptions row 62, the same figure the growth model charges each recruit); teacher savings appear ONLY on the teachers-cut lever, so staffing is never counted twice. The busing high leg is capped at half its bottom-up maximum ($95,000); a property-value lever priced in an interim draft is removed while the PVA records request is pending; the missing-students lever prices the statistical band's quartiles (117/136/154), with the signed-survey floor of 74 kept as hard evidence below every priced leg. v5.0 lever weights: triangular 1-2-1 on every three-leg lever, uniform on teachers. Weighted median -$428,627 (the median scenario LOSES money); middle half -$519,765 to -$340,021; EVERY weighted scenario is negative, the best case still loses $11,030; range -$847,825 to -$11,030 (unweighted median -$426,503). One-time transition costs $100K-$300K in year one are additional. The v4.5 grid (5,832 scenarios, weighted median -$20,007, 55 percent negative) and the v3.9 grid (2,916, median +$21,571) are retained in the version history.", NOTE, wrap=True)
put(c, "A52", "HOSTILE PAPER CASE, PUBLISHED WITH ITS REFUTATION", SEC)
put(c, "A53", "Every absorbed student priced at the $9,848 slope (withdrawn in v3.9; kept for the record)")
put(c, "B53", "=Assumptions!B14*Assumptions!B11-128*9848-137500-10*Assumptions!B6", BLK, CUR)
put(c, "C53", "About $1.03M on paper, and the $9,848 slope itself was withdrawn in v3.9 (Redistricting row 111). Kept because the refutation stands either way: it requires Cane Ridge at 525 students against its approved 422 rating and Bourbon Central at 555 against 521: exactly the overcrowding the $14M renovation exists to cure. The savings pre-spend the bond.", NOTE, wrap=True)
put(c, "A55", "MILLERSBURG, CLOSED 2006 (2006-07 CCD FILE; STUDENTS TO THE 2007 CANE RIDGE ADDITION): THE COUNTY'S OWN PRECEDENT (backs Figure 6)", SEC)
put(c, "A56", "Millersburg Elementary final enrollment (fall 2005; last operated SY 2005-06, closed status in the 2006-07 federal CCD file; distinct from the private military institute, which closed permanently July 2006)"); put(c, "B56", 119, BLUE, NUM)
put(c, "C56", "NCES CCD; series 153-133-145-139-137-127-129-119; closed 2007; students to Cane Ridge (2007 addition)", NOTE, wrap=True)
put(c, "A57", "Millersburg population 1980/1990/2000/2010/2020"); put(c, "B57", "987 / 937 / 842 / 792 / 747", NOTE)
put(c, "C57", "Decennial census; down 11% from 2000 while the county grew 4.6% (19,360 to 20,252)", NOTE, wrap=True)
put(c, "A58", "Town-budget blow from the 2013 Joy Global closure"); put(c, "B58", 100000, BLUE, CUR)
put(c, "C58", "Community Ventures account: 197 jobs lost; $100K/yr payroll tax, half the town budget", NOTE, wrap=True)
put(c, "A59", "Records asks: the 2006 closure's savings analysis and realized savings; the Millersburg Elementary building's deed, sale price, and current condition; PVA assessed-value history Millersburg vs county.", NOTE, wrap=True)

# ================= GROWTH_MODEL =================
ex = sheet("Exodus_Model", [56, 14, 14, 14, 14, 60])
put(ex, "A1", "THE EXODUS MODEL: WHAT LEAVING FAMILIES COST, FROM THE AUGUST 2026 SCHOOL-CHOICE SURVEY", SEC)
put(ex, "A2", "Every figure reproduced by build/exodus_model.py from the anonymized survey (build/survey_school_choice_2026_08_anonymized.csv). Names and dates are never published.", NOTE, wrap=True)
put(ex, "A4", "THE SURVEY, CLEANED", SEC)
put(ex, "A5", "Households answering / children covered"); put(ex, "B5", 38, BLUE, NUM); put(ex, "C5", 85, BLUE, NUM)
put(ex, "A6", "Leaving households / their children"); put(ex, "B6", 31, BLUE, NUM); put(ex, "C6", 70, BLUE, NUM)
put(ex, "A7", "Staying children / already left the district"); put(ex, "B7", 13, BLUE, NUM); put(ex, "C7", 2, BLUE, NUM)
put(ex, "A8", "Current student population (last official SAAR end-of-year count, 2024-25)"); put(ex, "B8", 128, BLUE, NUM)
put(ex, "A9", "Entering class per year (midpoint of the recent 19-24 SAAR per-grade range)"); put(ex, "B9", 21.5, BLUE, NUM); put(ex, "F9", "The ten-year average NMES kindergarten (SAAR school files, 2015-16 to 2025-26 observed seasons) is 22.2; the survey's hand-coded class years mark only who is enrolled now, and no per-class math is used (v5.0 method change)", NOTE, wrap=True)
put(ex, "A11", "EFFECTIVE YEARS PER LOST CHILD (district grade-to-grade survival, SAAR school files)", SEC)
put(ex, "A12", "Survival vs own 5th-grade class: 6th / 7th / 8th"); put(ex, "B12", 0.972, BLUE, NUM); put(ex, "C12", 0.947, BLUE, NUM); put(ex, "D12", 0.974, BLUE, NUM)
put(ex, "A13", "9th / 10th / 11th"); put(ex, "B13", 0.980, BLUE, NUM); put(ex, "C13", 0.960, BLUE, NUM); put(ex, "D13", 0.960, BLUE, NUM)
put(ex, "A14", "12th (early graduates, dropouts, moves)"); put(ex, "B14", 0.826, BLUE, NUM)
put(ex, "A15", "Effective enrollment-years per lost child (6 elementary + weighted secondary)")
put(ex, "B15", "=6+SUM(B12:D12)+SUM(B13:D13)+B14", BLK, NUM, bold=True)
put(ex, "F15", "12.62 instead of a flat 13: the model does not count years a child would not have been enrolled anyway", NOTE, wrap=True)
put(ex, "A17", "THE SHARE, APPLIED: TODAY'S STUDENTS, THEN THE WHOLE FEEDER STREAM", SEC)
put(ex, "A18", "Estimate", BOLDW, fill=HDR); put(ex, "B18", "Leave share", BOLDW, fill=HDR); put(ex, "C18", "Of today's 128", BOLDW, fill=HDR); put(ex, "D18", "Per year, full feeder", BOLDW, fill=HDR); put(ex, "E18", "$/yr (feeder)", BOLDW, fill=HDR); put(ex, "F18", "Basis", BOLDW, fill=HDR)
put(ex, "A19", "Band low (25th pct)"); put(ex, "B19", 0.42975, BLUE, NUM); put(ex, "C19", "=ROUND(B19*$B$8,0)", BLK, NUM); put(ex, "D19", "=ROUND(B19*$B$9*$B$15,0)", BLK, NUM); put(ex, "E19", "=D19*(Assumptions!B6+500)", BLK, CUR); put(ex, "F19", "Middle half of the corrected posterior; the closure grid's low leg", NOTE, wrap=True)
put(ex, "A20", "Median"); put(ex, "B20", 0.50075, BLUE, NUM); put(ex, "C20", "=ROUND(B20*$B$8,0)", BLK, NUM); put(ex, "D20", "=ROUND(B20*$B$9*$B$15,0)", BLK, NUM); put(ex, "E20", "=D20*(Assumptions!B6+500)", BLK, CUR); put(ex, "F20", "The closure grid's central leg and the calculator default", NOTE, wrap=True)
put(ex, "A21", "Band high (75th pct)"); put(ex, "B21", 0.56725, BLUE, NUM); put(ex, "C21", "=ROUND(B21*$B$8,0)", BLK, NUM); put(ex, "D21", "=ROUND(B21*$B$9*$B$15,0)", BLK, NUM); put(ex, "E21", "=D21*(Assumptions!B6+500)", BLK, CUR); put(ex, "F21", "The closure grid's high leg", NOTE, wrap=True)
put(ex, "A22", "95th percentile bound"); put(ex, "B22", 0.65525, BLUE, NUM); put(ex, "C22", "=ROUND(B22*$B$8,0)", BLK, NUM); put(ex, "D22", "=ROUND(B22*$B$9*$B$15,0)", BLK, NUM); put(ex, "E22", "=D22*(Assumptions!B6+500)", BLK, CUR); put(ex, "F22", "Not priced in the grid; the honest upper bound", NOTE, wrap=True)
put(ex, "A24", "Leave shares are the posterior of a response-propensity model: across the sampled window the raw split is 62 leavers to 13 stayers; leaving families are assumed 3.3x to about 9x likelier to answer (log-normal prior centered on 3.5x and FLOORED at 3.3x, the lower of Pew's high-salience benchmarks, chosen for a small, emotionally charged respondent pool reached through a campaign-circulated form), and the shares above are the resulting quartiles. Class-year codes are not assumed accurate; the evidence window spans current enrollment plus the next three entering classes (62 leavers to 13 stayers of 75), so a miscoded child counts either way, and the correction still assumes the largest silent pool the coding allows, leaning the band low. Corroboration outside the survey: SAAR 2025-26 kindergarten of 12 against a ten-year average of 22; end-of-year 141 / 128 / 115 across 2023-24 to 2025-26. Losses build from six grade cohorts in year one to all thirteen by year eight (141 of 169 cohort-years across a 13-year window).", NOTE, wrap=True)
put(ex, "A27", "COST RESPONSE: WHAT STOPS BEING SPENT AS STUDENTS LEAVE", SEC)
put(ex, "A28", "Supplies and materials per departed student (scales with students; the growth model charges recruits the same figure)"); put(ex, "B28", "=Assumptions!B62", GRN, CUR)
put(ex, "A29", "Teacher savings are priced ONLY on the Closure_Model teachers-cut lever (the district's own 0 to 3 positions), never here, so staffing savings cannot be counted twice. Even with the supplies credit and the teacher lever at its friendliest, every priced closure scenario loses money.", NOTE, wrap=True)

gr = sheet("Growth_Model", [50, 14, 14, 14])
put(gr, "A1", "Grow the Kings: Nonresident Enrollment Model (HB 563 / KRS 157.350; legacy single-point scenario, published grid summarized at row 17)", TITLE)
put(gr, "A2", "SEEK funding follows nonresident students since July 2022; no agreement from the home district is required.", NOTE)
for col, yr in zip("BCD", ["FY2027", "FY2028", "FY2029"]):
    put(gr, f"{col}4", yr, BOLDW, fill=HDR)
put(gr, "A4", "", BOLDW, fill=HDR)
put(gr, "A5", "Transfer students (scenario)"); 
put(gr, "B5", "=Assumptions!B59", GRN, NUM); put(gr, "C5", "=Assumptions!B60", GRN, NUM); put(gr, "D5", "=Assumptions!B61", GRN, NUM)
put(gr, "A6", "Enrolled (capped at open seats)")
for col in "BCD":
    put(gr, f"{col}6", f"=MIN({col}5,Assumptions!$B$13)", BLK, NUM)
put(gr, "A7", "SEEK base per pupil")
put(gr, "B7", "=Assumptions!B6", GRN, CUR); put(gr, "C7", "=Assumptions!B7", GRN, CUR); put(gr, "D7", "=Assumptions!B8", GRN, CUR)
put(gr, "A8", "New SEEK revenue (base only)")
for col in "BCD":
    put(gr, f"{col}8", f"={col}6*{col}7", BLK, CUR)
put(gr, "A9", "Variable cost of added students")
for col in "BCD":
    put(gr, f"{col}9", f"={col}6*Assumptions!$B$62", BLK, CUR)
put(gr, "A10", "Added teacher (once threshold passed)")
for col in "BCD":
    put(gr, f"{col}10", f"=IF({col}6>Assumptions!$B$63,49150,0)", BLK, CUR)
put(gr, "A11", "Net new recurring revenue", bold=True)
for col in "BCD":
    cc = put(gr, f"{col}11", f"={col}8-{col}9-{col}10", BLK, CUR, bold=True); cc.border = TOPLINE
put(gr, "A12", "Cumulative")
put(gr, "B12", "=B11", BLK, CUR); put(gr, "C12", "=B12+C11", BLK, CUR); put(gr, "D12", "=C12+D11", BLK, CUR)
put(gr, "A14", "Upside excluded from this model: tuition, SEEK add-on weights, preschool and day-care expansion, and in-county boundary redistricting, which fills seats with students the district already serves.", NOTE, wrap=True)
put(gr, "A15", "Context: NMES enrolled 160 students as recently as 2019-20 (see School_Data) - the growth targets restore recent history, they do not exceed it.", NOTE, wrap=True)
put(gr, "A17", "THE PUBLISHED GROWTH GRID (build/growth_grid.py; the headline model behind the site calculator)", SEC)
put(gr, "A18", "19,683 scenarios: added students 10 to 90; classroom-indexed hiring beyond the 25 open seats at 1 per 18/21/24; teacher cost $41,718/$49,150/$56,583 (certified schedule entry-to-mid rows); support staff none, 1 per 75 at $28,500, or 1 per 50 at $37,000 (staff-per and staff-cost enumerated as independent levers); busing $0/$500/$1,000 per recruit; marginal cost $400/$700/$1,000; SEEK add-ons $0/$500/$1,000. Weighted median +$142,220 (the site default, 30 students at $4,736, lands at +$142,080, within $140 of it); middle half +$94,720 to +$183,354; floor +$4,131; ceiling +$387,804; zero negative scenarios.", NOTE, wrap=True)
put(gr, "A19", "The three-year table above is the legacy v3 transfer scenario, kept for continuity; its teacher line is priced at the certified schedule mid-row ($49,150), consistent with the published grid, rather than the $85,000 all-in comparison figure.", NOTE, wrap=True)

# ================= REDISTRICTING =================
rd = sheet("Redistricting", [56, 15, 58])
put(rd, "A1", "Fill the Kings' Seats: Boundary Rebalancing and Cross-County Scenario", TITLE)
put(rd, "A2", "A planning scenario, not a routing study. The July 2026 records response produced current bus routes but answered N/A for GIS files, routing-software reports, and ride-time analyses; the optimization inputs do not exist in public form.", NOTE, wrap=True)

put(rd, "A4", "CURRENT ELEMENTARY MAP (2024-25 counts as cited in report Sections 4 and 9)", SEC)
put(rd, "A5", "NMES enrollment"); put(rd, "B5", "=Assumptions!B11", GRN, NUM)
put(rd, "A6", "NMES rated capacity"); put(rd, "B6", "=Assumptions!B12", GRN, NUM)
put(rd, "A7", "NMES open seats"); put(rd, "B7", "=B6-B5", BLK, NUM)
put(rd, "A8", "Bourbon Central enrollment"); put(rd, "B8", 459, BLUE, NUM); put(rd, "C8", "Report Section 4; confirm against current-year infinite campus counts", NOTE)
put(rd, "A9", "Cane Ridge enrollment"); put(rd, "B9", 453, BLUE, NUM); put(rd, "C9", "Report Section 4", NOTE)

put(rd, "A11", "SCENARIO LEVERS (yellow = judgment calls the district's data should replace)", SEC)
put(rd, "A12", "Students rezoned to NMES from the eastern edges of the Paris-area zones", )
put(rd, "B12", 30, BLUE, NUM, fill=YEL); put(rd, "C12", "Chosen from families already living closer to NMES than to their assigned school; requires the geocoded counts", NOTE, wrap=True)
put(rd, "A13", "Cross-county transfers under HB 563 (KRS 157.350)")
put(rd, "B13", 16, BLUE, NUM, fill=YEL); put(rd, "C13", "SEEK funding follows each transfer; no home-district agreement required since July 2022", NOTE, wrap=True)
put(rd, "A14", "NMES enrollment after"); put(rd, "B14", "=B5+B12+B13", BLK, NUM)
put(rd, "A15", "Fill check vs capacity"); put(rd, "B15", "=B6-B14", BLK, NUM); put(rd, "C15", "Zero = exactly full", NOTE)
put(rd, "A16", "Bourbon Central after (half the rezone)"); put(rd, "B16", "=B8-B12/2", BLK, NUM)
put(rd, "A17", "Cane Ridge after (half the rezone)"); put(rd, "B17", "=B9-B12/2", BLK, NUM)

put(rd, "A19", "CLASSROOMS", SEC)
put(rd, "A20", "NMES homerooms in use"); put(rd, "B20", 6, BLUE, NUM); put(rd, "C20", "One per grade K-5; NCES lists 9.41 classroom-teacher FTE, which also counts Title I, special ed and part-time certified", NOTE)
put(rd, "A21", "Average class size today"); put(rd, "B21", "=B5/B20", BLK, '0.0')
put(rd, "A22", "Average class size at capacity"); put(rd, "B22", "=B14/B20", BLK, '0.0')
put(rd, "A23", "Statutory caps (KRS 157.360): 24 in K-3, 28 in grade 4, 29 in grades 5-6. At an even mix, 174 students exceed the single-section caps in most grades, so the lever below prices new NMES sections honestly (v3.8 correction; earlier versions omitted this charge).", NOTE, wrap=True)
put(rd, "A24", "New sections needed at NMES under the caps (0 if the rezone is drawn grade-by-grade; 2 at a fully even mix)"); put(rd, "B24", 1, BLUE, NUM, fill=YEL)

put(rd, "A25", "RECURRING DOLLARS", SEC)
put(rd, "A26", "New SEEK revenue from cross-county transfers (FY2027 base)"); put(rd, "B26", "=B13*Assumptions!B6", GRN, CUR)
put(rd, "A27", "Variable cost of all added students"); put(rd, "B27", "=(B12+B13)*Assumptions!B62", GRN, CUR)
put(rd, "A28", "Sections avoided or redeployed at receiving schools, low (count)"); put(rd, "B28", 1, BLUE, NUM, fill=YEL)
put(rd, "A29", "Sections avoided or redeployed at receiving schools, high (count)"); put(rd, "B29", 2, BLUE, NUM, fill=YEL)
put(rd, "A30", "Net recurring benefit, low (avoided and added sections both priced at the GF-borne $60K central; v3.8 subtracts the NMES additions)", bold=True); b30 = put(rd, "B30", "=B26-B27+(B28-B24)*Assumptions!B69", BLK, CUR, bold=True); b30.border = TOPLINE
put(rd, "A31", "Net recurring benefit, high (two avoided sections, same NMES debit)", bold=True); put(rd, "B31", "=B26-B27+(B29-B24)*Assumptions!B69", BLK, CUR, bold=True)
put(rd, "A32", "Corner cases for the record: two added and none avoided gives about -$64,000; none added and two avoided gives about +$176,000. The published range uses the levers above.", NOTE, wrap=True)

put(rd, "A33", "PER-STUDENT ARITHMETIC (the number the closure argument leans on)", SEC)
put(rd, "A34", "NMES site spending today"); put(rd, "B34", "=Assumptions!B14*Assumptions!B11", GRN, CUR)
put(rd, "A35", "Per student today"); put(rd, "B35", "=Assumptions!B14", GRN, CUR)
put(rd, "A36", "Per student at capacity (site cost plus variable cost, over 174)")
put(rd, "B36", "=(B34+(B12+B13)*Assumptions!B62)/B14", BLK, CUR)
put(rd, "A37", "Change"); put(rd, "B37", "=B36/B35-1", BLK, PCT)

put(rd, "A45", "FILLED-TO-CAPACITY PER-PUPIL COST VS THE RECEIVING SCHOOLS (KDE 2023-24 filings)", SEC)
put(rd, "A46", "Bourbon Central per pupil today"); put(rd, "B46", 18131, BLUE, CUR); put(rd, "C46", "District's own KDE school-level filing, 2023-24", NOTE)
put(rd, "A47", "Cane Ridge per pupil today"); put(rd, "B47", 18670, BLUE, CUR); put(rd, "C47", "Same filing; Cane Ridge runs 31 students over its rated capacity", NOTE)
put(rd, "A48", "NMES per pupil today"); put(rd, "B48", "=Assumptions!B14", GRN, CUR)
put(rd, "A49", "NMES filled to 174 (state-approved rating), no added teachers")
put(rd, "B49", "=(B34+(B6-B5)*Assumptions!B62)/B6", BLK, CUR)
put(rd, "C49", "Six existing homerooms hold 153 at the district's own Appendix B caps (24 K-3, 28 fourth, 29 fifth; 25 seats open today); filling to 174 needs one added section, priced in the row below", NOTE, wrap=True)
put(rd, "A50", "NMES filled to 174, conservative (two teachers added anyway)")
put(rd, "B50", "=(B34+2*Assumptions!B41+(B6-B5)*Assumptions!B62)/B6", BLK, CUR)
put(rd, "C50", "Adds two loaded certified positions even though class sizes do not require them", NOTE, wrap=True)
put(rd, "A51", "NMES filled to 154 (2026 draft plan rating), no added teachers")
put(rd, "B51", "=(B34+(154-B5)*Assumptions!B62)/154", BLK, CUR)
put(rd, "C51", "Same math at the draft plan's lower rating; class sizes average 17.1", NOTE, wrap=True)
put(rd, "A52", "NMES at 174 vs cheapest receiving school", bold=True)
b52 = put(rd, "B52", "=B49-MIN(B46,B47)", BLK, CUR, bold=True); b52.border = TOPLINE
put(rd, "C52", "Negative = NMES becomes the cheapest elementary school in the district", NOTE, wrap=True)
put(rd, "A53", "Filling the school reverses the cost argument: the premium exists because fixed costs sit on 128 students; spread over 174 the same school costs less per child than either receiving school, and the transfer students free the exact seats Cane Ridge is over capacity by.", NOTE, wrap=True)

put(rd, "A55", "SENSITIVITY: IS THE CAPACITY RESULT STRUCTURAL, OR AN ARTIFACT OF ASSUMPTIONS?", SEC)
put(rd, "A56", "The only lever that matters is the marginal cost of each added student. Break-even values below are the marginal cost at which NMES at capacity stops being cheaper. Compare them to reality: added supplies and materials run a few hundred dollars; even hiring staff in proportion to students costs roughly half the average, because the average includes the principal, building, and staff already paid for.", NOTE, wrap=True)
put(rd, "A57", "Break-even marginal cost, NMES at 174 vs Bourbon Central")
put(rd, "B57", "=(B46*B6-B34)/(B6-B5)", BLK, CUR)
put(rd, "C57", "NMES stays cheaper than Bourbon Central for any marginal cost below this", NOTE, wrap=True)
put(rd, "A58", "Break-even marginal cost, NMES at 174 vs Cane Ridge")
put(rd, "B58", "=(B47*B6-B34)/(B6-B5)", BLK, CUR)
put(rd, "A59", "Break-even marginal cost, NMES at 154 vs Bourbon Central")
put(rd, "B59", "=(B46*154-B34)/(154-B5)", BLK, CUR)
put(rd, "A60", "Break-even marginal cost, NMES at 154 vs Cane Ridge")
put(rd, "B60", "=(B47*154-B34)/(154-B5)", BLK, CUR)
put(rd, "A61", "Reference: NMES's ENTIRE state and local spending per pupil (fixed costs included)")
put(rd, "B61", "=Assumptions!B15", GRN, CUR)
put(rd, "C61", "Even if every added student cost this full amount, which is impossible since it includes the fixed costs already paid, NMES at 174 still comes in under both receiving schools", NOTE, wrap=True)
put(rd, "A63", "PER-PUPIL GRID ACROSS MARGINAL-COST ASSUMPTIONS", SEC)
put(rd, "A64", "Marginal cost per added student"); put(rd, "B64", "At 174", BOLDW, fill=HDR); put(rd, "C64", "At 154", BOLDW, fill=HDR)
for i, mc in enumerate([400, 2500, 5000, 10000, 14173]):
    rr = 65 + i
    put(rd, f"A{rr}", mc, BLUE, CUR)
    put(rd, f"B{rr}", f"=(B$34+(B$6-B$5)*A{rr})/B$6", BLK, CUR)
    put(rd, f"C{rr}", f"=(B$34+(154-B$5)*A{rr})/154", BLK, CUR)
put(rd, "A70", "Receiving schools today: Bourbon Central $18,131, Cane Ridge $18,670. At the state-approved 174 rating, NMES undercuts both across the entire grid, including the impossible full-average case. At the draft plan's 154 rating it holds for any marginal cost under about $12,140, roughly three times the class-cap marginal this workbook prices (about $3,670). Verdict: structural at 174, robust at 154.", NOTE, wrap=True)
put(rd, "A71", "Current enrollment is 128 (NCES CCD, 2024-25); filling to 174 adds 46 students, the figure used throughout this workbook and the grid above.", NOTE, wrap=True)

put(rd, "A73", "FIVE YEARS OF THE SAME TEST (KDE school-level filings, 2019-20 to 2023-24; backs Figure 6)", SEC)
put(rd, "A74", "Year", BOLDW, fill=HDR); put(rd, "B74", "NMES members", BOLDW, fill=HDR); put(rd, "C74", "NMES $/pupil", BOLDW, fill=HDR)
put(rd, "D74", "BCES $/pupil", BOLDW, fill=HDR); put(rd, "E74", "CRES $/pupil", BOLDW, fill=HDR); put(rd, "F74", "NMES at 174", BOLDW, fill=HDR); put(rd, "G74", "NMES at 154", BOLDW, fill=HDR)
hist5 = [("2019-20", 166, 12903, 12159, 12168), ("2020-21", 160, 15406, 14011, 14621),
         ("2021-22", 146, 19080, 15619, 16137), ("2022-23", 144, 19003, 17410, 17403),
         ("2023-24", 128, 19348, 18131, 18670)]
for i, (yy, nn, npp, bpp, cpp) in enumerate(hist5):
    rr = 75 + i
    put(rd, f"A{rr}", yy); put(rd, f"B{rr}", nn, BLUE, NUM); put(rd, f"C{rr}", npp, BLUE, CUR)
    put(rd, f"D{rr}", bpp, BLUE, CUR); put(rd, f"E{rr}", cpp, BLUE, CUR)
    put(rd, f"F{rr}", f"=(C{rr}*B{rr}+(174-B{rr})*Assumptions!B62)/174", BLK, CUR)
    if nn < 154:
        put(rd, f"G{rr}", f"=(C{rr}*B{rr}+(154-B{rr})*Assumptions!B62)/154", BLK, CUR)
    else:
        put(rd, f"G{rr}", "n/a: enrolled above 154", NOTE)
put(rd, "A81", "Honest reading (corrected 7/26): at 174 the counterfactual is within 1 to 3 percent of the cheapest school in 2019-22, a tie, and decisively below both schools in 2022-23 and 2023-24. At the draft plan's 154 rating it is cheapest on the 2023-24 filing; in 2019-20 and 2020-21 the school enrolled above 154, the capacity the draft now assigns it. NMES's total site spending grew about 16 percent over the five years, against 35 to 37 percent at Bourbon Central and 46 to 47 percent at Cane Ridge, on the state's per-student filings times each school's reported enrollment; the bands carry the two base-year counts the record offers, the district's 2021 facility plan enrollment and the federal fall 2019 count. Decomposed, NMES's per-student rise is roughly one third spending and two thirds divisor (total up about 16 percent over membership down about 23 percent); at Bourbon Central the same decomposition is mostly spending. That is the finding: at this school the divisor did most of the work.", NOTE, wrap=True)

put(rd, "A83", "SYMMETRY CHECK: THE SENDERS' PER-PUPIL AFTER THE MOVE (2023-24 basis)", SEC)
put(rd, "A84", "Per-pupil rises at the senders for the same denominator reason it falls at NMES; a pure shuffle leaves total district spending nearly unchanged in either direction. The comparison below is the honest post-move one.", NOTE, wrap=True)
put(rd, "A85", "Bourbon Central after sending 15 rezoned students")
put(rd, "B85", "=(8902321-15*Assumptions!B62)/(491-15)", BLK, CUR)
put(rd, "C85", "2023-24 filing: $8,902,321 across 491 members; loses only variable cost per departing student", NOTE, wrap=True)
put(rd, "A86", "Cane Ridge after sending 15 rezoned students")
put(rd, "B86", "=(8606870-15*Assumptions!B62)/(461-15)", BLK, CUR)
put(rd, "A87", "Stress case: all 46 drawn in-county, 23 from each sender")
put(rd, "B87", "=(8902321-23*Assumptions!B62)/(491-23)", BLK, CUR)
put(rd, "C87", "Bourbon Central; Cane Ridge in the next row", NOTE)
put(rd, "A88", "Stress case, Cane Ridge")
put(rd, "B88", "=(8606870-23*Assumptions!B62)/(461-23)", BLK, CUR)
put(rd, "A91", "VARIABLE-COST VALIDATION, BOTH DIRECTIONS (added 7/26)", SEC)
put(rd, "A92", "Class-cap staffing check (KRS 157.360: 24 in K-3, 28 in grade 4, 29 in grade 5). At an even grade mix, 174 students need about 11 sections against 9 existing rooms, and 154 need about 10; the supplies-only counterfactual is therefore the BEST case, and the staffed cases below are the honest base cases. The July 2026 records response provided an elementary capacity graphic but no grade-by-grade files (building maps withheld for security); a mix weighted to grades 4-5 needs fewer sections.", NOTE, wrap=True)
put(rd, "A93", "NMES at 174 with two added sections (base case)")
put(rd, "B93", "=(B34+2*Assumptions!B41+(B6-B5)*Assumptions!B62)/B6", BLK, CUR)
put(rd, "C93", "Same as the conservative row above; cheapest of the three on the 2023-24 filing", NOTE, wrap=True)
put(rd, "A94", "NMES at 154 with one added section (base case)")
put(rd, "B94", "=(B34+Assumptions!B41+(154-B5)*Assumptions!B62)/154", BLK, CUR)
put(rd, "A95", "Long-run bound at 174: every added student at the full $14,173 S/L average")
put(rd, "B95", "=(B34+(B6-B5)*Assumptions!B15)/B6", BLK, CUR)
put(rd, "C95", "Still under both receiving schools on the 2023-24 filing. At 154 the same bound gives $18,474 and fails against Bourbon Central, consistent with the break-even rows above", NOTE, wrap=True)
put(rd, "A96", "Senders after 15 out each WITH one consolidated section (favorable to senders)")
put(rd, "B96", "=(8902321-15*Assumptions!B62-Assumptions!B41)/(491-15)", BLK, CUR)
put(rd, "C96", "Bourbon Central; Cane Ridge next row. Consolidation lowers sender per-pupil AND is the source of the district-level saving; do not count it twice", NOTE, wrap=True)
put(rd, "A97", "Cane Ridge, same treatment")
put(rd, "B97", "=(8606870-15*Assumptions!B62-Assumptions!B41)/(461-15)", BLK, CUR)
put(rd, "A98", "The same discipline applies to the closure direction: 128 arriving students trigger the same class caps at the receiving schools, adding sections in several grades, which is exactly why the closure's net effect runs minus $847,825 to minus $11,030 with a weighted median of minus $428,627, and not the school's $2.5M gross cost.", NOTE, wrap=True)

put(rd, "A100", "FAIR TEST: EVERY SCHOOL FILLED TO ITS RATED CAPACITY, SEVEN CAPACITY SETS (backs Figure 6)", SEC)
put(rd, "A101", "Step costs: $400 per student added or removed, plus or minus $85,000 per section vs today's staffing (even K-5 mix under KRS 157.360). Section deltas below are precomputed from that rule; capacities from the named documents, all archived in build/.", NOTE, wrap=True)
put(rd, "A102", "Scenario (NMES/BCES/CRES caps)", BOLDW, fill=HDR); put(rd, "B102", "NMES", BOLDW, fill=HDR); put(rd, "C102", "BCES", BOLDW, fill=HDR); put(rd, "D102", "CRES", BOLDW, fill=HDR); put(rd, "E102", "Cheapest", BOLDW, fill=HDR)
fair = [
 ("Actual today (2023-24 filing)", 19348, 18131, 18670, "BCES"),
 ("2013 plan, KBE approved (198/564/500)", 13937, 16137, 17245, "NMES"),
 ("2017 plan, KBE approved (152/611/550)", 16915, 15483, 16023, "BCES"),
 ("2021 plan, KBE approved, in force (174/521/422)", 15316, 17273, 19553, "NMES"),
 ("Peak enrollment 2005-2025 (224/620/495)", 12366, 15264, 17415, "NMES"),
 ("2026 KFICS architect slides (154/499/397)", 16701, 17847, 20759, "NMES"),
 ("2026 draft DFP table, unapproved (154/640/547)", 16701, 14800, 16108, "BCES"),
]
for i, (lbl, a, b2, c2, wname) in enumerate(fair):
    rr = 103 + i
    put(rd, f"A{rr}", lbl); put(rd, f"B{rr}", a, BLUE, CUR); put(rd, f"C{rr}", b2, BLUE, CUR); put(rd, f"D{rr}", c2, BLUE, CUR); put(rd, f"E{rr}", wname, NOTE)
put(rd, "A110", "Same-building ratings across four consecutive plans: NMES 198-152-174-154, Bourbon Central 564-611-521-640, Cane Ridge 500-550-422-547. Swings up to 128 seats with no major construction after 2009. The 2017 plan (recovered from the Internet Archive; KBE minutes June 7, 2017 corroborate) lists NMES at 154 enrolled against 152 capacity: OVER capacity.", NOTE, wrap=True)
put(rd, "A111", "VALIDATION AGAINST ACTUALS: withdrawn in v3.9. The two-school cost slope used here depended on memberships (491 and 461, matching the state 2024-25 SAAR end-of-year membership file (the draft plan labels the column 2023-24; the label is corrected in v5.0), Facility_Plans rows 30-31) that contradict the fall counts in B8 and B9 above, and its sign flips from plus $9,848 to minus $941 to minus $22,564 across the three plausible membership pairs. The honest bound is the break-even marginal cost already computed live in rows 57 to 60.", NOTE, wrap=True)
put(rd, "A114", "Extreme bound: NMES's own year-over-year cost change 2022-23 to 2023-24 ($259,888 lower with 16 fewer students, about $16,243 per student, which folds in deliberate staffing cuts) gives $18,527 at 174: under Cane Ridge, marginally over Bourbon Central. Even the most hostile actuals-derived number does not restore the cost case.", NOTE, wrap=True)

put(rd, "A116", "THE RECRUITMENT POOL: WHERE FILL-THE-SEATS STUDENTS CAN COME FROM (measured; year noted per row; sources archived under build/)", SEC)
put(rd, "A117", "Registered homeschoolers, Bourbon Co district (2022-23; district letter-of-intent counts via the Washington Post records project)"); put(rd, "B117", 236, BLUE, NUM)
put(rd, "A118", "Registered homeschoolers, Paris Independent (2022-23)"); put(rd, "B118", 23, BLUE, NUM)
put(rd, "A119", "County total registered homeschoolers (a floor; KY letter compliance is incomplete)"); put(rd, "B119", "=B117+B118", BLK, NUM, bold=True)
put(rd, "A120", "County school-age (5-17) residents in private school OR homeschool (ACS 2019-23; the Census asks these as one combined answer; margin of error +/-299)"); put(rd, "B120", 1135, BLUE, NUM)
put(rd, "A121", "Same figure, ACS 2014-18 window: the pool nearly tripled in five years"); put(rd, "B121", 422, BLUE, NUM)
put(rd, "A122", "Of the current pool, seats at St. Mary (Paris, PK-5; federal Private School Survey 2023-24). Bourbon Christian Academy (Millersburg, K-12) is absent from that voluntary survey, so its students are uncounted and the pool stays a floor"); put(rd, "B122", 96, BLUE, NUM)
put(rd, "A123", "Bourbon Co residents enrolled in OUT-OF-COUNTY public districts (KDE Non-Resident Student report, SY2024-25)"); put(rd, "B123", 76, BLUE, NUM)
put(rd, "A124", "  of which Cloverport Independent, host of the statewide virtual academy: a 150-mile 'commute' that is really virtual enrollment"); put(rd, "B124", 10, BLUE, NUM)
put(rd, "A125", "Out-of-county students already enrolled in Bourbon Co schools (same KDE report)"); put(rd, "B125", 131, BLUE, NUM)
put(rd, "A126", "  of which from Fayette County"); put(rd, "B126", 54, BLUE, NUM)
put(rd, "A127", "Net nonresident import, Bourbon Co district (436 enrolled in vs 247 residents out, Paris flows included; kept for context: the published pool counts the 247 exports as recoverable rather than netting them against imports)"); put(rd, "B127", 189, BLUE, NUM)
put(rd, "A129", "LEVER: returning homeschool or private-school students recruited (site planner slider; default zero so the published fill-planner numbers are unchanged)"); put(rd, "B129", 0, BLUE, NUM, fill=YEL)
put(rd, "A130", "New SEEK per returning student (FY2027 base; same cell the leaver cost uses, so the symmetry is honest)"); put(rd, "B130", "=Assumptions!B6", GRN, CUR)
put(rd, "A131", "Net per returning student after variable cost"); put(rd, "B131", "=Assumptions!B6-Assumptions!B62", BLK, CUR)
put(rd, "A132", "Recurring SEEK revenue if the lever is used"); put(rd, "B132", "=B129*B131", BLK, CUR)
put(rd, "A133", "Recurring SEEK revenue if all 46 open seats fill from this pool alone"); put(rd, "B133", "=46*Assumptions!B6", BLK, CUR, bold=True)
put(rd, "A134", "Capture rate that requires: share of registered homeschoolers / share of the full ACS pool"); put(rd, "B134", "=46/B119", BLK, PCT); put(rd, "C134", "=46/B120", BLK, PCT)
put(rd, "A136", "Residents enrolled in another district (KDE Non-Resident report SY2024-25: 171 at Paris Independent + 76 out of county)"); put(rd, "B136", 247, BLUE, NUM)
put(rd, "A137", "Documented Bourbon-specific floor: homeschool filings plus residents enrolled elsewhere"); put(rd, "B137", "=B117+B136", BLK, NUM, bold=True)
put(rd, "A138", "Published planning pool: Bourbon County Schools kids in homeschool, private school, or another district"); put(rd, "B138", "450 to 550", BLUE)
put(rd, "A139", "Pool priced at the full SEEK base, revenue not collected today (450 and 550 x FY2027 base; the published $2.1 to $2.5 million)"); put(rd, "B139", "=450*Assumptions!B6", BLK, CUR, bold=True); put(rd, "C139", "=550*Assumptions!B6", BLK, CUR, bold=True)
put(rd, "A140", "Same pool net of supplies, the basis the plan calculator credits per recovered student ($1.9 to $2.3 million)"); put(rd, "B140", "=450*(Assumptions!B6-Assumptions!B62)", BLK, CUR); put(rd, "C140", "=550*(Assumptions!B6-Assumptions!B62)", BLK, CUR)
put(rd, "A135", "Total added students from all sources (rezone + transfers + returns) is capped at 46 by the rated 174; the site planner enforces the cap. Rezoned students bring no new SEEK (already enrolled in-district); transfers and returns each bring the full base. Sources: KDE Non-Resident Student report SY24-25 and the Washington Post home-school district file, both archived under build/; ACS table B14003 (Bourbon County); NCES Private School Universe Survey 2023-24.", NOTE, wrap=True)

put(rd, "A89", "NMES at 174 stays the cheapest of the three after the move, by a wider margin than before it (about $4,400 vs $3,800). The district-level cash case is booked separately and conservatively above as the net recurring benefit: consolidated sections at the senders, HB 563 SEEK revenue from out-of-county transfers, and capacity relief at Cane Ridge, which runs 31 students over its rating. The same symmetry runs the other way: closing NMES would lower the receiving schools' per-pupil optics while saving almost nothing in total, which is the closure case's weakness in one sentence.", NOTE, wrap=True)
put(rd, "A39", "ASSUMPTIONS THE DISTRICT'S DATA SHOULD REPLACE", SEC)
put(rd, "A40", "Rezoned students are drawn only from homes closer to NMES than to their assigned school, so bus routes shorten or hold even; the district's routing data would settle it.", NOTE, wrap=True)
put(rd, "A41", "Receiving-school relief is booked only as one to two avoided or redeployed sections; the July 2026 records response provided a capacity graphic; grade-by-grade room files were not shared (building maps withheld for security), so relief stays conservatively booked.", NOTE, wrap=True)
put(rd, "A42", "SEEK for rezoned in-county students is unchanged (same district); only cross-county transfers add revenue.", NOTE, wrap=True)
put(rd, "A43", "Per-student figures pair the latest published spending year (2023-24) with the current enrollment count; refresh when the 2024-25 spending data posts.", NOTE, wrap=True)

# ================= TRANSPORT_GEO =================
tg = sheet("Transport_Geo", [58, 15, 58])
put(tg, "A1", "Transportation and Geography: Density, Route Miles, and What Closure Adds", TITLE)
put(tg, "A2", "Zone geometry is official: NCES School Attendance Boundary Survey, 2015-16 collection (build/sabs_zones.json in the repository). Cost inputs remain labeled estimates; the district's annual T-1 transportation report and geocoded counts replace them.", NOTE, wrap=True)

put(tg, "A4", "GEOGRAPHY AND STUDENT DENSITY", SEC)
put(tg, "A5", "Bourbon County land area (square miles)"); put(tg, "B5", 290, BLUE, NUM); put(tg, "C5", "U.S. Census: 289.7 land square miles", NOTE)
put(tg, "A6", "Paris city population, 2020"); put(tg, "B6", 10171, BLUE, NUM); put(tg, "C6", "2020 Census", NOTE)
put(tg, "A7", "Millersburg population, 2020"); put(tg, "B7", 747, BLUE, NUM); put(tg, "C7", "2020 Census", NOTE)
put(tg, "A8", "North Middletown population, 2020"); put(tg, "B8", "=Demographics!B29", GRN, NUM); put(tg, "C8", "Demographics tab", NOTE)
put(tg, "A9", "NMES zone share of county area"); put(tg, "B9", 0.38, BLUE, PCT); put(tg, "C9", "Official: NCES SABS 2015-16, NMES zone 110.3 sq mi of the 289.1 sq mi zone total", NOTE)
put(tg, "A10", "NMES zone area (sq mi)"); put(tg, "B10", "=B5*B9", BLK, '0')
put(tg, "A11", "Paris-area zones (sq mi)"); put(tg, "B11", "=B5-B10", BLK, '0')
put(tg, "A12", "NMES elementary students"); put(tg, "B12", "=Assumptions!B11", GRN, NUM)
put(tg, "A13", "Paris-area elementary students"); put(tg, "B13", "=Redistricting!B8+Redistricting!B9", GRN, NUM)
put(tg, "A14", "Students per square mile, NMES zone"); put(tg, "B14", "=B12/B10", BLK, '0.0')
put(tg, "A15", "Students per square mile, Paris-area zones"); put(tg, "B15", "=B13/B11", BLK, '0.0')
put(tg, "A16", "Students per square mile, district elementary overall"); put(tg, "B16", "=(B12+B13)/B5", BLK, '0.0')
put(tg, "A17", "State law (KRS 157.370) funds transportation on transported pupils per square mile: low density earns a higher per-pupil allotment because it costs more to serve. Funding history: below the formula for two decades, restored to 90 then 100 percent (on lagged FY2023 costs) in the 2024-2026 budget, then frozen again below formula in the 2026-2028 budget.", NOTE, wrap=True)
put(tg, "A18", "STATE REVENUE EFFECT: with the appropriation frozen at flat dollars computed on lagged costs, the marginal state reimbursement on NEW busing miles is zero, so closure's added routes are district money. Rebalancing changes no transported-pupil count, so the add-on is unchanged; and the district is not required to transport nonresident transfer students at all (board policy decides).", NOTE, wrap=True)

put(tg, "A19", "WHAT CLOSURE ADDS: ROUTE-MILE ARITHMETIC (yellow = replace with district T-1 data)", SEC)
put(tg, "A20", "Share of NMES students riding the bus"); put(tg, "B20", 0.85, BLUE, PCT, fill=YEL)
put(tg, "A21", "Riders"); put(tg, "B21", "=ROUND(B20*B12,0)", BLK, NUM)
put(tg, "A22", "Rural routes serving NMES today"); put(tg, "B22", 3, BLUE, NUM, fill=YEL)
put(tg, "A23", "Added distance to the Paris schools, one way (miles)"); put(tg, "B23", 10, BLUE, NUM); put(tg, "C23", "US 460, North Middletown to Paris", NOTE)
put(tg, "A24", "Added bus-miles per route per day (out and back, AM and PM)"); put(tg, "B24", "=B23*4", BLK, NUM)
put(tg, "A25", "School days per year"); put(tg, "B25", 170, BLUE, NUM, fill=YEL)
put(tg, "A26", "Added bus-miles per year"); put(tg, "B26", "=B22*B24*B25", BLK, NUM)
put(tg, "A27", "Marginal cost per bus-mile, low"); put(tg, "B27", 2.50, BLUE, '0.00', fill=YEL); put(tg, "C27", "Fuel, maintenance, driver time; replace with district cost data", NOTE)
put(tg, "A28", "Marginal cost per bus-mile, high"); put(tg, "B28", 4.50, BLUE, '0.00', fill=YEL)
put(tg, "A29", "Added cost, mileage basis, low"); put(tg, "B29", "=B26*B27", BLK, CUR)
put(tg, "A30", "Added cost, mileage basis, high"); put(tg, "B30", "=B26*B28", BLK, CUR)
put(tg, "A31", "Additional buses if route tiers break (high case)"); put(tg, "B31", 1, BLUE, NUM, fill=YEL)
put(tg, "A32", "All-in cost per additional bus-year"); put(tg, "B32", 55000, BLUE, CUR, fill=YEL)
put(tg, "A33", "Bottom-up added busing cost, low", bold=True); c33 = put(tg, "B33", "=B29", BLK, CUR, bold=True); c33.border = TOPLINE
put(tg, "A34", "Bottom-up added busing cost, high", bold=True); put(tg, "B34", "=B30+B31*B32", BLK, CUR, bold=True)
put(tg, "A35", "Report's planning range (Closure_Model offset basis)"); put(tg, "B35", "Between $75,000 and $200,000", NOTE)
put(tg, "A36", "The bottom-up estimate lands inside the planning range. The $137,500 midpoint survives only in the legacy single-point rows; the published grid's central busing figure is the bottom-up $63,000 (Closure_Model row 42). Note what closure does not remove: every square mile of the eastern county stays in the coverage area, with longer rides on it, roughly 15 to 20 added minutes each way on US 460.", NOTE, wrap=True)

put(tg, "A38", "WHAT REBALANCING CHANGES (Redistricting tab scenario)", SEC)
put(tg, "A39", "Students rezoned to NMES"); put(tg, "B39", "=Redistricting!B12", GRN, NUM)
put(tg, "A40", "Stem miles saved per affected route, one way"); put(tg, "B40", 3, BLUE, NUM, fill=YEL); put(tg, "C40", "Rezoned families live closer to NMES than to their assigned school", NOTE)
put(tg, "A41", "Affected routes"); put(tg, "B41", 2, BLUE, NUM, fill=YEL)
put(tg, "A42", "Bus-miles saved per year"); put(tg, "B42", "=B41*B40*4*B25", BLK, NUM)
put(tg, "A43", "Transport saving, low"); put(tg, "B43", "=B42*B27", BLK, CUR)
put(tg, "A44", "Transport saving, high"); put(tg, "B44", "=B42*B28", BLK, CUR)
put(tg, "A45", "Rebalancing is transport-neutral to modestly positive, the opposite sign of closure.", NOTE)

put(tg, "A47", "DISTRICT-WIDE CONTEXT", SEC)
put(tg, "A48", "Transportation expense, FY2025"); put(tg, "B48", "=Assumptions!B42", GRN, CUR)
put(tg, "A49", "Average Daily Attendance, FY2025"); put(tg, "B49", 2242.5, BLUE, '0.0'); put(tg, "C49", "FY2025 audit", NOTE)
put(tg, "A50", "Transportation cost per student in attendance"); put(tg, "B50", "=B48/B49", BLK, CUR)
put(tg, "A51", "Optimization potential at 5 to 10 percent (Alternatives menu)"); put(tg, "B51", "=Assumptions!B42*Assumptions!B43", GRN, CUR)
put(tg, "A52", ""); put(tg, "B52", "=Assumptions!B42*Assumptions!B44", GRN, CUR)

put(tg, "A54", "ACTUAL DISTANCES, COMPUTED FROM THE OFFICIAL ZONE GEOMETRY (build/zone_distances.py)", SEC)
put(tg, "A55", "Straight-line, NMES to the Paris schools (miles)"); put(tg, "B55", 8.871, BLUE, '0.0'); put(tg, "C55", "Great-circle between the school points; unrounded so derived cells match build/zone_distances.json", NOTE)
put(tg, "A56", "Measured road distance, US 460 (miles)"); put(tg, "B56", 10.0, BLUE, '0.0'); put(tg, "C56", "The one pair that can be measured exactly", NOTE)
put(tg, "A57", "Implied road factor on the measured pair"); put(tg, "B57", "=B56/B55", BLK, '0.00')
put(tg, "A58", "Road factor applied elsewhere (the measured US 460 pair implies 1.13)"); put(tg, "B58", 1.2, BLUE, '0.00', fill=YEL)
put(tg, "A59", "Area-average added distance under closure, one way, straight-line"); put(tg, "B59", 3.271, BLUE, '0.0'); put(tg, "C59", "Grid average over the official NMES zone (812 sample points); unrounded input, displays to one decimal", NOTE)
put(tg, "A60", "Area-average added distance, road"); put(tg, "B60", "=B59*B58", BLK, '0.0')
put(tg, "A61", "Farthest corner of the zone, straight-line to Paris / to NMES"); put(tg, "B61", 15.1, BLUE, '0.0'); put(tg, "C61", "8.3 miles to NMES; near the Nicholas County line", NOTE)
put(tg, "A62", "Farthest corner by road: to Paris / to NMES (miles)"); put(tg, "B62", "=ROUND(B61*B58,0)", BLK, '0'); put(tg, "C62", "About 18 versus 10; the district's routing data would give exact times", NOTE)
put(tg, "A63", "Share of the zone's area closer to NMES than to Paris"); put(tg, "B63", 0.78, BLUE, PCT); put(tg, "C63", "Computed on the official boundary; the map's whole point in one number", NOTE)

# ================= ALTERNATIVES =================
al = sheet("Alternatives", [46, 14, 14, 52, 20, 42])
put(al, "A1", "Revenue and Cost Alternatives (no school closed)", TITLE)
put(al, "A2", "Type separates new recurring revenue from recurring cost reductions; confidence names what would move each line from estimate to plan.", NOTE)
hdrs = ["Measure", "Low ($/yr)", "High ($/yr)", "Basis", "Type", "Confidence / what firms it up"]
for i, h in enumerate(hdrs):
    put(al, f"{get_column_letter(i+1)}3", h, BOLDW, fill=HDR)
alts = [
 ("KRS 160.470 four percent revenue mechanics (context; the published lead lever is the 2018 restore, Tax_History rows 70-91)", "=Tax_History!B50", 375000, "Low = year-one 4% revenue growth on the General Fund levied base; high allows base growth. A revenue cap, not a rate move.", GRN, BLUE,
  "New revenue", "Mechanics only; the published revenue ask anchors on restoring the 2018 rate"),
 ("Improve delinquent-tax recovery (partial)", 60000, 120000, "25-50% of FY2025 delinquency of $239,126 (2.4% of certified yield)", BLUE, BLUE,
  "New revenue", "Medium; needs an aging and collection analysis"),
 ("Attendance recovery (+1-2% ADA)", 100000, 200000, "Approx. SEEK value per 1% of ADA", BLUE, BLUE,
  "New revenue", "Medium; needs an attendance improvement plan"),
 ("Attrition-based staffing alignment", "=Assumptions!B48*Assumptions!B41", 425000, "Low = positions x loaded cost", GRN, BLUE,
  "Cost reduction", "Medium; needs a position-level staffing plan"),
 ("Administrative restraint", "=Assumptions!B47*(Assumptions!B46-Assumptions!B45)", 450000, "Low = rollback share of 2-yr district-admin growth", GRN, BLUE,
  "Cost reduction", "Medium; needs position- and vendor-level detail"),
 ("Transportation optimization", "=Assumptions!B42*Assumptions!B43", "=Assumptions!B42*Assumptions!B44", "5-10% of FY2025 transport expense", GRN, GRN,
  "Cost reduction", "Medium; needs a local route model (T-1 data requested)"),
 ("Energy performance contracting", 50000, 150000, "10-25% of utilities; authorized by 702 KAR 4:160", BLUE, BLUE,
  "Cost reduction", "Medium; contracts are structured to self-fund"),
 ("Fill NMES to capacity (rebalance + transfers, net)", "=Redistricting!B30", "=Redistricting!B31", "Boundary rebalancing and cross-county scenario, Redistricting tab", GRN, GRN,
  "New revenue, net of costs", "High; board boundary authority, math on Redistricting tab"),
 ("District-wide recruitment beyond NMES's 46 seats (homeschool, private-school, nonresident incentives; v3.8)", "=25*(Assumptions!B6-Assumptions!B62)", "=50*(Assumptions!B6-Assumptions!B62)", "25-50 additional students at $4,236 net; pool measured on Redistricting rows 116-135 (236 in the district's own homeschool files, 247 residents enrolled in other districts, 450-550 in all with private school); 62 open seats exist at Bourbon Central's approved rating", GRN, GRN,
  "New revenue", "Medium; needs an enrollment marketing plan and incentive design"),
]
r = 4
for label, lo, hi, basis, flo, fhi, typ, conf in alts:
    put(al, f"A{r}", label)
    put(al, f"B{r}", lo, flo, CUR)
    put(al, f"C{r}", hi, fhi, CUR)
    put(al, f"D{r}", basis, NOTE)
    put(al, f"E{r}", typ, NOTE)
    put(al, f"F{r}", conf, NOTE)
    r += 1
put(al, f"A{r}", "Total identified (ranges overlap; not additive to the penny)", bold=True)
b = put(al, f"B{r}", f"=SUM(B4:B{r-1})", BLK, CUR, bold=True); b.border = TOPLINE
cc = put(al, f"C{r}", f"=SUM(C4:C{r-1})", BLK, CUR, bold=True); cc.border = TOPLINE
tot = r
put(al, f"A{tot+2}", "COMPARISON", SEC)
put(al, f"A{tot+3}", "Package midpoint (raw sum of ranges)"); put(al, f"B{tot+3}", f"=(B{tot}+C{tot})/2", BLK, CUR)
put(al, f"A{tot+4}", "Published band, low (raw row sums, no haircut)"); put(al, f"B{tot+4}", f"=B{tot}", BLK, CUR)
put(al, f"A{tot+5}", "Published band, high (raw row sums, no haircut; the pre-v4.4 conservative $1.7M high is retired)"); put(al, f"B{tot+5}", f"=C{tot}", BLK, CUR)
put(al, f"A{tot+6}", "Band midpoint (used in Runway sheet)"); put(al, f"B{tot+6}", f"=(B{tot+4}+B{tot+5})/2", BLK, CUR)
put(al, f"A{tot+7}", "Average annual GF drawdown (FY2024-25)"); put(al, f"B{tot+7}", "=GF_Summary!D16", GRN, CUR)
put(al, f"A{tot+8}", "Closure net saving (LEGACY single-point base case, superseded by the grid median of -$428,627 at Closure_Model row 50)"); put(al, f"B{tot+8}", "=Closure_Model!B20", GRN, CUR)
put(al, f"A{tot+10}", "Reading: the raw-row band is $1.39M to $2.34M with no haircut (v4.4 review). The published headline is now the 2018 restore plus the counted-once cost package, $2.5M to $3.0M a year (transformative check below). Ranges overlap and are not additive to the penny, and each line carries its own confidence rating in column F. Medicaid and reimbursement recovery were removed from the menu in v4.2 review; shared services with Paris Independent was removed in v4.4 review. Coverage is reported against both yardsticks: the $2.65M structural gap before transfers and the roughly $1.15M net drawdown after transfers (Closure_Model row 21 carries both for closure).", NOTE, wrap=True)

put(al, f"A{tot+12}", "THE GROWTH PATH: THE SAME MENU AS A DISTRICT-WIDE RECOVERY PLAN (v3.8; backs the site card and Section 9)", SEC)
put(al, f"A{tot+13}", "Move 1: inspect fixed costs (every non-teaching position via attrition, administrative restructuring, transport, energy)")
put(al, f"B{tot+13}", "=SUM(B7:B10)", BLK, CUR, bold=True); put(al, f"C{tot+13}", "=SUM(C7:C10)", BLK, CUR, bold=True)
put(al, f"A{tot+14}", "Move 2: grow enrollment instead of shrinking it (attendance recovery, fill NMES, district-wide recruitment)")
put(al, f"B{tot+14}", "=B6+B11+B12", BLK, CUR, bold=True); put(al, f"C{tot+14}", "=C6+C11+C12", BLK, CUR, bold=True)
put(al, f"A{tot+15}", "Move 3: the honest revenue conversation, year one (this cell keeps the 4 percent + delinquency mechanics for continuity; the published lead lever is the 2018 restore, about $1.5M a year, Tax_History rows 82-89)")
put(al, f"B{tot+15}", "=B4+B5", BLK, CUR, bold=True); put(al, f"C{tot+15}", "=C4+C5", BLK, CUR, bold=True)

put(al, f"A{tot+17}", "THE TRANSFORMATIVE CHECK (v4.5: enrollment lever re-based on recovered leakage students at $4,236 each; gap re-based on the trending fiscal 2026 ledger; capacity anchored on the district advisor's June 2026 presentation)", SEC)
put(al, f"A{tot+18}", "Plan levers at the website defaults: 275 of 550 leakage students recovered (275 x $4,236 = $1,164,900) + costs at the $760K low end"); put(al, f"B{tot+18}", 1922150, BLUE, CUR)
put(al, f"C{tot+18}", "CORRECTED twice: an earlier release used $1.11M-$3.33M for the enrollment lever; the Move 2 rows price the near-term band at $260K-$530K; the website slider now prices the lever directly as recovered leakage students, 0 to the measured 550-student pool, at the $4,236 net-of-supplies cell in B49 legs. Every 100 recovered add $423,600 a year on top of this check.", NOTE)
put(al, f"A{tot+19}", "Full 2018 rate restore (live from Tax_History D79, certified real base; v4.6 correction from the blended $1,699,479)"); put(al, f"B{tot+19}", "=Tax_History!D79", BLK, CUR)
put(al, f"A{tot+20}", "Trending structural gap, fiscal 2026 (June 2026 year-end ledger, before transfers)"); put(al, f"B{tot+20}", 1738653, BLUE, CUR)
put(al, f"C{tot+20}", "District's own June 2026 GL: $20,694,287 of revenue against $22,432,940 of spending before transfers (fund balance $3,328,472 to $2,954,484 after $1,409,590 of transfers in; on-behalf cancels in the gap). The audited fiscal 2025 gap, Assumptions!B24-B21, was $2,648,086.", NOTE)
put(al, f"A{tot+21}", "Recurring surplus after the trending gap (website defaults + full restore)"); put(al, f"B{tot+21}", f"=B{tot+18}+B{tot+19}-B{tot+20}", BLK, CUR, bold=True)
put(al, f"A{tot+22}", "5% raise, every certified teacher (GF certified payroll x 5% x 1.0145 employer Medicare)")
put(al, f"B{tot+22}", "=10000388*0.05*1.0145", BLK, CUR)
put(al, f"C{tot+22}", "GF object 0110 total, FY2026 working budget (archived build/fy2026_working_budget.pdf)", NOTE)
put(al, f"A{tot+23}", "Left for new debt service"); put(al, f"B{tot+23}", f"=B{tot+21}-B{tot+22}", BLK, CUR)
put(al, f"A{tot+24}", "New GF-leveraged bonds supported (4.5%, 20 years; same basis as Debt_Service)")
put(al, f"B{tot+24}", f"=B{tot+23}*(1-1.045^-20)/0.045", BLK, CUR, bold=True)
put(al, f"A{tot+25}", "Bonding capacity per the district's own advisor (Baird, June 2026; build/baird_lpc_june2026.pdf)"); put(al, f"B{tot+25}", 32000000, BLUE, CUR)
put(al, f"C{tot+25}", "From $3,252,893 of FY2027 bondable restricted revenues. Real only if the restricted stream pays for buildings: the $1.32M-a-year capital-to-operations sweep consumes about $17M of this capacity.", NOTE)
put(al, f"A{tot+26}", "Building capacity, together"); put(al, f"B{tot+26}", f"=B{tot+24}+B{tot+25}", BLK, CUR, bold=True)
put(al, f"A{tot+27}", "Reading: at the website defaults (half the pool recovered, costs at the low end, full restore) the plan runs about $1.66M ahead of the trending fiscal 2026 gap, funds the 5 percent certified raise, and leaves about $1.16M for debt: about $15.0M of new GF-leveraged bonds plus the advisor's $32M, about $47 million of building capacity with every school open. The zero-recovery floor still clears the gap with about $500K to spare, within $7,000 of the full raise; two recovered students close the difference. At the slider top (all 550 recovered, levers high) the plan runs about $3.4M ahead: the raise plus about $37M of bonds, about $69M of capacity. Every 100 recovered leakage students move the surplus by $423,600 a year. The earlier 10-percent-raise / $52M top-end claim is withdrawn with the lever correction.", NOTE, wrap=True)
put(al, f"A{tot+28}", "Baird sensitivities, their own June 2026 numbers: minus 50 students drops capacity to $31M (about $20,000 of bonding capacity per student); rates 100bp lower raise it to $35M. Unexpired SFCC offers of $126,250 a year expire January 2028 through January 2034.", NOTE, wrap=True)
put(al, f"A{tot+16}", "Growth plan total (equals the raw sum above; the published band is the conservative cut of the same rows)")
put(al, f"B{tot+16}", f"=B{tot+13}+B{tot+14}+B{tot+15}", BLK, CUR, bold=True); put(al, f"C{tot+16}", f"=C{tot+13}+C{tot+14}+C{tot+15}", BLK, CUR, bold=True)
put(al, f"A{tot+29}", "NMES is not the reason salaries cannot rise (its honest excess cost is about $156,000, six tenths of a percent of the budget) and not the reason capital waits (about $17.6M of restricted capacity sits unused while the sweep drains the building fund). The structural problem is district-wide; so is the fix.", NOTE, wrap=True)

# ================= DEBT_SERVICE =================
d = sheet("Debt_Service", [16, 16, 14, 12, 18, 44])
put(d, "A1", "Outstanding Bonds: Bourbon County School District Finance Corporation", TITLE)
put(d, "A2", "Source: FY2025 audited financial statements, Note 4. Facility funds and debt are restricted; they cannot pay operating costs.", NOTE)
for i, h in enumerate(["Series", "Original", "Rate", "Maturity", "Outstanding 6/30/25", "Note"]):
    put(d, f"{get_column_letter(i+1)}4", h, BOLDW, fill=HDR)
bonds = [
 ("2013", 2255000, "1.90-2.10%", "2026", 348000, ""),
 ("2013R", 468000, "2.75-4.05%", "2033", 585000, "Audit figures internally inconsistent; maturity typo in audit; district to correct"),
 ("2016", 5700000, "1.00-3.00%", "2029", 3145000, "Refunded $5,315,000 of 2009 bonds; NPV savings $314,834"),
 ("2018", 1850000, "3.50%", "2038", 1560000, ""),
 ("2020", 3620000, "0.50-1.85%", "2031", 3405000, "Refunded $3,410,000 of 2011 bonds; NPV savings $106,627"),
 ("2023", 810000, "3.65-4.00%", "2034", 755000, "Purpose to be confirmed from official statement"),
 ("2024", 6055000, "4.00-5.00%", "2044", 5945290, "Funds active construction program; purpose to be published"),
]
r = 5
for s, orig, rate, mat, out, nn in bonds:
    put(d, f"A{r}", s)
    put(d, f"B{r}", orig, BLUE, CUR)
    put(d, f"C{r}", rate)
    put(d, f"D{r}", mat)
    put(d, f"E{r}", out, BLUE, CUR)
    put(d, f"F{r}", nn, NOTE, wrap=True)
    r += 1
put(d, f"A{r}", "Total", bold=True)
bb = put(d, f"B{r}", f"=SUM(B5:B{r-1})", BLK, CUR, bold=True); bb.border = TOPLINE
ee = put(d, f"E{r}", f"=SUM(E5:E{r-1})", BLK, CUR, bold=True); ee.border = TOPLINE
r += 2
put(d, f"A{r}", "ANNUAL DEBT SERVICE", SEC); r += 1
put(d, f"A{r}", "District-paid, FY2025"); put(d, f"B{r}", 1150216, BLUE, CUR); r += 1
put(d, f"A{r}", "District-paid, FY2026"); put(d, f"B{r}", 1578700, BLUE, CUR); r += 1
put(d, f"A{r}", "Increase, FY2025 to FY2026"); put(d, f"B{r}", f"=B{r-1}-B{r-2}", BLK, CUR); r += 1
put(d, f"A{r}", "FY2026 total including state (SFCC) share"); put(d, f"B{r}", 1846159, BLUE, CUR); r += 1
put(d, f"A{r}", "SFCC-paid principal over life of bonds"); put(d, f"B{r}", 1568809, BLUE, CUR)
r += 2
put(d, f"A{r}", "BONDING CAPACITY: WHAT CLOSURE CAN AND CANNOT CHANGE", SEC); r += 1
put(d, f"A{r}", "Average Daily Attendance, FY2025 (SEEK basis)"); put(d, f"B{r}", 2242.5, BLUE, '0.0'); put(d, f"F{r}", "FY2025 audit", NOTE); ada_r = r; r += 1
put(d, f"A{r}", "Capital outlay allotment per year (KRS 157.420, $100 per ADA)"); put(d, f"B{r}", f"=100*B{ada_r}", BLK, CUR); co_r = r; r += 1
put(d, f"A{r}", "Bondable share of capital outlay (702 KAR 4:160 safety factor)"); put(d, f"B{r}", 0.8, BLUE, PCT); sh_r = r; r += 1
put(d, f"A{r}", "Building-fund and debt-fund property tax, FY2025 (the 'nickel' stream)"); put(d, f"B{r}", "=Tax_History!B33", GRN, CUR); bf_r = r; r += 1
put(d, f"A{r}", "District-paid debt service, FY2026"); put(d, f"B{r}", 1578700, BLUE, CUR); ds_r = r; r += 1
put(d, f"A{r}", "Annual restricted margin available for new debt (illustrative)")
mm = put(d, f"B{r}", f"=B{co_r}*B{sh_r}+B{bf_r}-B{ds_r}", BLK, CUR); mm.border = TOPLINE
put(d, f"F{r}", "Simplified: KDE's official bonding potential statement is the authority and should be published", NOTE, wrap=True); r += 1
put(d, f"A{r}", "Unused bonding capacity stated in the FY2024 audit"); put(d, f"B{r}", 23500000, BLUE, CUR); put(d, f"F{r}", "FY2024 audit note", NOTE); r += 2
put(d, f"A{r}", "Why closing NMES does not create bonding capacity: capacity is built from the streams above, none of which", NOTE); r += 1
put(d, f"A{r}", "grows when a school closes. Each student who leaves the district subtracts $100 per year from capital outlay", NOTE); r += 1
put(d, f"A{r}", "and the SEEK base from operations. Sale proceeds are one-time and restricted to capital use. What a closure", NOTE); r += 1
put(d, f"A{r}", "changes is the facility plan's priority list, which steers SFCC offers (KRS 157.622) toward other projects.", NOTE); r += 1
put(d, f"A{r}", "That is a choice about priorities, not a gain in capacity, and it should be argued openly with the BG-1,", NOTE); r += 1
put(d, f"A{r}", "the official statement, and the bonding potential statement all public.", NOTE)
r += 2
put(d, f"A{r}", "THE $14 MILLION PLAN (JULY 15, 2026 PLANNING COMMITTEE)", SEC); r += 1
put(d, f"A{r}", "Stated plan: free up $800,000 to $1,000,000 a year of operating money to bond $14 million", NOTE); r += 1
put(d, f"A{r}", "Proposed bond amount"); put(d, f"B{r}", 14000000, BLUE, CUR); bond_r = r; r += 1
put(d, f"A{r}", "Assumed interest rate"); put(d, f"B{r}", 0.045, BLUE, PCT); rate_r = r; r += 1
put(d, f"A{r}", "Assumed term, years"); put(d, f"B{r}", 20, BLUE, '0'); term_r = r; r += 1
put(d, f"A{r}", "Annual debt service on the proposed bond")
pmt = put(d, f"B{r}", f"=B{bond_r}*B{rate_r}/(1-(1+B{rate_r})^-B{term_r})", BLK, CUR); pmt.border = TOPLINE
put(d, f"F{r}", "Payment is approximately the operating amount the plan frees up; compare the savings scenarios below", NOTE, wrap=True); pmt_r = r; r += 2
put(d, f"A{r}", "WHAT EACH SAVINGS ESTIMATE COULD ACTUALLY BOND", SEC); r += 1
put(d, f"A{r}", "Bond principal supported = annual savings x present-value annuity factor at the rate and term above", NOTE); r += 1
sav_rows = [
 ("District's own KDE-filed excess cost of NMES vs peer elementaries", 121220, "Above the closure model's central case, which is itself negative (-$410,806); the weighted median loses $428,627"),
 ("Closure model median (v5.0 weighted grid)", -428627, "Closure_Model tab, 972-combination grid; the median scenario loses money"),
 ("Closure model best case (favorable tail)", -11030, "Closure_Model tab B49"),
 ("Administration's claim, July 15, 2026", 900000, "Unpublished derivation; reconcile with KDE-filed school-level spending"),
]
for label, sv, note6 in sav_rows:
    put(d, f"A{r}", label); put(d, f"B{r}", sv, BLUE, CUR)
    put(d, f"C{r}", f"=B{r}*(1-(1+B${rate_r})^-B${term_r})/B${rate_r}", BLK, CUR)
    put(d, f"F{r}", note6, NOTE, wrap=True); r += 1
put(d, f"A{r}", "Only the administration's own number reaches $14 million. The audited excess cost supports about $1.6 million.", NOTE); r += 2
put(d, f"A{r}", "DISTRICT-PAID DEBT SERVICE SCHEDULE (FY2025 AUDIT, NOTE 4)", SEC); r += 1
put(d, f"A{r}", "Payments fall as the 2013, 2016, and 2020 series retire; this falling schedule is the room a wrap-around structure fills", NOTE); r += 1
for yr_label, amt in [("FY2026", 1578700), ("FY2027", 1575060), ("FY2028", 1578719), ("FY2029", 1577258), ("FY2030", 1578340),
                      ("FY2031-35 average", 1321112), ("FY2036-40 average", 398915), ("FY2041-45 average", 260708)]:
    put(d, f"A{r}", yr_label); put(d, f"B{r}", amt, BLUE, CUR); r += 1
r += 1
put(d, f"A{r}", "THE LEVERS THAT DO NOT CLOSE A SCHOOL", SEC); r += 1
put(d, f"A{r}", "Remaining restricted capacity (FY2024 audit $23.5M less local share of the 2024 issue, approximate)")
put(d, f"B{r}", 17600000, BLUE, CUR); put(d, f"F{r}", "Exact figure is the fiscal agent's bonding potential statement; demand it", NOTE, wrap=True); rrc_r = r; r += 1
put(d, f"A{r}", "Certified real and personal property assessment, FY2025"); put(d, f"B{r}", 1843569625, BLUE, CUR); asmt_r = r; r += 1
put(d, f"A{r}", "Recallable nickel status: ALREADY LEVIED August 17, 2023, inside the existing rate"); put(d, f"B{r}", "levied", NOTE)
put(d, f"F{r}", "KDE Nickel Levy Chart (March 2024) dates Bourbon's recallable levy 8/17/2023; KDE's levied-rates file (April 30, 2026) decomposes the 52.4-cent rate as 41.0 general fund + 5.7 FSPK + 5.7 recallable. Paris Independent, for scale: 71.5 cents with 17.4 recallable", NOTE, wrap=True); r += 1
put(d, f"A{r}", "New annual state equalization on the recallable nickel (full value, FY2027 schedule)"); put(d, f"B{r}", 276246, BLUE, CUR)
put(d, f"F{r}", "Phasing in on KDE SEEK schedules: $82,866 FY2025, $55,515 FY2026, $276,246 FY2027. New restricted revenue, no board action required", NOTE, wrap=True); nick_r = r; r += 1
put(d, f"A{r}", "Bonds the new equalization alone supports")
put(d, f"B{r}", f"=B{nick_r}*(1-(1+B{rate_r})^-B{term_r})/B{rate_r}", BLK, CUR)
put(d, f"F{r}", "Additive to the FY2024 audit's bonding potential, which predates the equalization", NOTE, wrap=True); r += 1
put(d, f"A{r}", "SFCC offers of assistance, typical cycle"); put(d, f"B{r}", 1750000, BLUE, CUR); put(d, f"F{r}", "State already pays $1,568,809 of current principal", NOTE, wrap=True); r += 2
put(d, f"A{r}", "FY2026 YEAR-END, DISTRICT'S OWN KDE BUDGET MONITORING TOOL (UNAUDITED)", SEC); r += 1
put(d, f"A{r}", "June 2026 financial packet, board agenda July 16, 2026; MUNIS run July 15, 2026. Audit will finalize these figures.", NOTE); r += 1
put(d, f"A{r}", "General Fund revenue, FY2026 actual (excludes carryforward and on-behalf)"); put(d, f"B{r}", 22103877, BLUE, CUR); fy26rev_r = r; r += 1
put(d, f"A{r}", "General Fund expenditures, FY2026 actual"); put(d, f"B{r}", 22477866, BLUE, CUR); fy26exp_r = r; r += 1
put(d, f"A{r}", "Net General Fund change, FY2026 (unaudited)")
nn26 = put(d, f"B{r}", f"=B{fy26rev_r}-B{fy26exp_r}", BLK, CUR); nn26.border = TOPLINE
put(d, f"F{r}", "Compare net changes of -$1,065,657 in FY2024 and -$1,225,465 in FY2025 (audited)", NOTE, wrap=True); r += 1
put(d, f"A{r}", "Revenues over budget per the monitoring tool"); put(d, f"B{r}", 2225835, BLUE, CUR); r += 1
put(d, f"A{r}", "Salaries under budget per the monitoring tool"); put(d, f"B{r}", 587592, BLUE, CUR); r += 1
put(d, f"A{r}", "Salaries below FY2025 actuals"); put(d, f"B{r}", 223974, BLUE, CUR); r += 1
put(d, f"A{r}", "Caveat 1: miscellaneous revenue (object 1990) budgeted at zero, received"); put(d, f"B{r}", 1567829, BLUE, CUR)
put(d, f"F{r}", "Unbudgeted even in the final amended budget; the district should identify this receipt on the record", NOTE, wrap=True); misc_r = r; r += 1
put(d, f"A{r}", "  of which booked in June (period 12) alone"); put(d, f"B{r}", 1413929, BLUE, CUR)
put(d, f"F{r}", "Months 1-11 produced $153,900 combined, a normal run rate; the balance sheet shows no receivable behind the June entry", NOTE, wrap=True); jmisc_r = r; r += 1
put(d, f"A{r}", "Caveat 2: restricted capital money transferred INTO the General Fund in June 2026"); put(d, f"B{r}", 1320939, BLUE, CUR)
put(d, f"F{r}", "Object 5210: $1,098,663 from the Building Fund (320) + $222,276 from Capital Outlay (310), which ended the year at $0 ($1,098,663 + $222,276 = $1,320,939 per the GF receipts ledger; the packet's fund 320 page records $1,098,633, a $30 internal discrepancy in the district's own packet). Lawful under budget-act capital funds flexibility. The Building Fund piece was budgeted from September 2025 ($1,120,203); the $222,276 Capital Outlay piece appears in no version of the FY2026 budget", NOTE, wrap=True); xfer_r = r; r += 1
put(d, f"A{r}", "Net change excluding the capital transfer")
put(d, f"B{r}", f"=B{fy26rev_r}-B{fy26exp_r}-B{xfer_r}", BLK, CUR)
put(d, f"F{r}", "About -$1.69 million: the operating result on operating revenue alone", NOTE, wrap=True); r += 1
put(d, f"A{r}", "Net change excluding the transfer and the unidentified June receipt")
put(d, f"B{r}", f"=B{fy26rev_r}-B{fy26exp_r}-B{xfer_r}-B{jmisc_r}", BLK, CUR)
put(d, f"F{r}", "About -$3.11 million: the same range as the audited years", NOTE, wrap=True); r += 1
put(d, f"A{r}", "June 2026 General Fund revenue vs June 2025 (district's monitoring tool)"); put(d, f"B{r}", 3923096, BLUE, CUR); put(d, f"C{r}", 1130736, BLUE, CUR)
put(d, f"F{r}", "The two June entries above are 98 percent of the year-over-year June difference", NOTE, wrap=True); r += 1
put(d, f"A{r}", "Variance identity from the packet: projected fund balance reconciles to the dollar")
put(d, f"B{r}", "=1489853+2225835+587592-224087", BLK, CUR)
put(d, f"F{r}", "Contingency + revenue over budget + salary savings - expense overrun = $4,079,193, the tool's projected balance. The June misc entry is 63 percent of the revenue beat", NOTE, wrap=True); r += 1
put(d, f"A{r}", "BUILDING FUND FLOWS, FY2026 (fund 320, same packet)", SEC); r += 1
put(d, f"A{r}", "Building Fund transfers out, FY2026 total"); put(d, f"B{r}", 2481394, BLUE, CUR); bfout_r = r; r += 1
put(d, f"A{r}", "  to Debt Service (fund 400)"); put(d, f"B{r}", 1382761, BLUE, CUR); bfds_r = r; r += 1
put(d, f"A{r}", "  to the General Fund (June 2026)")
put(d, f"B{r}", f"=B{bfout_r}-B{bfds_r}", BLK, CUR)
put(d, f"F{r}", "The residual after debt service. This is the restricted stream a $14 million bond would lean on; in FY2026 it plugged the operating budget instead", NOTE, wrap=True); bfres_r = r; r += 1
put(d, f"A{r}", "Bond capacity that residual supports at the same rate and term")
put(d, f"B{r}", f"=B{bfres_r}*(1-(1+B{rate_r})^-B{term_r})/B{rate_r}", BLK, CUR)
put(d, f"F{r}", "About $14 million: the plan's own target, carried by restricted money. The closure's role is only to replace the sweep, and any recurring million dollars does that", NOTE, wrap=True)
r += 2
put(d, f"A{r}", "SCENARIO: BALANCE THE BUDGET AND EXPAND BONDING CAPACITY, NO CLOSURE", SEC); r += 1
put(d, f"A{r}", "Two cases differ only in the operating gap assumed. Conservative treats the FY2026 $1.57M receipt as one-time;", NOTE); r += 1
put(d, f"A{r}", "trend uses the unaudited FY2026 net change. The sweep is ended in both cases; only genuine remainder services debt.", NOTE); r += 1
put(d, f"B{r}", "Conservative", BOLDW, fill=HDR); put(d, f"C{r}", "FY2026 trend", BOLDW, fill=HDR); r += 1
put(d, f"A{r}", "Operating gap to close first")
put(d, f"B{r}", 1787918, BLUE, CUR); put(d, f"C{r}", 373989, BLUE, CUR)
put(d, f"F{r}", "Editable. $1.9M = FY2026 result excluding the unidentified receipt; $373,989 = FY2026 unaudited net change", NOTE, wrap=True); gap_r = r; r += 1
put(d, f"A{r}", "New recurring revenue: 4 percent levy taken three years running (mechanics case; with the published 2018 restore instead, the gap clears in every case: Alternatives transformative check)")
put(d, f"B{r}", "=Tax_History!B32*(1.04^3-1)", GRN, CUR); put(d, f"C{r}", "=Tax_History!B32*(1.04^3-1)", GRN, CUR)
put(d, f"F{r}", "Same base and compounding as the levy tab; year 1 is $313,162, year 3 cumulative $977,568", NOTE, wrap=True); lev_r = r; r += 1
put(d, f"A{r}", "Recurring cost reductions, not from closing a school")
put(d, f"B{r}", 1000000, BLUE, CUR); put(d, f"C{r}", 1000000, BLUE, CUR)
put(d, f"F{r}", "Editable. For scale from the FY2025 audit: district administration grew $221K in one year, salaries fell $224K through attrition, spending ran $859K under budget", NOTE, wrap=True); cut_r = r; r += 1
put(d, f"A{r}", "End the capital-to-GF sweep (required before the building-fund residual can be pledged to new bonds)")
put(d, f"B{r}", 1320939, BLUE, CUR); put(d, f"C{r}", 1320939, BLUE, CUR)
put(d, f"F{r}", "The FY2026 sweep is General Fund revenue today; pledging the residual means replacing it. Without this row the restricted stream would be counted twice", NOTE, wrap=True); swp_r = r; r += 1
put(d, f"A{r}", "Operating room left for debt service after the budget balances and the sweep ends")
for col in ("B", "C"):
    put(d, f"{col}{r}", f"=MAX(0,{col}{lev_r}+{col}{cut_r}-{col}{gap_r}-{col}{swp_r})", BLK, CUR)
put(d, f"F{r}", "Trend case balances with about $283K to spare; the conservative case is still about $1.24M short, so its total below leans on equalization and existing restricted capacity only", NOTE, wrap=True)
room_r = r; r += 1
put(d, f"A{r}", "General-fund bond capacity from that room")
for col in ("B", "C"):
    put(d, f"{col}{r}", f"={col}{room_r}*(1-(1+B${rate_r})^-B${term_r})/B${rate_r}", BLK, CUR)
gfb_r = r; r += 1
put(d, f"A{r}", "New state equalization on the already-levied recallable nickel")
put(d, f"B{r}", 276246, BLUE, CUR); put(d, f"C{r}", 276246, BLUE, CUR)
put(d, f"F{r}", "FY2027 SEEK schedule; the recallable nickel itself is already levied and its local yield is already inside the building-fund stream above", NOTE, wrap=True)
nks_r = r; r += 1
put(d, f"A{r}", "Bond capacity from the new equalization")
for col in ("B", "C"):
    put(d, f"{col}{r}", f"={col}{nks_r}*(1-(1+B${rate_r})^-B${term_r})/B${rate_r}", BLK, CUR)
nkb_r = r; r += 1
put(d, f"A{r}", "Existing unused restricted capacity, approximate")
put(d, f"B{r}", 17600000, BLUE, CUR); put(d, f"C{r}", 17600000, BLUE, CUR)
put(d, f"F{r}", "FY2024 audit $23.5M less local share of the 2024 issue; fiscal agent's statement is the authority", NOTE, wrap=True); ex_r = r; r += 1
put(d, f"A{r}", "Total capacity available without closing a school", bold=True)
for col in ("B", "C"):
    cc = put(d, f"{col}{r}", f"={col}{gfb_r}+{col}{nkb_r}+{col}{ex_r}", BLK, CUR, bold=True); cc.border = TOPLINE
r += 1
put(d, f"A{r}", "Roughly $22.2 million in the conservative case and $42.1 million on the FY2026 trend, with the budget balanced", NOTE); r += 1
put(d, f"A{r}", "first in both. The administration's plan reaches $32 million at face value and leaves the deficit in place.", NOTE); r += 1
put(d, f"A{r}", "Floor check: even with zero general-fund room, the restricted capacity plus new equalization alone are about $21 million.", NOTE); r += 1
put(d, f"A{r}", "Pledging the building-fund residual to new bonds requires ending the capital-to-operations sweep documented above,", NOTE); r += 1
put(d, f"A{r}", "which is exactly what the administration's plan implies; the only question is which recurring million replaces it.", NOTE)

# ================= RUNWAY =================
rw = sheet("Runway", [52, 14, 14, 14, 14])
put(rw, "A1", "Reserve Runway: Where the Fund Balance Goes Under Each Path", TITLE)
put(rw, "A2", "Simplified straight-line projection from the FY2025 ending balance; excludes raises, inflation, and one-time items.", NOTE)
for col, yr in zip("BCDE", ["FY2026", "FY2027", "FY2028", "FY2029"]):
    put(rw, f"{col}4", yr, BOLDW, fill=HDR)
put(rw, "A4", "Projected ending General Fund balance", BOLDW, fill=HDR)
put(rw, "A5", "Status quo (current drawdown continues)")
put(rw, "B5", "=GF_Summary!D9-GF_Summary!$D$16", BLK, CUR)
for col, prev in zip("CDE", "BCD"):
    put(rw, f"{col}5", f"={prev}5-GF_Summary!$D$16", BLK, CUR)
put(rw, "A6", "With alternatives package (conservative midpoint; half effect FY2027, full after)")
put(rw, "B6", "=GF_Summary!D9-GF_Summary!$D$16", BLK, CUR)
put(rw, "C6", "=B6-GF_Summary!$D$16+0.5*Alternatives!$B$19", BLK, CUR)
put(rw, "D6", "=C6-GF_Summary!$D$16+Alternatives!$B$19", BLK, CUR)
put(rw, "E6", "=D6-GF_Summary!$D$16+Alternatives!$B$19", BLK, CUR)
put(rw, "A7", "Closure only (central case, $52,514, from FY2027)")
put(rw, "B7", "=GF_Summary!D9-GF_Summary!$D$16", BLK, CUR)
for col, prev in zip("CDE", "BCD"):
    put(rw, f"{col}7", f"={prev}7-GF_Summary!$D$16+Closure_Model!$B$47", BLK, CUR)
put(rw, "A9", "Closure range check (v5.0): even the -$11,030 best case drains reserves faster than status quo, the -$428,627 weighted median much faster; every weighted scenario loses money.", NOTE, wrap=True)
put(rw, "A8", "2% contingency floor (approx., FY2025 basis)")
for col in "BCDE":
    put(rw, f"{col}8", "=GF_Summary!$D$14", GRN, CUR)
put(rw, "A10", "Reading: the alternatives package restores balance faster than closure, keeps every school open, and adds enrollment revenue rather than risking it.", NOTE, wrap=True)

# ================= SCENARIOS =================
sc = sheet("Scenarios", [50, 18, 18, 26, 52])
put(sc, "A1", "Three Complete Plans, Compared (illustrative five-year view)", TITLE)
put(sc, "A2", "Each row is a full operating plan under this workbook's stated assumptions; the district should replace them with actuals. "
              "One-time closure transition costs (moving, receiving-school additions, building carrying or disposal) have not been published and are not included. Coverage percentages use the $2.65M before-transfers gap; these balance projections use the net drawdown after transfers.", NOTE, wrap=True)
schdrs = ["Plan", "Recurring GF impact ($/yr, by yr 3)", "Projected FY2029 balance", "One-time costs", "What it requires and risks"]
for i, h in enumerate(schdrs):
    put(sc, f"{get_column_letter(i+1)}4", h, BOLDW, fill=HDR)
put(sc, "A5", "1. Districtwide status quo (change nothing anywhere)")
put(sc, "B5", 0, BLUE, CUR)
put(sc, "C5", "=Runway!E5", BLK, CUR)
put(sc, "D5", "None", NOTE)
put(sc, "E5", "No decisions; the districtwide drawdown, which NMES did not cause, simply continues on the straight line with or without the school", NOTE)
put(sc, "A6", "2. Close NMES and consolidate")
put(sc, "B6", -428627, BLUE, CUR)
put(sc, "C6", "=Runway!E7", BLK, CUR)
put(sc, "D6", "Unpublished", NOTE)
put(sc, "E6", "Closure vote; the weighted median LOSES $428,627 a year (B6 is the grid median; the legacy single-point Closure_Model!B20 is superseded); the grid central case itself loses $410,806, and even the best-case tail still loses $11,030 a year; longer rides; measured enrollment-loss risk", NOTE)
put(sc, "A7", "3. Districtwide recovery plan (menu plus levy; includes rebalancing and growing NMES)")
put(sc, "B7", "=Alternatives!B19", GRN, CUR)
put(sc, "C7", "=Runway!E6", BLK, CUR)
put(sc, "D7", "Varies by measure", NOTE)
put(sc, "E7", "Revenue votes, administrative rollback, boundary action, HB 563 recruitment and a signature program at NMES (advanced learners is one option), implementation discipline; every school stays open", NOTE)
put(sc, "A9", "Reading: the question before the board is not closure versus no closure. It is which complete plan produces the best verified five-year result. "
              "Plan 2 buys roughly one extra year of runway; Plan 3 restores balance while keeping every school open. The rebalance-and-grow scenario "
              "(Redistricting tab) is one line inside Plan 3's menu.", NOTE, wrap=True)




# ================= TAX_HISTORY =================
th = sheet("Tax_History", [36, 13, 13, 13, 13, 13, 48])
put(th, "A1", "Property Tax Rates, Fund Split, Delinquency, and the 4% Option", TITLE)
put(th, "A2", "Backs Section 10 and Figures 15 and 16 of the report. Rates in cents per $100. DOR rate books primary for 2023-2025; 2018-2022 verified secondary; 2005-2017 not retrieved and not interpolated.", NOTE)

put(th, "A4", "BOURBON COUNTY SCHOOLS, REAL ESTATE RATE BY TAX YEAR", SEC)
trates = [("2018", 61.3), ("2019", 60.6), ("2020", 55.9), ("2021", 54.2),
          ("2022", 49.2), ("2023", 52.4), ("2024", 52.4), ("2025", 52.4)]
r = 5
for yr, v in trates:
    put(th, f"A{r}", yr); put(th, f"B{r}", v, BLUE, "0.0"); r += 1
put(th, "G5", "2019 documented as the board taking the 4% option (Bourbon County Citizen, 9/4/2019); other years' rate type undetermined pending KDE levy files", NOTE, wrap=True)
put(th, "A14", "Tangible/personal rate (recent)"); put(th, "B14", 64.5, BLUE, "0.0")
put(th, "A15", "Motor vehicle rate"); put(th, "B15", 54.7, BLUE, "0.0")
put(th, "A16", "Utility gross receipts"); put(th, "B16", 0.03, BLUE, PCT)

put(th, "A18", "AREA DISTRICTS, LEVIED REAL ESTATE RATE 2024-25", SEC)
nbrs = [("Fayette County", 80.9), ("Paris Independent", 71.5), ("Clark County", 65.5),
        ("Bath County", 63.4), ("Scott County", 62.9), ("Harrison County", 57.7),
        ("Montgomery County", 52.5), ("Bourbon County", 52.4), ("Nicholas County", 43.1)]
r = 19
for name, v in nbrs:
    put(th, f"A{r}", name); put(th, f"B{r}", v, BLUE, "0.0"); r += 1
put(th, "A28", "Statewide school average (DOR Table II, 2025)"); put(th, "B28", 65.13, BLUE, "0.0")
put(th, "G19", "Fayette and Clark from local reporting of board votes; others from DOR rate books. Bourbon ranks second lowest of the nine. Nicholas shown at its real estate rate of 43.1 (tangible 43.7). Bath 63.4 is the 2025 rate (2024: 60.7).", NOTE, wrap=True)

put(th, "A31", "FY2025 PROPERTY TAX BY FUND (audited)", SEC)
put(th, "A32", "General Fund"); put(th, "B32", 7829060, BLUE, CUR)
put(th, "A33", "Building (FSPK) and debt service funds"); put(th, "B33", 2052786, BLUE, CUR)
put(th, "A34", "Total property tax, FY2025"); put(th, "B34", "=B32+B33", BLK, CUR)
put(th, "A35", "General Fund property tax, FY2024"); put(th, "B35", 7150498, BLUE, CUR)
put(th, "A36", "General Fund revenue, FY2025"); put(th, "B36", 26449318, BLUE, CUR)
put(th, "A37", "General Fund deficit before transfers, FY2025"); put(th, "B37", 2648086, BLUE, CUR)
put(th, "A38", "General Fund expenditures, FY2025"); put(th, "B38", "=B36+B37", BLK, CUR)
put(th, "A39", "Local levy share of General Fund spending"); put(th, "B39", "=B32/B38", BLK, PCT)
put(th, "G32", "GF vs building-fund CENT split unverified; dollar split is audited. Cent split is a records request in the report.", NOTE, wrap=True)

put(th, "A41", "DELINQUENCY CHECK (collections vs certified yield)", SEC)
put(th, "B41", "FY2024", BOLD); put(th, "C41", "FY2025", BOLD)
put(th, "A42", "Calculated yield at levied rates"); put(th, "B42", 10556809, BLUE, CUR); put(th, "C42", 9880143, BLUE, CUR)
put(th, "A43", "Actual collections"); put(th, "B43", 10168969, BLUE, CUR); put(th, "C43", 9641017, BLUE, CUR)
put(th, "A44", "Gap (ordinary delinquency)"); put(th, "B44", "=B42-B43", BLK, CUR); put(th, "C44", "=C42-C43", BLK, CUR)
put(th, "A45", "Gap as share of yield"); put(th, "B45", "=B44/B42", BLK, PCT); put(th, "C45", "=C44/C42", BLK, PCT)

put(th, "A47", "FOUR PERCENT OPTION, THREE-YEAR PATH (KRS 160.470)", SEC)
put(th, "A48", "Base: FY2025 General Fund real + personal collections"); put(th, "B48", "=B32", GRN, CUR)
put(th, "A49", "Annual revenue option"); put(th, "B49", 0.04, BLUE, PCT)
put(th, "A50", "Year 1 added recurring revenue"); put(th, "B50", "=B48*B49", BLK, CUR)
put(th, "A51", "Year 2 added recurring revenue"); put(th, "B51", "=(B48+B50)*B49", BLK, CUR)
put(th, "A52", "Year 3 added recurring revenue"); put(th, "B52", "=(B48+B50+B51)*B49", BLK, CUR)
put(th, "A53", "Cumulative added annual revenue by Year 3"); put(th, "B53", "=B50+B51+B52", BLK, CUR)
put(th, "A54", "As a share of the FY2025 structural deficit"); put(th, "B54", "=B53/(Assumptions!B24-Assumptions!B21)", BLK, PCT)
put(th, "G48", "Basis: 4% more revenue from EXISTING real + personal property than the compensating rate, on the GENERAL FUND levy only. The restricted building-fund (FSPK, $2,052,786) and debt levies are excluded because that money cannot pay operating costs, the same restricted-funds rule this report applies to closure. New property excluded (upside); motor vehicle separate; recall applies only above 4%.", NOTE, wrap=True)

# ================= DEMOGRAPHICS =================
dm = sheet("Demographics", [30, 12, 30, 6, 8, 8, 6, 8, 8, 44])
put(dm, "A1", "County Demographics and NMES Long-Run Enrollment", TITLE)
put(dm, "A2", "An honest picture: the county is flat to declining, so the growth plan relies on redistricting and cross-county enrollment, not a population rebound.", NOTE)

put(dm, "A4", "BOURBON COUNTY POPULATION (U.S. Census / FRED; KSDC projection)", SEC)
pop = [("1970", 18476), ("1980", 19405), ("1990", 19247), ("2000", 19352),
       ("2010", 19985), ("2020", 20252), ("2024 est.", 20337), ("2040 projection (KSDC)", 19352)]
r = 5
for yr, v in pop:
    put(dm, f"A{r}", yr)
    put(dm, f"B{r}", v, BLUE, NUM)
    r += 1
put(dm, "A13", "Change, 1970 to 2020"); put(dm, "B13", "=B10-B5", BLK, NUM)
put(dm, "A14", "Change as a share of 1970"); put(dm, "B14", "=B13/B5", BLK, PCT)
put(dm, "J5", "Sources: FRED series KYBOUR7POP; Envision 2040 plan citing the Kentucky State Data Center", NOTE, wrap=True)

put(dm, "A16", "ADJACENT COUNTIES: 2020 CENSUS AND OUTLOOK", SEC)
adj = [("Scott (Georgetown)", 57155, "+46.1% projected to 2050"),
       ("Fayette (Lexington)", 322570, "+9.8% projected to 2050"),
       ("Clark (Winchester)", 36972, "Slow growth"),
       ("Montgomery (Mt. Sterling)", 28000, "Slow growth"),
       ("Harrison (Cynthiana)", 18692, "Roughly flat"),
       ("Bath", 12500, "Flat to declining"),
       ("Nicholas (Carlisle)", 7537, "Flat to declining"),
       ("Bourbon", 20252, "About -4% by 2040 (KSDC)")]
r = 17
for name, v, note_t in adj:
    put(dm, f"A{r}", name)
    put(dm, f"B{r}", v, BLUE, NUM)
    put(dm, f"C{r}", note_t, NOTE)
    r += 1

put(dm, "A26", "NORTH MIDDLETOWN CITY POPULATION", SEC)
town = [("2000", 562), ("2010 (approx.)", 521), ("2020", 610), ("2024 est.", 679)]
r = 27
for yr, v in town:
    put(dm, f"A{r}", yr)
    put(dm, f"B{r}", v, BLUE, NUM)
    r += 1

put(dm, "A32", "NMES ENROLLMENT, 1989-2025 (spring of school year; backs Figure 14)", SEC)
hist = [261, 255, 234, 225, 202, 203, 182, 196, 208, 198, 205, 195, 195, 203,
        196, 206, 204, 199, 211, 224, 217, 177, 165, 167, 154, 154, 155, 154,
        131, 131, 160, 160, 148, 153, 145, 135, 128]
years = list(range(1989, 2026))
r = 33
for i in range(19):
    put(dm, f"E{r+i}", str(years[i]), BOLD)
    put(dm, f"F{r+i}", hist[i], BLUE, NUM)
for i in range(19, 37):
    put(dm, f"H{r+i-19}", str(years[i]), BOLD)
    put(dm, f"I{r+i-19}", hist[i], BLUE, NUM)
put(dm, "A33", "Peak enrollment (1988-89)")
put(dm, "B33", "=MAX(F33:F51,I33:I50)", BLK, NUM)
put(dm, "A34", "Latest official count (2024-25)")
put(dm, "B34", "=I50", BLK, NUM)
put(dm, "A35", "Decline from peak")
put(dm, "B35", "=B33-B34", BLK, NUM)
put(dm, "A36", "Decline as a share of peak")
put(dm, "B36", "=B35/B33", BLK, PCT)
put(dm, "A37", "Current rated capacity")
put(dm, "B37", "=Assumptions!B12", GRN, NUM)
put(dm, "J33", "1989-2014 compiled by PublicSchoolReview from federal data; 2015-2025 match NCES directly", NOTE, wrap=True)

put(dm, "A54", "HONEST BOTTOM LINE", SEC)
put(dm, "A55", "Bourbon County faces a real demographic headwind: flat for fifty years, aging, land-constrained, projected down about 4% by 2040, and outside the Scott and Fayette growth corridor. Organic enrollment growth at NMES is unlikely. The growth case therefore rests on redistricting within the district and cross-county enrollment on quality under House Bill 563, plus the school's academic record and community-anchor value.", NOTE, wrap=True)

# ================= SCHOOL_DATA =================
sd = sheet("School_Data", [30] + [7.2] * 18 + [44])
put(sd, "A1", "School Data Backing the Report Figures", TITLE)
put(sd, "A2", "Inputs (blue) transcribed from public sources. The 2007-2025 series below is SchoolDigger's normalized 0-100 index built from KDE test data (a third-party rendering, kept for context). KDE's own official record, pulled directly from the department's historical files, is in the KDE OFFICIAL HISTORY block below and governs every claim in the report.", NOTE)

put(sd, "A4", "NMES ENROLLMENT BY SCHOOL YEAR (backs Figure 14)", SEC)
eyears = ["'15-16", "'16-17", "'17-18", "'18-19", "'19-20", "'20-21", "'21-22", "'22-23", "'23-24", "'24-25"]
ecounts = [154, 131, 131, 160, 160, 148, 153, 145, 135, 128]
put(sd, "A5", "School year", bold=True)
put(sd, "A6", "Students")
for i, (y, v) in enumerate(zip(eyears, ecounts)):
    col = get_column_letter(2 + i)
    put(sd, f"{col}5", y, BOLD)
    put(sd, f"{col}6", v, BLUE, NUM)
put(sd, "T6", "NCES school-level data (as compiled by PublicSchoolReview); 2024-25 = NCES CCD official count", NOTE, wrap=True)
put(sd, "A8", "Rated capacity (2021 facility plan)"); put(sd, "B8", 174, BLUE, NUM)
put(sd, "A9", "Open seats, 2024-25"); put(sd, "B9", "=B8-K6", BLK, NUM)
put(sd, "A11", "Note: the superintendent has publicly said 'around 100'; a '118' figure could not be verified in any official record.", NOTE)

put(sd, "A13", "ELEMENTARY SCORES BY YEAR, 2007-2025 (backs Figures 1 and 2)", SEC)
syrs = [str(y) for y in range(2007, 2020)] + [str(y) for y in range(2021, 2026)]
put(sd, "A14", "School (district)", bold=True)
for i, y in enumerate(syrs):
    put(sd, f"{get_column_letter(2 + i)}14", y, BOLD)
put(sd, "T14", "SchoolDigger rendering of KDE test data. No statewide tests in 2020; NMES 2021 not reported. Assessment systems: CATS/KCCT 2007-11, K-PREP 2012-19, KSA 2021-25.", NOTE, wrap=True)
NAv = None
scores = [
 ("North Middletown (Bourbon Co.)", [56.5, 63.9, 68.6, 87.9, 85.8, 72.5, 67.6, 56.6, 56.9, 48.7, 49.3, 40.0, 50.4, NAv, 47.7, 32.1, 54.1, 58.2]),
 ("Bourbon Central (Bourbon Co.)", [77.5, 81.9, 72.6, 69.6, 63.0, 74.7, 67.6, 51.8, 52.1, 30.0, 34.0, 32.8, 39.9, 20.0, 29.9, 29.0, 23.8, 26.5]),
 ("Cane Ridge (Bourbon Co.)", [35.2, 50.9, 56.2, 65.5, 34.5, 34.0, 49.6, 51.0, 51.1, 57.5, 50.4, 41.4, 38.8, 23.8, 38.7, 34.6, 35.8, 19.3]),
 ("Paris Elementary (Paris Indep.)", [NAv]*16 + [16.8, 12.2]),
 ("Shearer (Clark Co.)", [NAv]*16 + [30.7, 42.3]),
 ("Justice (Clark Co.)", [NAv]*16 + [43.2, 39.3]),
 ("Strode Station (Clark Co.)", [NAv]*16 + [43.2, 34.2]),
 ("Conkwright (Clark Co.)", [NAv]*16 + [15.6, 17.5]),
 ("Northview (Montgomery Co.)", [NAv]*16 + [67.4, 68.9]),
 ("Mapleton (Montgomery Co.)", [NAv]*16 + [65.9, 65.1]),
 ("Nicholas County Elementary (Nicholas Co.)", [NAv]*16 + [15.9, NAv]),
]
rr = 15
for name, vals in scores:
    put(sd, f"A{rr}", name)
    for i, v in enumerate(vals):
        if v is not None:
            put(sd, f"{get_column_letter(2 + i)}{rr}", v, BLUE, "0.0")
    rr += 1

put(sd, "T25", "Nicholas County: 2024 value; 2025 not retrieved", NOTE)
put(sd, "A26", "Three-year average, 2023-2025: NMES / Bourbon Central / Cane Ridge")
put(sd, "B26", "=AVERAGE(Q15:S15)", BLK, "0.0")
put(sd, "C26", "=AVERAGE(Q16:S16)", BLK, "0.0")
put(sd, "D26", "=AVERAGE(Q17:S17)", BLK, "0.0")

put(sd, "A28", "NMES DETAIL, 2024-25 (backs Section 5)", SEC)
put(sd, "B29", "NMES", BOLD); put(sd, "C29", "Kentucky", BOLD)
detail = [
 ("Reading, proficient or better", 0.50, 0.50, PCT),
 ("Mathematics, proficient or better", 0.44, 0.44, PCT),
 ("Writing, proficient or better", 0.58, 0.38, PCT),
 ("Science, proficient or better", 0.53, 0.37, PCT),
 ("Composite: economically disadvantaged students", 57.1, None, "0.0"),
 ("Composite: female students", 85.7, None, "0.0"),
 ("Composite: male students", 28.8, None, "0.0"),
]
rr = 30
for label, nm, ky, fmt in detail:
    put(sd, f"A{rr}", label)
    put(sd, f"B{rr}", nm, BLUE, fmt)
    if ky is not None:
        put(sd, f"C{rr}", ky, BLUE, fmt)
    rr += 1
put(sd, "T30", "SchoolDigger/KDE; economically disadvantaged composite = 62nd percentile statewide; female = 91st", NOTE, wrap=True)

put(sd, "A33", "KDE 2024-25 KSA, PERCENT PROFICIENT OR DISTINGUISHED, ALL STUDENTS (build/kde_ksa_2024_25.json)", SEC)
kde_hdrs = ["Subject", "NMES", "Bourbon Central", "Cane Ridge", "Paris Elem.", "KY elementary avg"]
for i, h in enumerate(kde_hdrs):
    put(sd, f"{get_column_letter(i+1)}34", h, BOLDW, fill=HDR)
kde_rows = [("Reading", 41, 38, 37, 25, 49), ("Mathematics", 31, 28, 27, 29, 43),
            ("Science", 53, 26, "*", "*", 37), ("Social Studies", 36, 31, 27, 22, 38),
            ("Combined Writing", 56, 40, 27, 4, 43)]
rr = 35
for row in kde_rows:
    for i, v in enumerate(row):
        put(sd, f"{get_column_letter(i+1)}{rr}", v, NOTE if i == 0 else BLUE, None if isinstance(v, str) else NUM)
    rr += 1
put(sd, "A41", "NMES is first among all four county elementary schools in every subject and beats the statewide elementary average in science and writing. Asterisks = state-suppressed cells.", NOTE, wrap=True)

put(sd, "A43", "KDE OFFICIAL HISTORY, 2011-12 TO 2024-25 (build/kde_scores_history.json)", SEC)
kh_years = ["'11-12", "'12-13", "'13-14", "'14-15", "'15-16", "'16-17", "'17-18",
            "'18-19", "'20-21", "'21-22", "'22-23", "'23-24", "'24-25"]
put(sd, "A44", "Percent proficient or distinguished", bold=True)
for i, y in enumerate(kh_years):
    put(sd, f"{get_column_letter(2 + i)}44", y, BOLD)
put(sd, "T44", "KDE historical accountability files (all students, elementary level); 2017-19 recovered from Wayback captures of KDE's retired download endpoint. 2019-20 canceled statewide; 2020-21 was a limited COVID administration. CATS-era files (2006-2011) are available from KDE by data request.", NOTE, wrap=True)
kh_rows = [
 ("Reading: North Middletown", [51.9, 59.7, 63.2, 53.2, 56.1, 55.6, 42.3, 48.4, 21.4, 32, 43, 45, 41]),
 ("Reading: Bourbon Central", [62.1, 56.6, 61.2, 55.7, 47.9, 47.9, 49.3, 49.1, 27.1, 43, 37, 36, 38]),
 ("Reading: Cane Ridge", [42.9, 47.2, 54.5, 55.3, 59.3, 57.1, 52.0, 49.6, 26.5, 44, 40, 40, 37]),
 ("Reading: Paris Elementary", [31.0, 34.1, 42.9, 38.9, 41.2, 34.3, 34.4, 35.3, 25.0, 28, 25, 30, 25]),
 ("Mathematics: North Middletown", [53.2, 49.4, 55.3, 59.5, 62.1, 50.8, 47.9, 48.4, 17.9, 31, 30, 45, 31]),
 ("Mathematics: Bourbon Central", [47.1, 47.4, 43.0, 43.6, 40.0, 43.5, 41.8, 45.1, 31.7, 27, 28, 23, 28]),
 ("Mathematics: Cane Ridge", [33.3, 37.7, 43.1, 46.1, 54.4, 45.9, 41.5, 39.0, 25.2, 33, 35, 30, 27]),
 ("Mathematics: Paris Elementary", [29.7, 25.0, 26.8, 22.9, 41.9, 28.7, 30.0, 32.0, 15.3, 23, 32, 27, 29]),
]
rr = 45
for name, vals in kh_rows:
    put(sd, f"A{rr}", name)
    for i, v in enumerate(vals):
        put(sd, f"{get_column_letter(2 + i)}{rr}", v, BLUE, "0.0")
    rr += 1
put(sd, "A53", "NMES led the county in mathematics in every pre-COVID administration on record, 2011-12 through 2018-19: eight straight years.", NOTE, wrap=True)

put(sd, "A55", "Official overall score", bold=True)
for i, y in enumerate(kh_years):
    put(sd, f"{get_column_letter(2 + i)}55", y, BOLD)
put(sd, "T55", "Unbridled Learning overall score 2011-12 to 2015-16 (state classifications noted below); KSA overall indicator rate 2021-22 to 2024-25. The two scales are not comparable to each other. No overall score was issued 2016-17 through 2020-21 by design.", NOTE, wrap=True)
kc_rows = [
 ("North Middletown", [62.6, 68.8, 71.4, 72.1, 79.1, None, None, None, None, 51.9, 62.2, 74.5, 54.0]),
 ("Bourbon Central", [63.2, 63.0, 69.8, 67.8, 56.8, None, None, None, None, 52.8, 56.7, 50.3, 55.4]),
 ("Cane Ridge", [53.3, 58.9, 69.2, 68.5, 65.5, None, None, None, None, 54.3, 51.8, 60.7, 47.8]),
 ("Paris Elementary", [48.0, 49.9, 59.4, 54.8, 69.6, None, None, None, None, 46.1, 45.9, 40.7, 41.9]),
]
rr = 56
for name, vals in kc_rows:
    put(sd, f"A{rr}", name)
    for i, v in enumerate(vals):
        if v is not None:
            put(sd, f"{get_column_letter(2 + i)}{rr}", v, BLUE, "0.0")
    rr += 1
put(sd, "A60", "Classifications: NMES rated Proficient 2012-2015 and Distinguished in 2015-16 at 79.1, the county's only Distinguished elementary rating in the files retrieved. In 2023-24 NMES posted 74.5, first in the county by 14 points.", NOTE, wrap=True)
put(sd, "A62", "Third-party check: the SchoolDigger index above correlates about 0.9 with these official results for Bourbon Central and Cane Ridge but only weakly for NMES, and it names the wrong county leader in three of the ten overlapping years. The official record governs.", NOTE, wrap=True)

# ================= FACILITY_PLANS =================
fp = sheet("Facility_Plans", [42, 12, 12, 12, 12, 50])
put(fp, "A1", "What the District's Own Facility Plans Show", TITLE)
put(fp, "A2", "Sources: District Facilities Plan, KBE approval June 2013 (Wayback Machine capture) and District Facility Plan, KBE approval August 2021 "
              "(currently posted; the 2026 draft plan now pending would replace it). Both archived in this repository under build/. Figures read enrollment/capacity; the 2021 "
              "preschool line (272/200, with the plan's note of 80 full-day plus 192 half-day students) confirms the order.", NOTE, wrap=True)
fhdrs = ["School", "2013 enr", "2013 cap", "2021 enr", "2021 cap", "Note"]
for i, h in enumerate(fhdrs):
    put(fp, f"{get_column_letter(i+1)}4", h, BOLDW, fill=HDR)
frows = [
 ("Bourbon County High School", 881, 637, 799, 704, "Rating rose 637 to 704 between plans; the CTC addition was the 2022-24 in-biennium priority"),
 ("Bourbon County Middle School", 616, 515, 640, 641, ""),
 ("Bourbon Central Elementary", 602, 564, 535, 521, "Approved rating 521; the plan's 549 is a contingent To-Become figure tied to an expansion never built (GSF unchanged at 63,320 through the July 2026 KDE report)"),
 ("Cane Ridge Elementary", 461, 500, 480, 422, "Rated capacity written down 78 seats between plans"),
 ("North Middletown Elementary", 169, 198, 161, 174, "Written down again to 154 in the 2026 draft; building held 261 students at the 1989 peak"),
 ("Preschool/Head Start Center", 296, 180, 272, 200, "Over capacity in both plans"),
]
r = 5
for row in frows:
    for i, v in enumerate(row):
        put(fp, f"{get_column_letter(i+1)}{r}", v, NOTE if i in (0, 5) else BLUE, None if i in (0, 5) else NUM)
    r += 1
put(fp, "A12", "RECEIVING CAPACITY TODAY (current ratings vs 2024-25 enrollment)", SEC)
put(fp, "A13", "Bourbon Central open seats at the approved 521"); put(fp, "B13", "=521-Redistricting!B8", BLK, NUM)
put(fp, "A14", "Cane Ridge students over rated 422"); put(fp, "B14", "=Redistricting!B9-E8", BLK, NUM)
put(fp, "A15", "Net uncommitted seats at both receiving schools"); put(fp, "B15", "=B13-B14", BLK, NUM, bold=True)
put(fp, "A16", "NMES students needing seats on closure"); put(fp, "B16", "=Assumptions!B11", GRN, NUM)
put(fp, "A17", "Shortfall if closure proceeds today"); put(fp, "B17", "=B16-B15", BLK, NUM, bold=True)
put(fp, "A20", "NMES INVESTMENT RECORD IN THE PLANS", SEC)
put(fp, "A21", "2013 plan: NMES major renovation priced (HVAC, media center, kitchen, security vestibule, gym)")
put(fp, "B21", 1594872, BLUE, CUR); put(fp, "C21", 239139, BLUE, CUR); put(fp, "D21", 86959, BLUE, CUR)
put(fp, "E21", "=SUM(B21:D21)", BLK, CUR); put(fp, "F21", "Base renovation + special ed classroom + family resource center", NOTE)
put(fp, "A22", "2021 plan: NMES need re-priced (life safety + accessibility ramp/elevator + major renovation)")
put(fp, "B22", 317660, BLUE, CUR); put(fp, "C22", 325000, BLUE, CUR); put(fp, "D22", 3617530, BLUE, CUR)
put(fp, "E22", "=SUM(B22:D22)", BLK, CUR); put(fp, "F22", "Life safety and accessibility ($642,660) within the 2022-24 biennium; the $3.62M major renovation after it; the biennium's headline priority was the HS Career & Technical Center", NOTE)
put(fp, "A23", "2021 plan: HS Career & Technical Center, in-biennium new construction")
put(fp, "B23", "=227149+1447694+1150886+1075170+772305+832878+1150886", BLK, CUR)
put(fp, "A24", "2021 plan: total district facility need"); put(fp, "B24", 43389464, BLUE, CUR)
put(fp, "A28", "THE 2026 DRAFT (presented July 15, 2026, before the committee's amendment; KDE approval date TBD; KFICS capacity basis)", SEC)
d26hdrs = ["School", "2024-25 SAAR end-of-year membership (the draft plan's enrollment column, relabeled in v5.0; state file build/saar_enrollment_2024_25.xls)", "Draft capacity (KFICS)", "Change vs 2021 rating"]
for i, h in enumerate(d26hdrs):
    put(fp, f"{get_column_letter(i+1)}29", h, BOLDW, fill=HDR)
put(fp, "A30", "Bourbon Central"); put(fp, "B30", 491, BLUE, NUM); put(fp, "C30", 640, BLUE, NUM); put(fp, "D30", "=C30-521", BLK, NUM)
put(fp, "A31", "Cane Ridge"); put(fp, "B31", 461, BLUE, NUM); put(fp, "C31", 547, BLUE, NUM); put(fp, "D31", "=C31-E8", BLK, NUM)
put(fp, "A32", "North Middletown"); put(fp, "B32", 128, GRN, NUM); put(fp, "C32", 154, BLUE, NUM); put(fp, "D32", "=C32-E9", BLK, NUM)
put(fp, "A33", "Paper seats added at the receiving schools, with the draft's new-construction sections reading None"); put(fp, "B33", "=D30+D31", BLK, NUM, bold=True)
put(fp, "A34", "NMES fill rate even at the draft's own 154 rating"); put(fp, "B34", "=B32/C32", BLK, '0.0%', bold=True)
put(fp, "A35", "As presented, the draft still listed North Middletown as Permanent (PS-5). Its headline 2026-28 priority is an $18,600,946 major renovation "
               "of the high school's 1968 and 1981 sections. Annotated attendee copy archived as build/dfp_2026_draft_excerpt.png.", NOTE, wrap=True)

put(fp, "A37", "KFICS ASSESSMENT (RossTarrant Architects, presented July 2026; slides archived as build/kfics_assessment.pdf)", SEC)
khdrs = ["Facility", "Condition need", "Instructional deficiency", "Total (as printed)", "Slide capacity"]
for i, h in enumerate(khdrs):
    put(fp, f"{get_column_letter(i+1)}38", h, BOLDW, fill=HDR)
krows = [
 ("Bourbon County High School", 22300342, 5188113, 27488455, 704),
 ("Bourbon County Middle School", 19132586, 3286105, 22418692, 544),
 ("Cane Ridge Elementary", 10313658, 4073936, 14387595, 397),
 ("Bourbon Central Elementary", 7214895, 1615951, 8840267, 499),
 ("North Middletown Elementary", 5648434, 2881659, 8530093, 154),
 ("Preschool/Head Start Center", 1518664, 6343032, 7861696, 160),
 ("Agriculture Building (part of BCHS)", 1825444, 0, 1825444, ""),
]
r = 39
for name, cond, instr, tot, cap in krows:
    put(fp, f"A{r}", name)
    put(fp, f"B{r}", cond, BLUE, CUR); put(fp, f"C{r}", instr, BLUE, CUR)
    put(fp, f"D{r}", tot, BLUE, CUR)
    if cap != "":
        put(fp, f"E{r}", cap, BLUE, NUM)
    r += 1
put(fp, "A46", "School centers total"); put(fp, "B46", "=SUM(D39:D45)", BLK, CUR, bold=True)
put(fp, "A47", "Support facilities total (slide)"); put(fp, "B47", 7089052, BLUE, CUR)
put(fp, "A48", "District total need"); put(fp, "B48", "=B46+B47", BLK, CUR, bold=True)
put(fp, "A49", "Receiving schools combined (Bourbon Central + Cane Ridge)"); put(fp, "B49", "=D41+D42", BLK, CUR, bold=True)
put(fp, "A50", "Net receiving seats at the slides' own capacities (499 and 397)"); put(fp, "B50", "=(499-Redistricting!B8)-(Redistricting!B9-397)", BLK, NUM, bold=True)
put(fp, "A52", "Notes: components and totals as printed on the slides; Bourbon Central's printed components sum $9,421 below its printed total. "
               "The slides' receiving-school capacities (499, 397) differ from the 2026 draft DFP table (640, 547) and from the approved 2021 plan "
               "(the approved 521, 422; the plan's contingent 549 depends on unbuilt work): three documents, three capacity sets for the same buildings. NMES total need is second lowest among the district's schools.", NOTE, wrap=True)

put(fp, "A55", "2017 PLAN, RECOVERED (KBE approved June 7, 2017; Wayback capture July 1, 2017, archived as build/dfp_wayback_20170701225631.pdf; KBE minutes corroborate)", SEC)
put(fp, "A56", "School", BOLDW, fill=HDR); put(fp, "B56", "2017 enr", BOLDW, fill=HDR); put(fp, "C56", "2017 cap", BOLDW, fill=HDR)
for i, (sname, enr, cap) in enumerate([("Bourbon County High School", 914, 671), ("Bourbon County Middle School", 607, 564),
                                        ("Bourbon Central Elementary", 633, 611), ("Cane Ridge Elementary", 482, 550),
                                        ("North Middletown Elementary", 154, 152), ("Preschool/Head Start Center", 125, 200)]):
    put(fp, f"A{57+i}", sname); put(fp, f"B{57+i}", enr, BLUE, NUM); put(fp, f"C{57+i}", cap, BLUE, NUM)
put(fp, "A63", "NMES listed at 154 enrolled against 152 capacity: OVER capacity in the district's own 2017 plan. Same-building rating trajectory across plans: NMES 198-152-174-154; Bourbon Central 564-611-521-640; Cane Ridge 500-550-422-547.", NOTE, wrap=True)

put(fp, "A65", "KFICS CONDITION INDEX, EVERY STATE REPORT PUBLISHED (KDE KFICS State Reports: official Oct 2023, official Oct 2025, updated report generated July 2, 2026)", SEC)
cihdrs = ["School", "Oct 2023 report", "Oct 2025 report", "Jul 2026 report", "4-yr needs (Jul 2026)", "Replacement value (Jul 2026)"]
for i, h in enumerate(cihdrs):
    put(fp, f"{get_column_letter(i+1)}66", h, BOLDW, fill=HDR)
cirows = [
 ("Bourbon Central Elementary (1988)", 0.887637, 0.819273, 0.823017, 4006243.38, 22636266.80),
 ("Cane Ridge Elementary (1992)", 0.812058, 0.811765, 0.728249, 4796308.00, 17649638.79),
 ("North Middletown Elementary (1948/64)", 0.694064, 0.702133, 0.773295, 3099147.93, 13670417.60),
 ("Bourbon County Middle School (1948)", 0.726249, 0.727501, 0.596145, 12167086.72, 30127350.00),
 ("Bourbon County High School (1968)", 0.809819, 0.802057, 0.792529, 10251532.33, 49411909.14),
]
r = 67
for name, c23, c25, c26, needs, crv in cirows:
    put(fp, f"A{r}", name)
    put(fp, f"B{r}", c23, BLUE, "0.000"); put(fp, f"C{r}", c25, BLUE, "0.000"); put(fp, f"D{r}", c26, BLUE, "0.000")
    put(fp, f"E{r}", needs, BLUE, CUR); put(fp, f"F{r}", crv, BLUE, CUR)
    r += 1
put(fp, "A72", "Check: Condition Index = 1 - (4-year renewal needs / replacement value), NMES"); put(fp, "B72", "=1-E69/F69", BLK, "0.000", bold=True)
put(fp, "A73", "NMES change between inspection cycles (2020-21 to April 2026)"); put(fp, "B73", "=D69-B69", BLK, "0.000", bold=True)
put(fp, "A74", "Notes: the Oct 2023 and Oct 2025 official reports rest on the same 2020-21 inspections (NMES and Cane Ridge Jan 5, 2021; Bourbon Central Apr 14, 2020) with costs updated between reports; "
               "the Jul 2026 report carries the first fresh inspections, completed April 2026 and reviewed by KDE. NMES is the only school whose index improved between inspection cycles, and its $3.1M "
               "four-year repair bill is the smallest of the district's five schools. NMES Educational Suitability prints 0.21725 in both the 2023 and Jul 2026 reports, identical to five decimal places, "
               "so that component appears carried forward rather than re-surveyed. Higher index = healthier building.", NOTE, wrap=True)

put(fp, "A26", "Reading: the receiving schools' approved ratings are 521 (Bourbon Central; the plan's 549 is contingent on an expansion never built) and 422 (Cane Ridge). At current enrollment that is 62 open at Bourbon Central and 31 over at Cane Ridge, "
               "a net 31 uncommitted seats for 128 children (59 only under the plan's contingent 549). NMES's major renovation was priced in 2013 and re-priced higher in 2021, each time scheduled after the then-current biennium. "
               "Its rated capacity fell 198 to 174 between the same two plans while its enrollment fell 169 to 128.", NOTE, wrap=True)

# ================= KY_CLOSURES =================
kc = sheet("KY_Closures", [46, 10, 12, 14, 14, 10, 14, 14, 12, 60])
put(kc, "A1", "Thirty Years of Kentucky Rural School Closures: The Record", TITLE)
put(kc, "A2", "Universe: federal Common Core of Data, every KY public school, every year, 1994-2023 (via the Urban Institute Education Data API). Screens: rural/small-town locale; "
              "enrollment above zero; renames, rebuilds under new IDs, and non-community programs removed; corrupt federal finance rows (12) excluded by an enrollment-plausibility screen. "
              "Full lists archived as build/ky_rural_closures_1995_2023.csv and build/ky_closure_dollar_cases.csv.", NOTE, wrap=True)
put(kc, "A4", "THE UNIVERSE", SEC)
put(kc, "A5", "Rural/small-town school closures since 1995"); put(kc, "B5", 339, BLUE, NUM)
put(kc, "A6", "Towns that lost their last public school (any district; successor-school check applied)"); put(kc, "B6", 72, BLUE, NUM)
put(kc, "A7", "Closure events with clean district finance data"); put(kc, "B7", 163, BLUE, NUM)
put(kc, "A8", "Median district per-pupil spending growth vs state, year before to 3 years after (pct pts)"); put(kc, "B8", -0.05, BLUE, "0.00")
put(kc, "A9", "Same, vs districts of similar size (within 40 percent of enrollment)"); put(kc, "B9", -0.24, BLUE, "0.00")
put(kc, "A10", "Districts growing slower than benchmark after closing"); put(kc, "B10", "=84/163", BLK, '0.0%')
put(kc, "A11", "Closures most like this plan (single small elementary, nothing built): count / raw median per displaced student (inside the plausible window the median is $541; see report Section 5)"); put(kc, "B11", 27, BLUE, NUM); put(kc, "C11", 8440, BLUE, CUR)
put(kc, "A13", "THE DOLLAR TEST: BUDGET GAP VS STATE TREND, CREDITED ENTIRELY TO THE CLOSURE (A GENEROUS CEILING), PER DISPLACED STUDENT", SEC)
kchdrs = ["Case", "Closed", "Students displaced", "Spending, year before", "Spending, 3 yrs after", "State growth", "Counterfactual", "Gap per year", "Gap per student", "Note"]
for i, h in enumerate(kchdrs):
    put(kc, f"{get_column_letter(i+1)}14", h, BOLDW, fill=HDR)
kcases = [
 ("Perry County", 2017, 825, 42356000, 44953000, 0.1323, "NEW West Perry Elementary built for the children moved; scores continued a pre-existing climb"),
 ("Johnson County", 2016, 178, 37827000, 39175000, 0.1339, "Meade Memorial El, pure absorption; scores flat: savings without improvement"),
 ("Pineville Independent", 2014, 508, 5355000, 4662000, 0.0667, "Same-town rebuild into one new combined campus; no community lost a school"),
 ("Leslie County", 2013, 144, 19840000, 18031000, 0.0374, "Middle school folded into existing campus; the record's one savings-plus-gains case"),
 ("Pike County", 2014, 123, 88648000, 90441000, 0.0667, "Majestic Knox Creek El, 1.3 percent of a 9,400-student district; the +8 score case"),
 ("Somerset Independent", 1999, 264, 9161000, 9316000, 0.2045, "City-district grade realignment of a 4-5 center, not a rural town; grades split between existing city schools"),
 ("Rowan County", 2001, 144, 19483000, 21010000, 0.1931, "Farmers El; gap exceeds plausible closure savings"),
 ("Webster County", 2012, 172, 19693000, 19475000, 0.0197, "Slaughters El; town lost its school; scores -2 vs state"),
 ("Butler County", 2003, 94, 14995000, 16806000, 0.2352, "Third District El; gap exceeds plausible closure savings"),
 ("Letcher County", 2007, 67, 30621000, 31987000, 0.1523, "Kingdom Come Settlement El; gap of $49K per child is 4x what a school costs: budget-wide, not closure"),
 ("Muhlenberg County", 2013, 91, 52941000, 45883000, 0.0374, "Career HS; $99K per child gap = post-construction and coal-era budget swings, not closure"),
 ("Wayne County", 2007, 381, 22129000, 23687000, 0.1523, "A J Lloyd Middle"),
 ("Lincoln County", 2016, 286, 40397000, 41252000, 0.1339, "Sixth-grade center reorganization"),
 ("Montgomery County", 2018, 737, 45709000, 46413000, 0.1366, "County-seat grade reshuffle of a 5-6 intermediate center, not a rural town or an elementary; opened NEW Northview Elementary the same year"),
 ("Metcalfe County", 2013, 817, 17028000, 15990000, 0.0374, "Closed 3 schools, half the district displaced; built new Primary and Intermediate centers; scores fell 10.5 vs state: the cautionary pair"),
 ("Adair County", 2006, 479, 22126000, 23240000, 0.2005, "3 rural elementaries; built NEW Adair County Elementary the same year; pre-closure spending spike reverting; gap is 94 percent of everything the district then spent per student"),
 ("Breckinridge County", 1997, 299, 17048000, 16425000, 0.1319, "2 rural K-8 schools, nothing built; gap of $9,606 per child EXCEEDS the district's $6,108 total cost per student then: impossible as closure savings, a data-edge artifact"),
]
r = 15
for name, cy, kids, pre, post, g, note in kcases:
    put(kc, f"A{r}", name); put(kc, f"B{r}", cy, BLUE, NUM); put(kc, f"C{r}", kids, BLUE, NUM)
    put(kc, f"D{r}", pre, BLUE, CUR); put(kc, f"E{r}", post, BLUE, CUR); put(kc, f"F{r}", g, BLUE, '0.0%')
    put(kc, f"G{r}", f"=D{r}*(1+F{r})", BLK, CUR); put(kc, f"H{r}", f"=G{r}-E{r}", BLK, CUR)
    put(kc, f"I{r}", f"=H{r}/C{r}", BLK, CUR, bold=True); put(kc, f"J{r}", note, NOTE, wrap=True)
    r += 1
put(kc, "A33", "THE YARDSTICK: WHAT THIS PLAN REQUIRES PER DISPLACED STUDENT", SEC)
put(kc, "A34", "Plan requirement, low / high ($800K-$1M over the 128 students displaced)")
put(kc, "B34", "=800000/Assumptions!B11", BLK, CUR, bold=True); put(kc, "C34", "=1000000/Assumptions!B11", BLK, CUR, bold=True)
put(kc, "A35", "This report's own model, per displaced student: median / central / best case")
put(kc, "B35", "=-428627/Assumptions!B11", BLK, CUR); put(kc, "C35", "=Closure_Model!B47/Assumptions!B11", BLK, CUR); put(kc, "D35", "=Closure_Model!B49/Assumptions!B11", BLK, CUR)
put(kc, "A36", "Reading: among rural ELEMENTARY closures, the one clean no-construction comparable (Webster 2012) paid $3,525 per displaced student; every case at or near the plan's band built a new school (Perry, Adair, Metcalfe) or was a city or county-seat grade reshuffle (Somerset, Montgomery). Gaps beyond roughly a school's own cost per student are flagged in the notes: they prove budget-wide causes, which is why this model prices closure bottom-up (positions, busing, SEEK) rather than from budget trends.", NOTE, wrap=True)

put(kc, "A38", "THE FULL DISTRIBUTION, PER DISPLACED STUDENT (all measurable events; whole budget gap credited to the closure)", SEC)
put(kc, "A39", "All 163 events: P25 / median / P75"); put(kc, "B39", -1875, BLUE, CUR); put(kc, "C39", 1102, BLUE, CUR); put(kc, "D39", 7520, BLUE, CUR)
put(kc, "A40", "Physically plausible magnitudes only (within $13K, 116 events): P25 / median / P75"); put(kc, "B40", -1433, BLUE, CUR); put(kc, "C40", 818, BLUE, CUR); put(kc, "D40", 4120, BLUE, CUR)
put(kc, "A41", "Share of districts spending MORE than trend after closing (all / plausible-only)"); put(kc, "B41", 0.40, BLUE, '0%'); put(kc, "C41", 0.41, BLUE, '0%')
put(kc, "A42", "Analogous-subset RAW median, published for transparency"); put(kc, "B42", 8440, BLUE, CUR)
put(kc, "A43", "  why $8,440 is an artifact, not a savings figure: dividing whole-district budget noise by 60-320 student denominators explodes per-child values; 11 of 27 such events exceed the $13K physical ceiling in one direction or the other. Same events, plausible window only, median:", NOTE, wrap=True); put(kc, "B43", 541, BLUE, CUR)
put(kc, "A44", "BRACKET CHECK (v3.9): the record's plausible median vs this model's independent bottom-up median per displaced student. These bracket rather than coincide and measure different things: the record credits a district's whole budget change to its closure (an upper bound by construction), the model prices only the levers a closure moves. Before the v3.9 fixed-cost rebuild the model read $713 and the two nearly coincided. Both remain far below the plan's $6,250-$7,813 requirement."); put(kc, "B44", 818, BLUE, CUR); put(kc, "C44", "=Closure_Model!B47/Assumptions!B11", BLK, CUR, bold=True)

put(kc, "A46", "OUTCOMES, WITHIN ONE TESTING SYSTEM (federal proficiency series, closures 2012-2016, change vs state 3 years out)", SEC)
put(kc, "A47", "Events measurable / improved 3+ pts / declined 3+ pts / flat"); put(kc, "B47", 42, BLUE, NUM); put(kc, "C47", 11, BLUE, NUM); put(kc, "D47", 10, BLUE, NUM); put(kc, "E47", 21, BLUE, NUM)
put(kc, "A48", "Spearman correlation, share of district displaced vs score change"); put(kc, "B48", -0.38, BLUE, "0.00")
put(kc, "A49", "Median score change: events displacing 15%+ of district / under 15%"); put(kc, "B49", -2.0, BLUE, "0.0"); put(kc, "C49", -0.2, BLUE, "0.0")
put(kc, "A50", "Events with BOTH clear savings and clear gains"); put(kc, "B50", 1, BLUE, NUM); put(kc, "C50", "Leslie County 2013: middle school folded into an existing campus in the same community", NOTE)
put(kc, "A51", "Cases of a rural town's ELEMENTARY closed with clear savings and clear gains"); put(kc, "B51", 0, BLUE, NUM, bold=True)
put(kc, "A53", "Honest limits: closures before 2012 and after 2016 cannot be score-tested across Kentucky's assessment-system changes; whole-district mergers are unobservable afterward; and same-size Kentucky towns that KEPT schools declined in population at nearly the same median rate as towns that lost them (-4.1 vs -3.5 percent, 2000-2020), so no claim is made that closure causes population decline. The claim is narrower: the record contains no measurable precedent for the savings or the improvement this plan promises.", NOTE, wrap=True)


put(th, "A56", "FOURTEEN YEARS OF LEVIED RATES, NINE AREA DISTRICTS (KDE Local District Tax Levies files, Total Real Estate column; cross-checked against DOR rate books 2024-2025, all nine reconcile exactly; archived as build/ky_levy_history_2012_2026.csv)", SEC)
put(th, "A57", "District", BOLDW, fill=HDR); put(th, "B57", "2012-13", BOLDW, fill=HDR); put(th, "C57", "2013-14", BOLDW, fill=HDR); put(th, "D57", "2014-15", BOLDW, fill=HDR); put(th, "E57", "2015-16", BOLDW, fill=HDR); put(th, "F57", "2016-17", BOLDW, fill=HDR); put(th, "G57", "2017-18", BOLDW, fill=HDR); put(th, "H57", "2018-19", BOLDW, fill=HDR); put(th, "I57", "2019-20", BOLDW, fill=HDR); put(th, "J57", "2020-21", BOLDW, fill=HDR); put(th, "K57", "2021-22", BOLDW, fill=HDR); put(th, "L57", "2022-23", BOLDW, fill=HDR); put(th, "M57", "2023-24", BOLDW, fill=HDR); put(th, "N57", "2024-25", BOLDW, fill=HDR); put(th, "O57", "2025-26", BOLDW, fill=HDR); put(th, "P57", "Change", BOLDW, fill=HDR)
put(th, "A58", "Bath"); put(th, "B58", 36.8, BLUE, "0.0"); put(th, "C58", 44.0, BLUE, "0.0"); put(th, "D58", 44.2, BLUE, "0.0"); put(th, "E58", 44.8, BLUE, "0.0"); put(th, "F58", 44.8, BLUE, "0.0"); put(th, "G58", 47.2, BLUE, "0.0"); put(th, "H58", 52.6, BLUE, "0.0"); put(th, "I58", 52.6, BLUE, "0.0"); put(th, "J58", 52.6, BLUE, "0.0"); put(th, "K58", 52.4, BLUE, "0.0"); put(th, "L58", 54.1, BLUE, "0.0"); put(th, "M58", 57.8, BLUE, "0.0"); put(th, "N58", 60.7, BLUE, "0.0"); put(th, "O58", 63.4, BLUE, "0.0"); put(th, "P58", "=(O58/B58-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A59", "Scott"); put(th, "B59", 45.3, BLUE, "0.0"); put(th, "C59", 47.2, BLUE, "0.0"); put(th, "D59", 47.6, BLUE, "0.0"); put(th, "E59", 49.0, BLUE, "0.0"); put(th, "F59", 49.6, BLUE, "0.0"); put(th, "G59", 56.4, BLUE, "0.0"); put(th, "H59", 56.4, BLUE, "0.0"); put(th, "I59", 57.1, BLUE, "0.0"); put(th, "J59", 57.6, BLUE, "0.0"); put(th, "K59", 58.1, BLUE, "0.0"); put(th, "L59", 58.1, BLUE, "0.0"); put(th, "M59", 62.8, BLUE, "0.0"); put(th, "N59", 62.9, BLUE, "0.0"); put(th, "O59", 62.9, BLUE, "0.0"); put(th, "P59", "=(O59/B59-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A60", "Harrison"); put(th, "B60", 43.0, BLUE, "0.0"); put(th, "C60", 45.2, BLUE, "0.0"); put(th, "D60", 47.2, BLUE, "0.0"); put(th, "E60", 47.3, BLUE, "0.0"); put(th, "F60", 49.0, BLUE, "0.0"); put(th, "G60", 50.5, BLUE, "0.0"); put(th, "H60", 50.5, BLUE, "0.0"); put(th, "I60", 51.7, BLUE, "0.0"); put(th, "J60", 51.7, BLUE, "0.0"); put(th, "K60", 57.7, BLUE, "0.0"); put(th, "L60", 57.7, BLUE, "0.0"); put(th, "M60", 57.7, BLUE, "0.0"); put(th, "N60", 57.7, BLUE, "0.0"); put(th, "O60", 57.7, BLUE, "0.0"); put(th, "P60", "=(O60/B60-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A61", "Clark"); put(th, "B61", 53.6, BLUE, "0.0"); put(th, "C61", 55.9, BLUE, "0.0"); put(th, "D61", 57.4, BLUE, "0.0"); put(th, "E61", 60.0, BLUE, "0.0"); put(th, "F61", 62.2, BLUE, "0.0"); put(th, "G61", 62.2, BLUE, "0.0"); put(th, "H61", 62.2, BLUE, "0.0"); put(th, "I61", 63.7, BLUE, "0.0"); put(th, "J61", 63.7, BLUE, "0.0"); put(th, "K61", 64.9, BLUE, "0.0"); put(th, "L61", 65.9, BLUE, "0.0"); put(th, "M61", 67.5, BLUE, "0.0"); put(th, "N61", 66.8, BLUE, "0.0"); put(th, "O61", 65.5, BLUE, "0.0"); put(th, "P61", "=(O61/B61-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A62", "Fayette"); put(th, "B62", 67.4, BLUE, "0.0"); put(th, "C62", 69.6, BLUE, "0.0"); put(th, "D62", 71.9, BLUE, "0.0"); put(th, "E62", 74.0, BLUE, "0.0"); put(th, "F62", 75.0, BLUE, "0.0"); put(th, "G62", 75.0, BLUE, "0.0"); put(th, "H62", 81.0, BLUE, "0.0"); put(th, "I62", 81.0, BLUE, "0.0"); put(th, "J62", 81.0, BLUE, "0.0"); put(th, "K62", 80.8, BLUE, "0.0"); put(th, "L62", 83.3, BLUE, "0.0"); put(th, "M62", 81.0, BLUE, "0.0"); put(th, "N62", 80.9, BLUE, "0.0"); put(th, "O62", 79.8, BLUE, "0.0"); put(th, "P62", "=(O62/B62-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A63", "Paris Ind"); put(th, "B63", 61.0, BLUE, "0.0"); put(th, "C63", 64.1, BLUE, "0.0"); put(th, "D63", 66.5, BLUE, "0.0"); put(th, "E63", 69.0, BLUE, "0.0"); put(th, "F63", 71.6, BLUE, "0.0"); put(th, "G63", 74.4, BLUE, "0.0"); put(th, "H63", 78.3, BLUE, "0.0"); put(th, "I63", 80.4, BLUE, "0.0"); put(th, "J63", 68.7, BLUE, "0.0"); put(th, "K63", 74.5, BLUE, "0.0"); put(th, "L63", 71.5, BLUE, "0.0"); put(th, "M63", 71.5, BLUE, "0.0"); put(th, "N63", 71.5, BLUE, "0.0"); put(th, "O63", 71.5, BLUE, "0.0"); put(th, "P63", "=(O63/B63-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A64", "Nicholas"); put(th, "B64", 38.4, BLUE, "0.0"); put(th, "C64", 38.7, BLUE, "0.0"); put(th, "D64", 38.5, BLUE, "0.0"); put(th, "E64", 38.3, BLUE, "0.0"); put(th, "F64", 38.5, BLUE, "0.0"); put(th, "G64", 39.5, BLUE, "0.0"); put(th, "H64", 39.6, BLUE, "0.0"); put(th, "I64", 41.0, BLUE, "0.0"); put(th, "J64", 40.1, BLUE, "0.0"); put(th, "K64", 40.2, BLUE, "0.0"); put(th, "L64", 41.4, BLUE, "0.0"); put(th, "M64", 42.5, BLUE, "0.0"); put(th, "N64", 43.6, BLUE, "0.0"); put(th, "O64", 43.1, BLUE, "0.0"); put(th, "P64", "=(O64/B64-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A65", "Montgomery"); put(th, "B65", 49.4, BLUE, "0.0"); put(th, "C65", 49.4, BLUE, "0.0"); put(th, "D65", 49.0, BLUE, "0.0"); put(th, "E65", 48.9, BLUE, "0.0"); put(th, "F65", 50.8, BLUE, "0.0"); put(th, "G65", 51.3, BLUE, "0.0"); put(th, "H65", 52.8, BLUE, "0.0"); put(th, "I65", 52.3, BLUE, "0.0"); put(th, "J65", 52.3, BLUE, "0.0"); put(th, "K65", 52.3, BLUE, "0.0"); put(th, "L65", 52.2, BLUE, "0.0"); put(th, "M65", 52.4, BLUE, "0.0"); put(th, "N65", 52.5, BLUE, "0.0"); put(th, "O65", 52.5, BLUE, "0.0"); put(th, "P65", "=(O65/B65-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A66", "Bourbon Co"); put(th, "B66", 55.4, BLUE, "0.0"); put(th, "C66", 57.6, BLUE, "0.0"); put(th, "D66", 57.3, BLUE, "0.0"); put(th, "E66", 59.1, BLUE, "0.0"); put(th, "F66", 58.8, BLUE, "0.0"); put(th, "G66", 60.0, BLUE, "0.0"); put(th, "H66", 61.3, BLUE, "0.0"); put(th, "I66", 60.6, BLUE, "0.0"); put(th, "J66", 55.9, BLUE, "0.0"); put(th, "K66", 54.2, BLUE, "0.0"); put(th, "L66", 49.2, BLUE, "0.0"); put(th, "M66", 52.4, BLUE, "0.0"); put(th, "N66", 52.4, BLUE, "0.0"); put(th, "O66", 52.4, BLUE, "0.0"); put(th, "P66", "=(O66/B66-1)", BLK, '+0.0%;-0.0%', bold=True)
put(th, "A68", "Bourbon County is the only district in the region whose levied rate is lower today than in tax year 2012 (55.4 down to 52.4, minus 5.4 percent); Bath rose 72.3 percent, Scott 38.9, Harrison 34.2, Clark 22.2, Fayette 18.4, Paris Independent 17.2. Honesty notes: these are levied RATES, not revenue effort; under HB 44 the compensating rate falls as assessments grow, so the chart nets each board's levy choices against its assessment growth. Bourbon took the 4 percent option in 5 of the last 12 years (KDE levied-type file); the rate fell anyway because the other years took compensating or less, including the 12-cent slide of 2018-2022. Neighbors that rose took the 4 percent 7 to 9 times over the same window.", NOTE, wrap=True)

put(th, "A70", "BEYOND THE 4 PERCENT: THE RECALLABLE LEVY OPTIONS (KRS 160.470; backs the levy card and Section 9)", SEC)
put(th, "A71", "General Fund share of the levied 52.4 cents (KDE levied-rates file: 41.0 GF + 5.7 FSPK + 5.7 recallable)"); put(th, "B71", 41.0, BLUE, "0.0")
put(th, "A72", "Yield per cent of REAL ESTATE rate, per year (certified real base / 10,000)"); put(th, "B72", "=1661885191/10000", BLK, CUR)
put(th, "G72", "CORRECTED in v4.6: the FY2025 audit's certified valuation of $1,843,569,625 splits, at its own calculated levy of $9,880,143 with rates 52.4 real / 64.5 tangible, into $1,661,885,191 real and $181,684,434 tangible; one real cent yields $166,189 at full collection. The earlier blended figure ($7,829,060 collections / 41.0 cents = $190,953) mixed tangible taxed at 64.5 cents into a real-rate move and overstated the options below by about 13 percent. Tangible (already 64.5, above every option) and motor vehicle rates do not move", NOTE, wrap=True)
put(th, "A73", "Median owner-occupied home value, Bourbon County (Census ACS 2019-2023, table B25077)"); put(th, "B73", 211600, BLUE, CUR)
put(th, "A74", "Cost per added cent to that household, per year"); put(th, "B74", "=B73*0.01/100", BLK, '"$"#,##0.00')
put(th, "G74", "$21.16 a year, $1.76 a month, per added cent. Homestead exemption shields about $46,000 of a senior homeowner's value; farmland is assessed at agricultural use value, not market; renters pay only what landlords pass through", NOTE, wrap=True)
put(th, "A76", "Option", BOLDW, fill=HDR); put(th, "B76", "New rate", BOLDW, fill=HDR); put(th, "C76", "Added cents", BOLDW, fill=HDR); put(th, "D76", "Revenue per year", BOLDW, fill=HDR); put(th, "E76", "Median-home cost/yr", BOLDW, fill=HDR); put(th, "F76", "Direct bond capacity", BOLDW, fill=HDR)
lv_opts = [("Match Harrison County (2025-26 levied rate)", "=B24"),
           ("Match the regional median (eight area districts, Fayette excluded)", "=MEDIAN(O58,O59,O60,O61,O63,O64,O65,O66)"),
           ("Restore Bourbon's own 2018 rate", "=B5"),
           ("Match Clark County (2025-26 levied rate)", "=O61")]
lvr = 77
for lv_label, lv_rate in lv_opts:
    put(th, f"A{lvr}", lv_label)
    put(th, f"B{lvr}", lv_rate, BLK, "0.0")
    put(th, f"C{lvr}", f"=B{lvr}-B12", BLK, "+0.0;-0.0")
    put(th, f"D{lvr}", f"=C{lvr}*$B$72", BLK, CUR)
    put(th, f"E{lvr}", f"=C{lvr}*$B$74", BLK, CUR)
    put(th, f"F{lvr}", f"=D{lvr}*(1-(1+Debt_Service!B{rate_r})^-Debt_Service!B{term_r})/Debt_Service!B{rate_r}", BLK, CUR)
    lvr += 1
put(th, "G77", "Rates pulled live from this tab: Harrison row 24, the 2025-26 columns of the levy table, and Bourbon's own 2018 row. Direct bond capacity uses the Debt_Service tab's 4.5 percent, 20-year assumption, the same one that prices the $14 million nickel-residual bond", NOTE, wrap=True)

put(th, "A82", "THE SEQUENCE: BALANCE THE BUDGET FIRST, THEN BOND", SEC)
put(th, "A83", "First call on new recurring money: operating gap (FY2026 trend) plus ending the capital-to-GF sweep")
put(th, "B83", f"=Debt_Service!C{gap_r}+Debt_Service!B{swp_r}", BLK, CUR)
put(th, "G83", "$373,989 gap plus the $1,320,939 sweep = $1,694,928. Ending the sweep is what frees the building-fund residual to service new bonds", NOTE, wrap=True)
put(th, "A84", "Revenue from restoring the 2018 rate"); put(th, "B84", "=D79", BLK, CUR)
put(th, "A85", "Margin after the gap is closed and the sweep ended"); put(th, "B85", "=B84-B83", BLK, CUR, bold=True)
put(th, "G85", "Negative about $216,000: the restore covers the operating close and most of the sweep; the remainder sits well inside the cost package's $760,000 floor. An earlier version showed +$4,551 on the blended per-cent basis, corrected in v4.6", NOTE, wrap=True)
put(th, "A86", "Capacity unlocked once the sweep ends: nickel-residual bond"); put(th, "B86", f"=Debt_Service!B{bfres_r+1}", BLK, CUR)
put(th, "A87", "  plus remaining restricted capacity (FY2024 audit less the 2024 issue)"); put(th, "B87", f"=Debt_Service!B{rrc_r}", BLK, CUR)
put(th, "A88", "  plus bonds the new nickel equalization supports"); put(th, "B88", f"=Debt_Service!B{nick_r+1}", BLK, CUR)
put(th, "A89", "Construction capacity unlocked without pledging a cent of the new levy"); put(th, "B89", "=B86+B87+B88", BLK, CUR, bold=True)
put(th, "G89", "Roughly $35 million, for $15.69 a month on the median home, with nothing closed. The Harrison and median options are honest partial steps, leaving about $814,000 and $382,000 a year to find from the alternatives menu before the sweep can end; the Clark option adds about $6.3 million more direct capacity from its surplus", NOTE, wrap=True)
put(th, "A93", "HOUSE BILL 44: THE 4 PERCENT IS A LIMIT ON REVENUE, NOT ON THE RATE (backs Section 11 and Question 2)", SEC)
put(th, "A94", "Certified real and personal property assessment, FY2025 (audit)"); put(th, "B94", 1843569625, BLUE, CUR)
put(th, "A95", "General Fund rate levied, cents per $100"); put(th, "B95", "=B71", BLK, "0.00")
put(th, "A96", "General Fund revenue at that rate on that assessment"); put(th, "B96", "=B94*B95/100/100", BLK, CUR)
put(th, "G96", "$7,558,635 on the certified assessment. Actual FY2025 GF tax collections were $7,829,060; the gap is motor vehicle and other classes plus collection timing, and it does not move any percentage below", NOTE, wrap=True)
put(th, "A98", "Assessment growth", BOLDW, fill=HDR); put(th, "B98", "Compensating rate", BOLDW, fill=HDR)
put(th, "C98", "4 percent option", BOLDW, fill=HDR); put(th, "D98", "Rate raised 4 percent", BOLDW, fill=HDR)
put(th, "E98", "Revenue at 4 percent option", BOLDW, fill=HDR); put(th, "F98", "Revenue at rate x 1.04", BOLDW, fill=HDR)
hb = 99
for g in (0.00, 0.03, 0.05, 0.07):
    put(th, f"A{hb}", g, BLUE, "0%")
    put(th, f"B{hb}", f"=$B$95/(1+A{hb})", BLK, "0.00")
    put(th, f"C{hb}", f"=B{hb}*1.04", BLK, "0.00")
    put(th, f"D{hb}", f"=$B$95*1.04", BLK, "0.00")
    put(th, f"E{hb}", f"=$B$94*(1+A{hb})*C{hb}/100/100", BLK, CUR)
    put(th, f"F{hb}", f"=$B$94*(1+A{hb})*D{hb}/100/100", BLK, CUR)
    put(th, f"G{hb}", f"=F{hb}-E{hb}", BLK, CUR, bold=True)
    hb += 1
put(th, "G98", "Above the ceiling, recallable", BOLDW, fill=HDR)
put(th, "A104", "Read row 101, five percent growth: the largest rate the board can levy with no recall exposure is 40.61 cents, BELOW the 41.0 it levies now, and it still collects four percent more revenue. Column E is flat at every growth rate because the four percent is a revenue limit; column D is flat because it is a rate rule. The naive rate overshoots the protected rate by exactly the assessment growth rate, so the two coincide only at zero growth. New-construction revenue sits outside the cap and is additional; any rate above the compensating rate requires a public hearing; the two facilities nickels are separate from this General Fund table. KRS 132.010, KRS 132.017, KRS 160.470, read as a layperson, not as legal advice.", NOTE, wrap=True)
put(th, "A106", "What this settles, and what it does not: a flat or falling levied rate is not evidence either way about whether the four percent option was taken, because taking it in a county with rising assessments produces a falling rate. KDE's levied-type file shows Bourbon took the option in five of the last twelve years against seven, eight and nine for the neighbors that rose. The document that closes the question is the certified compensating rate against the rate actually levied, each year, and it is on the records list.", NOTE, wrap=True)
put(th, "A108", "THE 2023 RATE MOVEMENT: 3.2 CENTS NET, 5.7 CENTS RESTRICTED IN", SEC)
put(th, "A109", "Total levied rate, tax year 2022"); put(th, "B109", 49.2, BLUE, "0.0")
put(th, "A110", "Total levied rate, tax year 2023 (held at 52.4 for three years since)"); put(th, "B110", 52.4, BLUE, "0.0")
put(th, "A111", "Net change"); put(th, "B111", "=B110-B109", BLK, "+0.0;-0.0")
put(th, "A112", "Recallable facilities nickel levied 8/17/2023 (KDE levied-rates file, current year)"); put(th, "B112", 5.7, BLUE, "0.0")
put(th, "A113", "Implied change in the rest of the rate"); put(th, "B113", "=B111-B112", BLK, "+0.0;-0.0", bold=True)
put(th, "A114", "Implied annual change in unrestricted revenue"); put(th, "B114", "=B113*B72", BLK, CUR, bold=True)
put(th, "G114", "About -2.5 cents, roughly -$477,000 a year. INFERENCE, NOT A DOCUMENT: KDE does not publish the year-by-year rate-type split, and a facilities nickel is an equivalent rate restated annually against the whole property base, so it drifts. The certified split by year is requested in Appendix B. What is not in doubt: the levy that rose in 2023 was restricted building money that cannot lawfully pay a teacher", NOTE, wrap=True)

put(th, "A91", "None of these is a recommendation of a particular number, and none is counted in the alternatives package total. The point is narrower: a menu of options exists between cut nothing and close a school, every one prices out larger than the most generous closure estimate, and every one carries a built-in democratic check. A levy above 4 percent can be recalled by the voters it taxes. A closed school cannot be recalled by the children it displaces. This community was offered that veto twice on the facilities nickels and twice declined to use it; it has never been offered the same vote on the operating levy that pays teachers.", NOTE, wrap=True)

# ================= SCHOOL COSTS AND BREAKEVENS (v3.8) =================
sc = sheet("School_Costs", [46, 13, 13, 13, 13, 13, 52])
put(sc, "A1", "School-Level Costs Across Every Reporting System, and Every Breakeven Construction", TITLE)
put(sc, "A2", "Backs the two v3.8 site cards and the Section 4 additions. Three reporting systems with different definitions; compare within a system and a year, never across.", NOTE)

put(sc, "A4", "DISTRICT REVENUE BY SOURCE, PER MEMBER (KDE funding files; archived build/bourbon_revenue_by_source_2020_2024.csv)", SEC)
put(sc, "A5", "Year", BOLDW, fill=HDR); put(sc, "B5", "Members", BOLDW, fill=HDR); put(sc, "C5", "Local", BOLDW, fill=HDR); put(sc, "D5", "State", BOLDW, fill=HDR); put(sc, "E5", "Federal", BOLDW, fill=HDR); put(sc, "F5", "Total/member", BOLDW, fill=HDR)
rev_rows = [("2019-20", 2620, 9427594, 18087210, 5573373), ("2020-21", 2561, 9934020, 17692714, 7900643),
            ("2021-22", 2483, 10662819, 18079609, 9985007), ("2022-23", 2454, 11808998, 20381341, 9550068),
            ("2023-24", 2406, 13172314, 17149768, 8132977)]
for i, (yy, mm, lo, st, fe) in enumerate(rev_rows):
    rr = 6 + i
    put(sc, f"A{rr}", yy); put(sc, f"B{rr}", mm, BLUE, NUM); put(sc, f"C{rr}", lo, BLUE, CUR); put(sc, f"D{rr}", st, BLUE, CUR); put(sc, f"E{rr}", fe, BLUE, CUR)
    put(sc, f"F{rr}", f"=(C{rr}+D{rr}+E{rr})/B{rr}", BLK, CUR)
put(sc, "G6", "State column includes on-behalf pension and insurance payments that never pass through the district's accounts; net SEEK cash is $7.8M-$9.4M of it. Federal 2020-21 to 2022-23 carries ESSER.", NOTE, wrap=True)

put(sc, "A13", "THE 300-STUDENT BREAKEVEN, RECONSTRUCTED (cited at the committee meeting; no worksheet published)", SEC)
put(sc, "A14", "Cost side: NMES 2023-24 report-card total per student"); put(sc, "B14", 19348, BLUE, CUR); put(sc, "G14", "KYRC24_FT_Spending_per_Student.csv; includes federal spending and embedded on-behalf", NOTE, wrap=True)
put(sc, "A15", "Times 128 students"); put(sc, "B15", "=B14*128", BLK, CUR, bold=True); put(sc, "G15", "Equals $2,476,544 exactly, zero-dollar difference from the figure the 300 implies", NOTE, wrap=True)
put(sc, "A16", "Implied revenue denominator for a 300 breakeven"); put(sc, "B16", "=B15/300", BLK, CUR)
put(sc, "A17", "State-only revenue per member, 2022-23"); put(sc, "B17", "=D9/B9", BLK, CUR); put(sc, "G17", "$8,305: the only revenue definition in the state's files that lands near the implied $8,255", NOTE, wrap=True)
put(sc, "A18", "Breakeven N at state-only revenue"); put(sc, "B18", "=B15/B17", BLK, '0.0', bold=True); put(sc, "G18", "298 students. Per SEEK AADA (2,490) instead of members: 303. The 300 sits between the two bases.", NOTE, wrap=True)
put(sc, "A19", "Same fraction with matching definitions: all-in cost / total revenue per member (2023-24)"); put(sc, "B19", "=B15/F10", BLK, '0.0')
put(sc, "G19", "155 students, not 300. The 300 construction counts local and federal dollars on the cost side and skips them on the revenue side; those two sources bring $8,855 per student, more than the state share.", NOTE, wrap=True)

put(sc, "A21", "THE CORRECTED TEST, APPLIED TO EVERY SCHOOL (all-in cost per pupil / total revenue per member, 2023-24)", SEC)
put(sc, "A22", "School", BOLDW, fill=HDR); put(sc, "B22", "Cost/pupil", BOLDW, fill=HDR); put(sc, "C22", "Enrolled", BOLDW, fill=HDR); put(sc, "D22", "Breakeven N", BOLDW, fill=HDR); put(sc, "E22", "Short by", BOLDW, fill=HDR)
bek = [("North Middletown", 19348, 128), ("Bourbon Central", 18131, 459), ("Cane Ridge", 18670, 453),
       ("Bourbon Co Middle", 16673, 590), ("Bourbon Co High", 17404, 766)]
for i, (nm, pp, en) in enumerate(bek):
    rr = 23 + i
    put(sc, f"A{rr}", nm); put(sc, f"B{rr}", pp, BLUE, CUR); put(sc, f"C{rr}", en, BLUE, NUM)
    put(sc, f"D{rr}", f"=B{rr}*C{rr}/F$10", BLK, '0'); put(sc, f"E{rr}", f"=D{rr}-C{rr}", BLK, '0')
put(sc, "A28", "Every school in the district fails, including both receiving schools; Cane Ridge falls short by 76 students and about $1.2 million on this test, nearly three times NMES. Statewide, 1,146 of 1,151 A1 schools with reported data (99.6 percent) spend more per student all-in than $8,255. An average-cost breakeven in a drawdown year is a district-budget thermometer, not a school test.", NOTE, wrap=True)

put(sc, "A30", "THE REAL BREAKEVEN: FIXED SITE BASE VS THE FUNDING A STUDENT CARRIES", SEC)
put(sc, "A31", "School", BOLDW, fill=HDR); put(sc, "B31", "Fixed base", BOLDW, fill=HDR); put(sc, "C31", "N (SEEK only)", BOLDW, fill=HDR); put(sc, "D31", "N if fed counted", BOLDW, fill=HDR); put(sc, "E31", "Enrolled", BOLDW, fill=HDR)
put(sc, "A32", "North Middletown"); put(sc, "B32", "=Assumptions!B51+Assumptions!B52", GRN, CUR)
put(sc, "A33", "Cane Ridge"); put(sc, "B33", "=Assumptions!B51+Assumptions!B52*Facility_Plans!F68/Facility_Plans!F69", BLK, CUR)
put(sc, "A34", "Bourbon Central"); put(sc, "B34", "=Assumptions!B51+Assumptions!B52*Facility_Plans!F67/Facility_Plans!F69", BLK, CUR)
for rr, en in ((32, 128), (33, 453), (34, 459)):
    put(sc, f"C{rr}", f"=B{rr}/(Assumptions!B6-Assumptions!B62)", BLK, '0')
    put(sc, f"D{rr}", f"=B{rr}/(Assumptions!B6+3380-Assumptions!B62)", BLK, '0')
    put(sc, f"E{rr}", en, BLUE, NUM)
put(sc, "G32", "B51 and B52 are now the measured lines themselves and plant is scaled by KFICS replacement value for the other two schools. The district's FY2026 working budget measures North Middletown's own General Fund lines, excluding state-paid on-behalf, at school administration $132,744, plant operations $96,107 (the General Fund portion of function 2600; $119,909 across all funds) and instructional staff support, the school's library and media line, $49,097. Administration plus plant measure $227,831 on the program view used by B51 and B52, or $228,851 on the function view, against the $290,000 this model assumed for the same two components through v3.8. Local carries zero at the margin, because the levy does not change with enrollment, and federal carries zero as well, because Title I is a district allocation driven by resident poverty and reappears at whichever school the child attends. Column D keeps the federal case only to show how far the answer moves when you count it.", NOTE, wrap=True)
put(sc, "A35", "NMES clears its bar at roughly two to two and a half times its enrollment; the receiving schools at roughly 5 to 6 times. The full grid version with position, busing, and leaver effects brackets NMES at 20 to 122 (KY_Closures yardstick rows). Every construction except all-source-cost-over-state-only-revenue says every school pays its way.", NOTE, wrap=True)

put(sc, "A37", "SCHOOL-LEVEL SPENDING PER STUDENT, OLD SRC SYSTEM 2011-12 TO 2016-17 (KDE Learning Environment files; archived CSV in build/)", SEC)
put(sc, "A38", "Year", BOLDW, fill=HDR); put(sc, "B38", "NMES", BOLDW, fill=HDR); put(sc, "C38", "NMES members", BOLDW, fill=HDR); put(sc, "D38", "BCES", BOLDW, fill=HDR); put(sc, "E38", "Cane Ridge", BOLDW, fill=HDR); put(sc, "F38", "NMES vs BCES", BOLDW, fill=HDR)
sric = [("2011-12", 9715, 177, 7351, 7740), ("2012-13", 19635, 158, 10281, 11383), ("2013-14", 9948, 154, 7123, 7718),
        ("2014-15", 10999, 153, 8651, 9222), ("2015-16", 11521, 149, 8705, 9655), ("2016-17", 13260, 133, 9197, 9815)]
for i, (yy, nn, mm, bb, cc) in enumerate(sric):
    rr = 39 + i
    put(sc, f"A{rr}", yy); put(sc, f"B{rr}", nn, BLUE, CUR); put(sc, f"C{rr}", mm, BLUE, NUM); put(sc, f"D{rr}", bb, BLUE, CUR); put(sc, f"E{rr}", cc, BLUE, CUR)
    put(sc, f"F{rr}", f"=B{rr}/D{rr}-1", BLK, '+0%;-0%')
put(sc, "G40", "2012-13 NMES prints $19,635, off both neighbors' trend by a factor of two in a renovation year: a capital charge booked to the site. Excluded from operating comparisons.", NOTE, wrap=True)
put(sc, "A45", "These are the deep-emptiness years (133-177 students): the premium is widest exactly when the building is emptiest. Two federal collections confirm the pattern independently: CRDC school-level salaries (archived) and the NCES school-level finance survey FY16-FY17 (archived).", NOTE, wrap=True)

put(sc, "A47", "THE 2000-01 REPORT CARDS: FOUR SCHOOLS, ONE SCALE CURVE (recovered from the Internet Archive; archive copies being added)", SEC)
put(sc, "A48", "School", BOLDW, fill=HDR); put(sc, "B48", "Students", BOLDW, fill=HDR); put(sc, "C48", "Reported $/student", BOLDW, fill=HDR); put(sc, "D48", "Curve prediction", BOLDW, fill=HDR)
c01 = [("Bourbon Central", 595, 3360), ("Cane Ridge", 312, 4053), ("North Middletown", 193, 4414), ("Millersburg", 145, 5200)]
for i, (nm, en, pp) in enumerate(c01):
    rr = 49 + i
    put(sc, f"A{rr}", nm); put(sc, f"B{rr}", en, BLUE, NUM); put(sc, f"C{rr}", pp, BLUE, CUR)
    put(sc, f"E{rr}", f"=1/B{rr}", BLK, '0.00000')
for i in range(4):
    rr = 49 + i
    put(sc, f"D{rr}", f"=B$55+B$54*E{rr}", BLK, CUR)
put(sc, "A54", "Fixed base per building (least-squares slope on 1/N)"); put(sc, "B54", "=SLOPE(C49:C52,E49:E52)", BLK, CUR, bold=True)
put(sc, "A55", "Variable cost per student (intercept)"); put(sc, "B55", "=INTERCEPT(C49:C52,E49:E52)", BLK, CUR, bold=True)
put(sc, "A56", "Fit quality (R-squared)"); put(sc, "B56", "=RSQ(C49:C52,E49:E52)", BLK, '0.000')
put(sc, "G49", "One formula, about $2,851 per student plus about $332,000 per building over the zone-assigned enrollment, predicts all four schools within 4 percent. No school on this table was mismanaged; the ranking is enrollment. Cane Ridge itself cost 21 percent more than Bourbon Central at 312 students. CATS-era cards may report prior-year spending; the within-year comparison is unaffected.", NOTE, wrap=True)

put(sc, "A58", "THE MILLERSBURG SYMMETRY: THE DISTRICT HAS RUN THIS EXPERIMENT", SEC)
put(sc, "A59", "Millersburg premium over its receiving school, 2000-01 (per student x students)"); put(sc, "B59", "=(C52-C50)*B52", BLK, CUR, bold=True)
put(sc, "G59", "$166,315 in FY2001 dollars. Closed 2006; the 30-year record (KY_Closures tab) shows no measurable district budget bend afterward.", NOTE, wrap=True)
put(sc, "A60", "NMES premium over the cheaper receiving school, 2023-24 (same computation)"); put(sc, "B60", "=(B14-18131)*128", BLK, CUR, bold=True)
put(sc, "G60", "$155,776. The same experiment at almost the same number, twenty-three years apart.", NOTE, wrap=True)

put(sc, "A62", "STAFFING RECORD (federal CCD, 1996-2019; archived build/bourbon_staffing_ratios_ccd.csv)", SEC)
put(sc, "A63", "NMES students per teacher at its 2007-09 enrollment peak (224 and 217 students): 16.0 and 15.5, vs 16.1-16.3 at Bourbon Central and Cane Ridge: parity. The ratio falls to 11.8 by 2012-13 as enrollment falls; the dollar premium peaks in the same years. 1998-2000 CCD teacher counts are corrupt for small Kentucky schools and excluded; single-year integer FTE wobbles about one student per teacher.", NOTE, wrap=True)

# ---- finish ----
del wb["Sheet"]
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = True
wb.save("/home/claude/nmes/NMES_Financial_Model.xlsx")
print("model written")
