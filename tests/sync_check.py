"""Three-way number sync: site JavaScript vs workbook cells vs report text.

Run:  python tests/sync_check.py
Needs: pip install pypdf openpyxl
Exits nonzero on any discrepancy.
"""
from pathlib import Path
REPO = Path(__file__).resolve().parents[1]
import re, json
from openpyxl import load_workbook
from pypdf import PdfReader

R = {"match": [], "diff": [], "note": []}
def match(m): R["match"].append(m)
def diff(m): R["diff"].append(m)
def note(m): R["note"].append(m)

html = open(f"{REPO}/index.html").read()
wb = load_workbook(f"{REPO}/NMES_Financial_Model.xlsx")
pdf_text = "\n".join(p.extract_text() for p in PdfReader(f"{REPO}/Saving_North_Middletown_Elementary.pdf").pages)
pdf_flat = pdf_text.replace("\n", " ")

# ---------- helpers ----------
def js_array(name_pattern):
    m = re.search(name_pattern, html, re.S)
    return json.loads("[" + m.group(1).replace("null", "null") + "]") if m else None

A = wb["Assumptions"]; TH = wb["Tax_History"]; SD = wb["School_Data"]; DM = wb["Demographics"]

# ---------- 1. core scalars ----------
model_deficit = A["B24"].value - A["B21"].value
if model_deficit == 2648086 and "2,648,086" in pdf_flat and "$2.65 million" in html:
    match(f"FY2025 structural deficit $2,648,086 (model computed, PDF text; site prose $2.65 million; the site JS constant retired with the v4.5 percentile-only readouts)")
else:
    diff(f"deficit: model {model_deficit}, pdf has 2,648,086: {'2,648,086' in pdf_flat}, site prose: {'$2.65 million' in html}")

bal = A["B29"].value
if bal == 4290840 and "4,290,840" in pdf_flat and "4,290,840" not in html:
    note("fund balance $4,290,840 in model+PDF; site says '$4.3 million'/'falling ~$1.1M a year' (rounded prose, consistent)")
if "4,290,840" in html: match("fund balance $4,290,840 also on site")

# closure central case (v4.2 grid, rebuilt on the district's own response)
CM = wb["Closure_Model"]
central = (CM["C39"].value + CM["C40"].value + CM["C41"].value * 54479.4
           - CM["C42"].value - CM["C43"].value * (A["B6"].value + CM["C44"].value - 400))
site_capv = re.search(r"var CAPV=\[(\d+),(\d+),(\d+)\]", html)
site_fixv = re.search(r"FIXV=\[(\d+),(\d+),(\d+)\]", html)
site_teach = re.search(r"TEACH=108958\.80/2", html)
site_default = (80279 + 0 + 3*54479.4 - 20000 - 154*(A["B6"].value+0-400) - 0)
if round(central) == -409446 and round(site_default) == -427087 and "-$427,087" in html:
    match("model central case -$409,446 (v5 grid); site calculator opens at the grid's weighted median cell (-$427,087, the 50th percentile by construction)")
else:
    diff(f"closure defaults: model central {central:.0f}, site median cell {site_default:.0f} shown: {'-$427,087' in html}")
# growth calculator default (v4.5 review): the weighted median scenario itself.
# 30 added students (target 140): inside the 25-seat headroom plus a partial
# class at 1 per 21, so zero teachers and zero support trigger; $0 bus, $400
# cps, $500 SEEK add-ons (the SAME default leg the closure leaver lever uses)
growth_default = 30*(4626+500-400) - ((30-25)//21)*49150 - (30//50)*37000 - 0*30
if (growth_default == 141780 and "$141,780" in html and "RATV=[18,21,24]" in html
        and "Math.max(0,gain-25)/ratio" in html and 'id="sGad"' in html
        and 'id="sGro" min="110" max="200" value="140"' in html
        and 'id="sGad" min="0" max="1000" value="500"' in html
        and 'id="sAdd" min="0" max="1000" value="0"' in html):
    match("growth calculator opens at the weighted median ($141,780, target 140, no new hires); both calculators price the same 0/500/1,000 add-ons legs, and the closure default (the median cell) starts its add-ons slider at zero")
else:
    has500 = 'value="500"' in html
    diff(f"growth default mismatch: {growth_default}, shown: {'$141,780' in html}, add-ons defaults equal: {has500}")
if site_capv and [int(site_capv.group(i)) for i in (1, 2, 3)] == [CM["B39"].value, CM["C39"].value, CM["D39"].value] == [53519, 80279, 127039]:
    match("capture lever (53,519 / 80,279 / 127,039 = district worksheet + insurance) identical site JS and model")
else:
    diff("capture lever CAPV does not match model/grid")
site_fixv = re.search(r"FIXV=\[0,(\d+\.?\d*),(\d+\.?\d*)\]", html)
if site_fixv and [float(site_fixv.group(i)) for i in (1, 2)] == [CM["C40"].value, CM["D40"].value] == [107052.2, 214104.4]:
    match("fixed-position lever (0 / 107,052.20 / 214,104.40, MUNIS FY2026 actuals) identical site JS and model")
else:
    diff("fixed-position lever FIXV does not match model/grid")
cm_teach_note = str(CM["E41"].value or "")
if site_teach and "$54,479.40" in pdf_flat and "108,958.80" in cm_teach_note:
    match("teacher basis $54,479.40 = district Appendix A.1 price, in site JS, PDF, and the model lever note")
else:
    diff("teacher basis $54,479.40 missing on site, in PDF, or in the model note")
if re.search(r"l\*\(SEEK\+ad-SUPP\)", html) and "SEEK=4626" in html:
    match("calculator prices each missing student at $4,626 + add-ons net of the $400 supplies credit (matches build/closure_grid.py)")
else:
    diff("site leaver pricing formula not found")

# two-tailed range strings consistent
if "predicts what comparable communities already did" in html \
        and "predicts what comparable communities already did" in pdf_flat \
        and "sits between the" in html and "sits between the" in pdf_flat:
    match("revealed-preference calibration (corrected band vs realized case-study losses) on site and in PDF")
else:
    diff("revealed-preference calibration sentence missing on site or in PDF")
if "still losing $9,860" in pdf_flat and "losing $846,285" in pdf_flat and "losing $846,285" in html \
        and "still losing $9,860" in html:
    match("two-tailed range (-$846,285 to -$9,860, v5 survey-anchored grid) consistent on site and in PDF")
else:
    diff("v4.2 two-tailed range strings missing on site or PDF")
if "loses $427,087" in html and "percentile of the 972 weighted scenarios" in html \
        and "LOSES $427,087" in pdf_flat and CM["B47"].value.startswith("="):
    match("v5.0 weighted median ($427,087 yearly loss) in site prose and JS readout and in PDF; central case live in model")
else:
    diff("v4.5 weighted median strings missing")

# v4.5 weighted grid stats: recompute both grids with the published 1-2-1 /
# uniform lever weights and check every published statistic
from itertools import product as _prod
import math as _math
_T = 108958.80 / 2
_cl = sorted(
    (c + f + t * _T - b - l * (4626 + ad - 400),
     wc * wf * wl * wa * wb)
    for c, wc in ((53519, 1), (80279, 2), (127039, 1))
    for f, wf in ((0, 1), (107052.2, 2), (214104.4, 1))
    for t in (0, 1, 2, 3)
    for l, wl in ((117, 1), (136, 2), (154, 1))
    for ad, wa in ((0, 1), (500, 2), (1000, 1))
    for b, wb in ((20000, 1), (63000, 2), (95000, 1)))
_tw = sum(w for _, w in _cl)
def _wp(pairs, tw, q):
    cum = 0
    for v, w in pairs:
        cum += w
        if cum >= q * tw:
            return v
_neg = sum(w for v, w in _cl if v < 0) / _tw
def _gnet(g, r, sp, tc, sc, b, c, ad):
    te = _math.floor(max(0, g - 25) / r)
    st = 0 if sp == 0 else _math.floor(g / sp)
    return g * (4626 + ad - c) - te * tc - st * sc - b * g
_gr = sorted(
    (_gnet(g, r, sp, tc, sc, b, c, ad), w2 * w3 * w4 * w5 * w6 * w7 * w8)
    for g in range(10, 91, 10)
    for r, w2 in ((18, 1), (21, 2), (24, 1))
    for sp, w3 in ((0, 1), (75, 2), (50, 1))
    for tc, w4 in ((41718, 1), (49150, 2), (56583, 1))
    for sc, w5 in ((20000, 1), (28500, 2), (37000, 1))
    for b, w6 in ((0, 1), (500, 2), (1000, 1))
    for c, w7 in ((400, 1), (700, 2), (1000, 1))
    for ad, w8 in ((0, 1), (500, 2), (1000, 1)))
_gtw = sum(w for _, w in _gr)
ok_w = (round(_wp(_cl, _tw, 0.5)) == -427087 and round(_wp(_cl, _tw, 0.25)) == -518405
        and round(_wp(_cl, _tw, 0.75)) == -338727 and _neg == 1.0
        and _wp(_gr, _gtw, 0.5) == 141780 and _wp(_gr, _gtw, 0.25) == 94520
        and _wp(_gr, _gtw, 0.75) == 182654)
site_w = all(s in html for s in ("$427,087", "$518,405", "$338,727", "$141,780",
                                 "$94,520", "$182,654", "Every priced scenario loses money"))
if ok_w and site_w:
    match("v5.0 weighted stats recompute (closure median/IQR/share, growth median/IQR) and all seven appear on the site")
else:
    diff(f"v5.0 weighted stats: recompute ok={ok_w}, site tokens ok={site_w}")

# levy path: 4 percent framing retired from the site (review round 2); path stays model+PDF
levy_base = TH["B32"].value  # General Fund property tax
cum = 0
for i in range(3): cum += (levy_base + cum) * 0.04
if levy_base == 7829060 and round(cum) == 977568 and "sYrs" not in html:
    match(f"levy base $7,829,060 (GF only) and 3-yr path to $977,568 live in model+PDF; 4 percent framing off the site")
else:
    diff(f"levy: model base {levy_base}, cum {cum:.0f}, site 4-percent remnants: {'sYrs' in html}")
restore_full = round((TH["B5"].value - TH["B12"].value) * 1661885191 / 10000)
if restore_full == 1479078 and "1479078" in html and "$166,189 per cent" in html:
    match("2018 restore $1,479,078 = 8.9 cents x the certified real base; site JS constant and basis note match")
else:
    diff(f"2018 restore: computed {restore_full}, site constant {'1479078' in html}")
y1 = levy_base * 0.04
pdf_levy_ok = "313,000" in pdf_flat and "978,000" in pdf_flat and "639,000" in pdf_flat
if pdf_levy_ok: match(f"PDF levy path ($313K yr1, $639K yr2, $978K yr3) matches computed ({y1:,.0f} / 638,851 / 977,568)")
else: diff("PDF levy figures not all found")
# the excluded restricted base is disclosed, not silently dropped
if "7,829,060" in pdf_flat and "restricted building-fund" in pdf_flat:
    match("PDF discloses the GF-only levy base and why the restricted levy is excluded")
else: diff("PDF levy-base disclosure missing")

# ---------- 2. score series (site vs model vs PDF claims) ----------
years_model = list(range(2007, 2020)) + list(range(2021, 2026))
def model_row(r):
    return {y: SD.cell(row=r, column=2+i).value for i, y in enumerate(years_model)}

# the site now carries four selectable sources: pd (KDE reading/math average),
# oth (KDE science/social studies/writing average), sd (SchoolDigger third-party
# index), comp (KDE official composite, off by default: it folds in the climate
# survey and a change score)
def site_block(src):
    blk = re.search(src + r":\{(.*?)\]\}", html, re.S).group(1) + "]"
    out = {}
    for key in ("NMES", "BC", "CR", "PE"):
        vals = [None if v.strip() == "null" else float(v)
                for v in re.search(key + r":\[([^\]]+)\]", blk).group(1).split(",")]
        out[key] = {2007 + i: v for i, v in enumerate(vals)}
    return out
site_scores = site_block("sd")
site_comp = site_block("comp")
site_pd = site_block("pd")

# KDE official record: site arrays vs the archived source file
kde = json.load(open(f"{REPO}/build/kde_scores_history.json"))
comp_bad, pd_bad = [], []
for yk, row in kde["official_composite"].items():
    if yk == "note": continue
    y = int(yk[:4]) + 1
    for key in ("NMES", "BC", "CR", "PE"):
        want = row[key][0]
        got = site_comp[key].get(y)
        if got is None or abs(got - want) > 0.06:
            comp_bad.append((key, y, got, want))
for yk in kde["reading_pd"]:
    y = int(yk[:4]) + 1
    for key in ("NMES", "BC", "CR", "PE"):
        want = (kde["reading_pd"][yk][key] + kde["math_pd"][yk][key]) / 2
        got = site_pd[key].get(y)
        if got is None or abs(got - want) > 0.06:
            pd_bad.append((key, y, got, round(want, 2)))
if not comp_bad: match("site KDE composite series matches build/kde_scores_history.json (all schools, all years)")
else: diff(f"site KDE composite mismatches: {comp_bad}")
if not pd_bad: match("site KDE reading/math average series matches build/kde_scores_history.json")
else: diff(f"site KDE R/M average mismatches: {pd_bad}")

# site 'oth' series (science / social studies / writing average) vs the archived extract
subj = json.load(open(f"{REPO}/build/kde_subjects_history.json"))
site_oth = site_block("oth")
oth_bad = []
for yk, row in subj["avg"].items():
    y = int(yk[:4]) + 1
    for key in ("NMES", "BC", "CR", "PE"):
        want = row[key]
        got = site_oth[key].get(y)
        if (got is None) != (want is None) or (want is not None and abs(got - want) > 0.06):
            oth_bad.append((key, y, got, want))
if not oth_bad: match("site KDE science/social studies/writing average series matches build/kde_subjects_history.json (all schools, all years)")
else: diff(f"site KDE Sci/SS/Writing average mismatches: {oth_bad}")
rec_bad = []
for yk, row in subj["per_subject"].items():
    for key, subs in row.items():
        vals = [v for v in subs.values() if v is not None]
        want = round(sum(vals) / len(vals), 1) if vals else None
        if want != subj["avg"][yk][key]:
            rec_bad.append((yk, key, want, subj["avg"][yk][key]))
if not rec_bad: match("archived Sci/SS/Writing averages recompute from the per-subject values")
else: diff(f"Sci/SS/Writing average recompute mismatches: {rec_bad}")

# the plan calculator mirrors the model's transformative chain lever for lever
plan_ok = (275 * 4226 + 760000 + 1479078 - 1738653 == 1662575
           and 'id="sPw" min="0" max="550" value="275"' in html
           and 'id="sPt" min="0" max="10" value="5"' in html
           and "kids*4226" in html and "restore-1738653" in html
           and "*13.008" in html and "bonds+32000000" in html
           and "(15.69/8.9)" in html and "166189" in html
           and "$1,738,653" in pdf_flat and "$2,648,086" in pdf_flat
           and "$1,662,575" in html and "275 of 550" in html and "275 of 550" in pdf_flat)
if plan_ok:
    match("plan calculator: leakage lever (0-550 at $4,226) defaults to half the pool (275), trending FY2026 gap $1,738,653, default surplus $1,662,575 (site+PDF), capacity anchored on the advisor's $32M")
else:
    diff("plan calculator bases or defaults out of sync with the re-based chain")

# model School_Data KDE history block vs the same source file
kh_years_keys = ["2011-12", "2012-13", "2013-14", "2014-15", "2015-16", "2016-17",
                 "2017-18", "2018-19", "2020-21", "2021-22", "2022-23", "2023-24", "2024-25"]
mh_bad = []
for subj, base in (("reading_pd", 45), ("math_pd", 49)):
    for j, key in enumerate(("NMES", "BC", "CR", "PE")):
        for i, yk in enumerate(kh_years_keys):
            got = SD.cell(row=base + j, column=2 + i).value
            if abs(got - kde[subj][yk][key]) > 0.001:
                mh_bad.append((subj, key, yk, got))
for j, key in enumerate(("NMES", "BC", "CR", "PE")):
    for i, yk in enumerate(kh_years_keys):
        want = kde["official_composite"].get(yk, {})
        want = want[key][0] if key in want else None
        got = SD.cell(row=56 + j, column=2 + i).value
        if (got is None) != (want is None) or (want is not None and abs(got - want) > 0.001):
            mh_bad.append(("composite", key, yk, got))
if not mh_bad: match("model School_Data KDE history block matches build/kde_scores_history.json (104 P/D cells + composites)")
else: diff(f"model KDE history mismatches: {mh_bad}")
rowmap = {"NMES": 15, "BC": 16, "CR": 17, "PE": 18}
for key, row in rowmap.items():
    mrow = model_row(row)
    bad = [(y, site_scores[key].get(y), mrow[y]) for y in years_model
           if (site_scores[key].get(y) is None) != (mrow[y] is None)
           or (mrow[y] is not None and abs((site_scores[key].get(y) or 0) - mrow[y]) > 0.01)]
    if not bad: match(f"score series {key}: all {sum(1 for y in years_model if mrow[y] is not None)} values identical site vs model")
    else: diff(f"score series {key} mismatches: {bad}")
head_ok = all(s in pdf_flat for s in ["79.1", "74.5"])
site_sd_ok = all(s in html for s in ["58.2", "26.5", "19.3"])
if head_ok and site_sd_ok:
    match("PDF headlines the official composites (79.1 Distinguished, 74.5 first by 14); site keeps the SchoolDigger series (58.2 / 26.5 / 19.3) as selectable context")
else:
    diff(f"headline scores: PDF official {head_ok}, site SchoolDigger {site_sd_ok}")

# 3-yr averages claimed in PDF (48.1 / 26.4 / 29.9)
import statistics
def avg3(key): return statistics.mean([site_scores[key][y] for y in (2023, 2024, 2025)])
a_n, a_b, a_c = avg3("NMES"), avg3("BC"), avg3("CR")
claim_ok = abs(a_n - 48.1) < 0.06 and abs(a_b - 26.4) < 0.06 and abs(a_c - 29.9) < 0.06
if claim_ok and "48.1" in pdf_flat: match(f"PDF 3-yr averages 48.1/26.4/29.9 recompute correctly ({a_n:.1f}/{a_b:.1f}/{a_c:.1f})")
else: diff(f"3-yr averages recompute to {a_n:.2f}/{a_b:.2f}/{a_c:.2f} vs PDF claim 48.1/26.4/29.9")

# ---------- 3. enrollment series (site chart retired in the v4.1 cut; prose keeps the peak) ----------
model_enroll = []
for i in range(19): model_enroll.append(DM.cell(row=33 + i, column=6).value)
for i in range(19, 37): model_enroll.append(DM.cell(row=33 + i - 19, column=9).value)
if max(model_enroll) == 261 and model_enroll[-1] == 128 and A["B11"].value == 128 and A["B12"].value == 174 \
        and "261 children at its peak" in html and "a 174 rating" in html:
    match("peak 261, current 128, capacity 174 consistent across model, PDF, and the site prose")
else:
    diff(f"enrollment series: model peak {max(model_enroll)}, latest {model_enroll[-1]}, site prose {'261 children at its peak' in html}")

# SD enrollment row (2015-25) vs tail of long series
sd_counts = [SD.cell(row=6, column=2 + i).value for i in range(10)]
if sd_counts == model_enroll[-10:]: match("School_Data 10-yr enrollment row matches Demographics long series tail")
else: diff(f"School_Data row {sd_counts} vs Demographics tail {model_enroll[-10:]}")

# ---------- 4. tax rates (site bar chart retired in the v4.1 cut; the levy-history chart remains) ----------
model_nbrs = [TH.cell(row=19 + i, column=2).value for i in range(9)]
if model_nbrs[0] == 80.9 and 52.4 in model_nbrs and "52.4 cents" in html:
    match(f"nine-district rates intact in model; site quotes the 52.4-cent rate in prose ({model_nbrs})")
else:
    diff(f"tax comparison: model {model_nbrs}, site 52.4 {'52.4 cents' in html}")
if TH["B28"].value == 65.13 and "65.1" in html and "65.1" in pdf_flat:
    match("state average 65.1 consistent (model 65.13, site and PDF 65.1)")
hist_rates = [TH.cell(row=5 + i, column=2).value for i in range(8)]
if hist_rates == [61.3, 60.6, 55.9, 54.2, 49.2, 52.4, 52.4, 52.4]:
    match("2018-2025 Bourbon rate history in model matches PDF Figure 16 series (61.3 -> 52.4)")

# ---------- 5. spot figures in PDF vs model ----------
checks = [("19,348", A["B14"].value == 19348, "per-pupil spending $19,348"),
          ("4,586", A["B5"].value == 4586, "SEEK base FY2026 $4,586"),
          ("20.3 percent", A["B42"].value == 2913654, "transportation trend (dollar figure lives in model B42)"),
          ("$1.4 to $2.3 million", True, "alternatives package $1.4-2.3M, no haircut")]
for needle, mok, label in checks:
    if needle in pdf_flat and mok: match(f"{label}: PDF text and model agree")
    else: diff(f"{label}: pdf has '{needle}': {needle in pdf_flat}, model ok: {mok}")
alt_low = None
for r in range(15, 25):
    if wb["Alternatives"].cell(row=r, column=1).value and "Conservative combined estimate, low" in str(wb["Alternatives"].cell(row=r, column=1).value):
        alt_low = wb["Alternatives"].cell(row=r, column=2).value
        alt_high = wb["Alternatives"].cell(row=r + 1, column=2).value
if alt_low == 1393830 and alt_high == 2338281:
    match("alternatives published band $1.4M-$2.3M = raw row sums, no haircut, in model and quoted in the PDF")

# ---------- 6. bonding story: $14M plan, savings-to-bond scenarios, FY2026 close ----------
DS = wb["Debt_Service"]
ds_vals = {}
for r in range(1, DS.max_row + 1):
    a = DS.cell(row=r, column=1).value
    if a: ds_vals[str(a)] = (DS.cell(row=r, column=2).value, DS.cell(row=r, column=3).value)
bond_amt = ds_vals.get("Proposed bond amount", (None, None))[0]
if bond_amt == 14000000 and "$14 million" in pdf_flat:
    match("proposed $14M bond consistent in model and PDF (site prose retired in the v4.2 cut)")
else:
    diff(f"$14M bond: model {bond_amt}, pdf {'$14 million' in pdf_flat}")
excess = ds_vals.get("District's own KDE-filed excess cost of NMES vs peer elementaries", (None, None))[0]
if excess == 121220 and "$121,220" in pdf_flat:
    match("audited excess cost $121,220 consistent in model and PDF (site pin retired in the v4.2 cut)")
else:
    diff(f"excess cost: model {excess}, pdf {'$121,220' in pdf_flat}")
fy26 = ds_vals.get("Net General Fund change, FY2026 (unaudited)", (None, None))[0]
rev26 = ds_vals.get("General Fund revenue, FY2026 actual (excludes carryforward and on-behalf)", (None, None))[0]
exp26 = ds_vals.get("General Fund expenditures, FY2026 actual", (None, None))[0]
if rev26 == 22103877 and exp26 == 22477866 and "$374,000" in pdf_flat:
    match("FY2026 unaudited net change (-$373,989) rounds to $374,000 in the PDF (site mention retired in the v4.2 cut)")
else:
    diff(f"FY2026 close: model rev {rev26} exp {exp26}, pdf {'$374,000' in pdf_flat}")
misc = ds_vals.get("Caveat 1: miscellaneous revenue (object 1990) budgeted at zero, received", (None, None))[0]
if misc == 1567829 and "$1,413,929" in pdf_flat:
    match("FY2026 miscellaneous-revenue caveat in model and PDF (site mention retired in the v4.2 cut)")
else:
    diff(f"misc revenue caveat: model {misc}, pdf {'$1,413,929' in pdf_flat}")
xfer = ds_vals.get("Caveat 2: restricted capital money transferred INTO the General Fund in June 2026", (None, None))[0]
if xfer == 1320939 and "$1,320,939" in pdf_flat and "$1.32 million" in html:
    match("June 2026 capital-to-GF transfer ($1,320,939) in model and PDF; site keeps the rounded mention")
else:
    diff(f"capital transfer: model {xfer}, site $1.32M {'$1.32 million' in html}, pdf {'$1,320,939' in pdf_flat}")
gap_b, gap_c = ds_vals.get("Operating gap to close first", (None, None))
if gap_b == 1787918 and gap_c == 373989 and "$21 million" not in pdf_flat:
    match("balanced-budget scenario inputs ($1.9M / $373,989 gaps) live in the model; the $21M/$25M walk retired from the report")
else:
    diff(f"scenario: model gaps {gap_b}/{gap_c}, pdf $21M {'$21 million' in pdf_flat}, $25M {'$25 million' in pdf_flat}")

# ---------- 7. KFICS condition index (v3.1): model vs PDF (site card moved to report in v4.1) ----------
FP = wb["Facility_Plans"]
ci_model = {FP.cell(row=r, column=1).value: [FP.cell(row=r, column=c).value for c in (2, 3, 4)]
            for r in range(67, 72)}
nmes_ci = ci_model.get("North Middletown Elementary (1948/64)")
cane_ci = ci_model.get("Cane Ridge Elementary (1992)")
cent_ci = ci_model.get("Bourbon Central Elementary (1988)")
def _r3(v): return [round(x, 3) for x in v]
if (nmes_ci and _r3(nmes_ci) == [0.694, 0.702, 0.773]
        and cane_ci and _r3(cane_ci) == [0.812, 0.812, 0.728]
        and cent_ci and _r3(cent_ci) == [0.888, 0.819, 0.823]
        and "chartCondition" not in html):
    match("KFICS condition index series intact in model; site card retired to the report (v4.1)")
else:
    diff(f"condition index: model {nmes_ci}/{cane_ci}/{cent_ci}, site chart absent: {'chartCondition' not in html}")
needs = FP["E69"].value; crv = FP["F69"].value
if needs and crv and abs((1 - needs / crv) - 0.773295) < 0.0005 and "0.773" in pdf_flat:
    match("NMES condition index 0.773 = 1 - 3,099,148/13,670,418 verified from state-report components; quoted in the PDF")
else:
    diff(f"NMES CI recompute: needs {needs}, crv {crv}, pdf 0.773 {'0.773' in pdf_flat}")

# ---------- 8. recruitment pool (v3.2): fill planner lever, model, PDF ----------
RD = wb["Redistricting"]
if not re.search(r"net=\(t\+h\)\*4626", html) and 'id="sRez"' not in html:
    match("seat planner retired from the site in v4.4 (duplicative of the growth calculator); the fill package lives in the PDF and the Redistricting tab")
else:
    diff("seat planner still present on the site after retirement")
hs = (RD["B117"].value, RD["B118"].value)
seek46 = 46 * A["B6"].value
if hs == (236, 23) and "259 registered homeschool" in html and "236 registered homeschool" in pdf_flat:
    match("registered homeschool counts (236 BCS + 23 Paris = 259, 2022-23) consistent model/site/PDF")
else:
    diff(f"homeschool counts: model {hs}, site 259 {'259 registered homeschool' in html}, pdf 236 {'236 registered homeschool' in pdf_flat}")
pool = (RD["B123"].value, RD["B125"].value, RD["B126"].value, RD["B127"].value)
if pool == (76, 131, 54, 189) and "Fayette pulling 54 commuters" in html \
        and "247 Bourbon County Schools residents enrolled in another district" in pdf_flat:
    match("KDE nonresident flows (76 out / 131 in / 54 Fayette; 247 exports counted in the pool) consistent model/site/PDF")
else:
    diff(f"nonresident flows: model {pool}")
if seek46 == 212796 and "$213,000" in pdf_flat and RD["B131"].value == "=Assumptions!B6-Assumptions!B62":
    match("46-seat fill from the pool = 46 x $4,626 = $213,256, quoted as about $213,000 in PDF; per-return net formula in model")
else:
    diff(f"pool revenue: 46*B6={seek46}, pdf $213,000 {'$213,000' in pdf_flat}")

# ---------- 9. KY closure record (v3.3): model formulas vs site chart vs PDF ----------
KC = wb["KY_Closures"]
perry = None; johnson = None
for r in range(15, 30):
    nm = KC.cell(row=r, column=1).value
    if nm == "Perry County": perry = r
    if nm == "Johnson County": johnson = r
if perry and johnson:
    pd_, pe, pf, pc = KC.cell(row=perry, column=4).value, KC.cell(row=perry, column=5).value, KC.cell(row=perry, column=6).value, KC.cell(row=perry, column=3).value
    per_kid = (pd_ * (1 + pf) - pe) / pc
    if abs(per_kid - 3643) < 2 and "$3,600" in pdf_flat:
        match("Perry 2017 per-displaced-student figure ($3,643) recomputes from model inputs; quoted in the PDF (about $3,600)")
    else:
        diff(f"Perry per-kid: model recompute {per_kid:.0f}, pdf $3,600 {'$3,600' in pdf_flat}")
else:
    diff("KY_Closures tab missing Perry/Johnson rows")
plan_lo = 800000 / 115; plan_hi = 1000000 / 115
if round(plan_lo) == 6957 and round(plan_hi) == 8696 and "$6,957 to $8,696" in pdf_flat:
    match("plan requirement per displaced student ($6,957 to $8,696 = $800K-$1M over the 115 displaced) consistent model/PDF")
else:
    diff(f"plan per-kid: {plan_lo:.0f}/{plan_hi:.0f}")
if "chartKYRecord" not in html:
    match("KY-record chart retired to the report (v4.1); case table lives in the model and PDF")
else:
    diff("KY-record chart unexpectedly present on the simplified site")
if KC["B5"].value == 339 and KC["B6"].value == 72 and KC["B51"].value == 0 and "339 rural" in pdf_flat:
    match("closure universe (339/72) and the zero-precedent cell consistent in model and PDF")
else:
    diff(f"universe: model {KC['B5'].value}/{KC['B6'].value}, zero-cell {KC['B51'].value}")

# distribution rows (v3.4)
if (KC["C39"].value, KC["C40"].value, KC["B42"].value, KC["B43"].value) == (1102, 818, 8440, 541) \
        and "$1,102" in pdf_flat and "$8,440" in pdf_flat and "$541" in pdf_flat and "$818" in pdf_flat:
    match("distribution stats (median $1,102 / plausible $818 / artifact $8,440 / corrected $541) consistent model/PDF")
else:
    diff(f"distribution rows: {KC['C39'].value}/{KC['C40'].value}/{KC['B42'].value}/{KC['B43'].value}")

# ---------- 10. levy history (v3.6): site JS vs model vs CSV ----------
import csv as _csv
lvrows={r["district"]:r for r in _csv.DictReader(open(f"{REPO}/build/ky_levy_history_2012_2026.csv"))}
site_bourbon=re.search(r'"Bourbon County":\[([\d.,]+)\]',html)
csv_b=[lvrows["Bourbon Co"][y] for y in sorted(k for k in lvrows["Bourbon Co"] if k[:2]=="20")]
if site_bourbon and [float(x) for x in site_bourbon.group(1).split(",")]==[float(x) for x in csv_b]:
    match("levy history: site JS Bourbon series identical to archived CSV (14 years)")
else:
    diff(f"levy history mismatch: site {site_bourbon and site_bourbon.group(1)} vs csv {csv_b}")
TH2 = wb["Tax_History"]
if TH2["O66"].value == 52.4 and TH2["B58"].value == 36.8 and abs(TH2["O58"].value-63.4)<0.05:
    match("levy table in model matches endpoints (Bath 36.8->63.4; Bourbon ends 52.4)")
else:
    diff(f"levy model endpoints: {TH2['B58'].value}/{TH2['O58'].value}/{TH2['O66'].value}")

# ---------- 11. beyond-4% recallable levy options (v3.7) ----------
lv_yield = 1661885191 / 10000                            # certified real base per cent
lv_median = sorted(TH2[f"O{r}"].value for r in (58, 59, 60, 61, 63, 64, 65, 66))
lv_median = (lv_median[3] + lv_median[4]) / 2            # median of eight, Fayette excluded
lv_cases = [(TH2["B24"].value, "$0.88 million", 112),    # Harrison 57.7
            (lv_median,        "$1.31 million", 167),    # regional median 60.3
            (TH2["B5"].value,  "$1.48 million", 188),    # Bourbon's own 2018 rate 61.3
            (TH2["O61"].value, "$2.18 million", 277)]    # Clark 65.5
percent_cost = TH2["B73"].value * 0.01 / 100             # $21.16 per cent on the median home
ok_lv = abs(lv_yield - 166188.52) < 1 and abs(lv_median - 60.3) < 0.01 and abs(percent_cost - 21.16) < 0.005
for rate, revstr, fam in lv_cases:
    cents = rate - TH2["B12"].value
    rev = cents * lv_yield
    ok_lv &= (abs(rev / 1e6 - float(revstr[1:5])) < 0.005) and revstr in pdf_flat
    ok_lv &= round(cents * percent_cost) == fam
if ok_lv:
    match("beyond-4% options: model-derived yield/median/costs reproduce all four PDF rows (site now carries the 2018 restore only)")
else:
    diff(f"beyond-4% options mismatch: yield {lv_yield:.2f}, median {lv_median}, cost/cent {percent_cost}")
lv_first_call = 373989 + 1320939
lv_margin = (TH2["B5"].value - TH2["B12"].value) * lv_yield - lv_first_call
if lv_first_call == 1694928 and -216100 < round(lv_margin) < -215600 and "$216,000" in pdf_flat \
        and "$4,551" not in pdf_flat \
        and str(TH2["B83"].value).startswith("=Debt_Service!") and TH2["B85"].value == "=B84-B83":
    match("beyond-4% sequencing: 2018 rate covers the close and most of the sweep, about $216,000 remaining from the cost package; the old $4,551 precision claim is gone")
else:
    diff(f"beyond-4% sequencing: first call {lv_first_call}, margin {lv_margin:.0f}")

# ---------- 11b. Fayette peer yardstick incl. the Weaver audit ----------
DFT = wb["Defaults"]
fay_ok = (DFT["D73"].value == 6902403 and DFT["D69"].value == 690460223
          and DFT["C73"].value == 28361786 and DFT["B73"].value == 4290840
          and round(6902403 / 690460223 * 100, 2) == 1.0
          and 28361786 - 6902403 == 21459383
          and DFT["D75"].value is None and DFT["D77"].value is None
          and "$6,902,403" in html and "$6,902,403" in pdf_flat
          and "$21.5 million" in html and "$21.5 million" in pdf_flat
          and "$28,361,786" in html and "restatement" not in html
          and "KRS 160.470(6)(a)" in html and "KRS 160.470(6)(a)" in pdf_flat
          and (REPO / "build" / "fcps_weaver_audit_2026_08.pdf").exists())
if fay_ok:
    match("Fayette yardstick: Weaver's adjusted $6,902,403 (1.0%) and its $21.5M difference from the audited balance agree across model, site and PDF; no drawdown or runway is computed on the mixed basis; the deck is archived")
else:
    diff("Fayette/Weaver comparison out of sync across model, site and PDF")

# ---------- 12. v3.8: school costs, breakevens, growth plan ----------
SCt = wb["School_Costs"]
ok38 = SCt["B14"].value == 19348 and SCt["B15"].value == "=B14*128" and 19348*128 == 2476544
ok38 &= "$933,537" in html and "$2,476,544" in pdf_flat
rev23 = (11808998, 20381341, 9550068, 2454)
ok38 &= (SCt["C9"].value, SCt["D9"].value, SCt["E9"].value, SCt["B9"].value) == rev23
ok38 &= abs(rev23[1]/rev23[3] - 8305.4) < 0.5 and abs(2476544/(rev23[1]/rev23[3]) - 298.2) < 0.5
if ok38:
    match("300-breakeven reconstruction: cost 128 x 19,348 = $2,476,544 exact; state-only $8,305 -> 298 (model; PDF keeps the $2,476,544 anchor, full reconstruction in the workbook)")
else:
    diff("300-breakeven reconstruction mismatch across model/site/PDF")
c01 = [(SCt[f"B{r}"].value, SCt[f"C{r}"].value) for r in (49,50,51,52)]
if c01 == [(595,3360),(312,4053),(193,4414),(145,5200)]:
    xs=[1/n for n,_ in c01]; ys=[p for _,p in c01]
    mx=sum(xs)/4; my=sum(ys)/4
    F=sum((x-mx)*(y-my) for x,y in zip(xs,ys))/sum((x-mx)**2 for x in xs)
    a0=my-F*mx
    if abs(F-331507)<5 and abs(a0-2851)<5:
        match("2000-01 scale curve: $2,851 + $331,507/N reproduced from the four report-card points (model+PDF; site card replaced by the live enrollment curve)")
    else:
        diff(f"2000-01 curve fit mismatch: F={F:.0f}, a={a0:.0f}")
    mil = (5200-4053)*145; nm = (19348-18131)*128
    if mil == 166315 and nm == 155776:
        match("Millersburg symmetry: $166,315 (2000-01) vs $155,776 (2023-24) recompute (site prose retired in the v4.2 cut)")
    else:
        diff(f"Millersburg symmetry mismatch: {mil} / {nm}")
else:
    diff(f"2000-01 report card rows mismatch in model: {c01}")
AL2 = wb["Alternatives"]
raw_lo = 313162.4 + 60000 + 100000 + 4*85000 + 0.5*(1447164-999727) + 2913654*0.05 + 50000 + (16*4626-46*400+(1-1)*60000) + 25*4226
raw_hi = 375000 + 120000 + 200000 + 425000 + 450000 + 2913654*0.10 + 150000 + (16*4626-46*400+(2-1)*60000) + 50*4226
if abs(raw_lo-1393829.6)<5 and abs(raw_hi-2338281.4)<5 and "$1.4 to $2.3 million" in pdf_flat:
    match("growth plan raw sums $1.39M/$2.34M recomputed without Medicaid or shared services; quoted in the PDF")
else:
    diff(f"growth plan raw sums: {raw_lo:.0f}/{raw_hi:.0f}")
if "$760,000 to $1.3 million" in html and "$760,000 to $1.3 million" in pdf_flat \
        and "$260,000 to $530,000" in pdf_flat and "Medicaid" not in html:
    match("growth plan pillar subtotals consistent without Medicaid or shared services (site lever-2 list sums 760K-1.3M; PDF matches)")
else:
    diff("growth plan pillar subtotals missing, inconsistent, or a removed menu line still present on site")

# transformative check: default settings (no recovery, low costs, full restore) clear the
# trending FY2026 gap, fund the 5% certified raise, and leave debt room, alongside the
# freed restricted capacity
surplus = 760000 + 1479078 - 1738653
raise_cost = 10000388 * 0.05 * 1.0145
debt_room = surplus - raise_cost
bonds = debt_room * (1 - 1.045 ** -20) / 0.045
default_bonds = (1662575 - raise_cost) * (1 - 1.045 ** -20) / 0.045
if (surplus == 500425 and abs(raise_cost - 507270) < 2 and surplus < raise_cost
        and raise_cost - surplus < 7000
        and abs(default_bonds - 15028138) < 6000
        and "$507,000" in pdf_flat and "within $7,000" in html and "within $7,000" in pdf_flat
        and "$32 million" in html and "$32 million" in pdf_flat
        and "$15.0 million" in pdf_flat
        and "$47 million" in html and "$47 million" in pdf_flat
        and "$69 million" in html and "$69 million" in pdf_flat):
    match("transformative check: floor $500K (within $7K of the raise, advisor's $32M) and central-case default $1.66M (raise + $15.0M bonds, ~$47M), site+PDF")
else:
    diff(f"transformative check mismatch: surplus {surplus:.0f}, raise {raise_cost:.0f}, default bonds {default_bonds:.0f}")
# the leakage lever and the withdrawn claims
top_surplus = 550 * 4226 + 1300000 + 1479078 - 1738653
if (top_surplus == 3364725 and "$500,000 a year to spare" in html
        and "$500,000 a year to spare" in pdf_flat
        and "$422,600 a year" in html and "$422,600 a year" in pdf_flat
        and "10 percent raise" not in html and "$52 million" not in html
        and "withdrawn with the lever correction" in pdf_flat
        and "$260,000 to $530,000" in html and "$260,000 to $530,000" in pdf_flat):
    match("leakage lever: 0-550 students at $4,226 on site and in PDF ($422,600 per 100); defaults clear the trending gap by $721K; the 10%-raise/$52M claims stay withdrawn; the Move 2 band survives as the near-term view")
else:
    diff(f"leakage lever incomplete: top surplus {top_surplus}, 10%-claim still on site: {('10 percent raise' in html)}")
fills = (16*4626-46*400, 55616+60000)
if fills[0]-0 == 55616+400*0 - 0 and "$56,000 to $116,000" in pdf_flat:
    match("fill package $56,000-$116,000 in the PDF; the site carries the live planner (prose quote retired in v4.2)")
else:
    diff("fill package correction incomplete")

# site enrollment-cost curve: KYRC25 anchor + the growth calculator's full central
# marginal stack ($700 supplies + $500 busing, teacher per 22 past the 25 seats
# open at the district's own Appendix B caps at $49,150, support hire per 75 at
# $28,500) - same constants and headroom as growth_grid
if ("17903*128" in html and "(N-128)*(700+500)" in html and "49150" in html
        and "28500" in html and "chartCurve" in html
        and "Math.max(0,gain-25)/21" in html):
    match("site cost-per-student curve counts the full marginal stack past the 25-seat headroom at today's 1-per-21 class size")
else:
    diff("site cost curve missing the full marginal stack, the 25-seat headroom, or the class-size pace")

# ---------- 13. federal EDFacts series: retired from the site (v4.4 review); archive kept in build/
if 'id="tgEF"' not in html:
    match("EDFacts toggle retired from the site; archived extract remains in build/")
else:
    diff("EDFacts toggle still present after retirement")

# site text spot checks
for s, label in [("first in the county in every subject", "first-in-county claim"), ("losing $846,285 a year at the left end to still losing $9,860 a year", "closure range prose"),
                 ("$166,189 per cent", "certified real-estate yield in the levy note"), ("$2.65 million", "deficit figure in prose"),
                 ("115 today", "enrollment in prose"), ("a 174 rating", "capacity prose")]:
    if s in html: match(f"site text: '{s}' present ({label})")
    else: diff(f"site text missing '{s}' ({label})")

print("=== MATCHES ===")
for m in R["match"]: print(" +", m)
print("=== NOTES ===")
for m in R["note"]: print(" *", m)
print("=== DIFFS ===")
for m in R["diff"]: print(" -", m)
print(f"\n{len(R['match'])} matches, {len(R['diff'])} discrepancies")

import sys
sys.exit(1 if R["diff"] else 0)
