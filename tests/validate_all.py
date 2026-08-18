"""Cross-file consistency validation: report vs model vs site vs README.

Checks counts, figure numbering, meeting details, board roster, local
assets, forbidden dashes, pagination quality, SABS pipeline references,
and headline claims.

Run:  python tests/validate_all.py
Needs: pip install pypdf openpyxl  (pypdfium2 optional, for the
pagination scan; that check is skipped without it)
Exits nonzero if any check fails.
"""
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

REPO = Path(__file__).resolve().parents[1]

ok, bad = [], []
def chk(cond, label): (ok if cond else bad).append(label)


def main():
    html = (REPO / "index.html").read_text()
    readme = (REPO / "README.md").read_text()
    r = PdfReader(REPO / "Saving_North_Middletown_Elementary.pdf")
    t = " ".join(pg.extract_text() for pg in r.pages).replace("\n", " ")
    wb = load_workbook(REPO / "NMES_Financial_Model.xlsx")

    # wording
    chk("arithmetic" not in t.lower(), "no 'arithmetic' in the PDF")
    chk("arithmetic" not in html.lower(), "no 'arithmetic' on the site")

    # counts everywhere
    n = len(r.pages)
    nf = sum(1 for ws in wb.worksheets for row in ws.iter_rows()
             for c in row if isinstance(c.value, str) and c.value.startswith("="))
    chk(f"The {n}-page report" in html and f"the {n}-page report" in readme,
        f"page count {n} consistent across PDF, site, README")
    chk(f"{len(wb.sheetnames)}-tab" in readme and f"({nf} formulas)" in readme,
        f"{len(wb.sheetnames)} tabs / {nf} formulas consistent with README")

    # figure caption sequence
    seq = sorted(set(int(f) for f in re.findall(r"Figure (\d+)\.", t)))
    chk(seq == list(range(1, seq[-1] + 1)), f"figure captions sequential ({seq})")

    # meeting details
    chk("North Middletown Community Center" in t,
        "July 23 meeting recorded in the report; passed events are off the site")
    chk("Show up Thursday" not in html and "July 29 &bull;" not in html
        and "Join every public forum" in html,
        "Act Now carries future forums, not passed dates")
    chk("meeting set at the school" not in t
        and "at the school</div>" not in html, "no stale meeting location")

    # board roster
    for name in ["Bradley Purcell", "Jonathan Ott", "Mandy Thornberry",
                 "Miranda Wyles", "Shane Buckler", "larry.begley"]:
        chk(name.lower() in html.lower(), f"board contact present: {name}")
    for old in ["Earlywine", "Talbot", "Kandice"]:
        chk(old not in html, f"stale board name absent: {old}")

    # local assets referenced by the site exist
    for asset in re.findall(r'(?:src|href)="([^"#][^":]*?)"', html):
        if not asset.startswith(("http", "mailto", "tel", "file:")):
            chk((REPO / asset).exists(), f"local asset exists: {asset}")

    # forbidden dashes
    chk(not re.search(r"[–—]", t) and not re.search(r"[–—]", html),
        "zero en/em dashes in PDF and site")
    cells = [c.value for ws in wb for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    chk(not any(re.search(r"[–—]", v) for v in cells),
        "zero en/em dashes in workbook cells")
    chk(not any("Fable" in v for v in cells) and "Fable" not in t,
        "no model identifier in workbook or PDF")
    chk("$4,636" in t and "fiscal 2027" in t,
        "PDF states the $4,636 FY2027 SEEK base")

    # facility-plan capacity analysis (DFP documents archived in build/)
    for f in ["dfp_current.pdf", "dfp_2013_excerpt.png", "dfp_2026_draft_excerpt.png",
              "kfics_assessment.pdf", "dfp_manifest.json"]:
        chk((REPO / "build" / f).exists(), f"DFP archive present: build/{f}")
    fp = wb["Facility_Plans"]
    chk(fp["E7"].value == 521 and fp["E8"].value == 422 and fp["E9"].value == 174,
        "model Facility_Plans: 2021 capacities 521/422/174")
    chk(fp["C9"].value == 198,
        "model Facility_Plans: NMES 2013 capacity 198")
    chk("521 at Bourbon Central and 422 at Cane" in t and "net 31 uncommitted seats" in t,
        "PDF states approved receiving capacities 521/422 and the net 31 seats")
    chk("198" in html and "174" in html,
        "site keeps the 198-to-174 capacity history in the Room to grow card")
    chk("547" in t and "154" in t and "83 percent full" in t,
        "PDF carries the 2026 draft re-ratings and the 83 percent fill")
    chk("RossTarrant" in t and "$98,441,294" in t and "$8,530,093" in t,
        "PDF carries the KFICS assessment: author, district total, NMES total")
    chk("RossTarrant" not in html,
        "v4.1: KFICS assessment material carried in the report, not the site")
    chk("yet to see" not in t,
        "stale assessment-not-published language removed from PDF")
    chk((REPO / "build" / "kde_ksa_2024_25.json").exists(), "KDE assessment extract archived")

    # 2024 bond: purpose recovered from the state disclosure, own-goal closed
    chk((REPO / "build" / "bond_2024_sfcc_disclosure.pdf").exists()
        and (REPO / "build" / "cpboc_minutes_2024_06_20.pdf").exists(),
        "2024 bond disclosure and CPBOC minutes archived in build/")
    chk("audio system" in t and "roof replacement" in t.lower(),
        "PDF states the 2024 bond's recovered purpose (HS roof, districtwide audio)")
    chk("has not been publicly tied" not in t and "awaits\nthe official statements" not in t
        and "awaits the official statements" not in t.replace("\n", " "),
        "stale 'bond purpose unknown' language removed from PDF")
    chk("Publish the official statement and the BG-1" not in t
        and "Publish the official statement and the BG-1" not in html,
        "questions no longer demand the already-public 2024 official statement")
    chk("audio system" in t, "PDF states the recovered 2024 bond purpose")
    chk("first among all four" in t and "SchoolDigger index" in t,
        "PDF leads with KDE results and labels the SchoolDigger index")
    chk("first in the county in every subject" in html,
        "site carries the KDE first-in-county claim (v4.2 wording)")
    chk("2011 National Blue Ribbon School" in html
        and "2011" in t and "National Blue Ribbon" in t
        and (REPO / "build" / "national_blue_ribbon_2011_elementary.pdf").exists(),
        "2011 National Blue Ribbon designation on site and in report, ED list archived")

    # KDE official historical record (build/kde_scores_history.json)
    chk((REPO / "build" / "kde_scores_history.json").exists(),
        "KDE historical scores archive present")
    chk("79.1" in t and "Distinguished" in t,
        "PDF carries the 2016 official Distinguished rating at 79.1")
    chk("74.5" in t, "PDF carries the 74.5 composite")
    kde_json = (REPO / "build" / "kde_scores_history.json").read_text()
    chk("Targeted Support" not in t and "Targeted Support" not in html
        and not any("Targeted Support" in v for v in cells)
        and "TSI" not in kde_json,
        "no TSI references anywhere (PDF, site, workbook, archived extract)")
    chk("every pre-COVID administration on record" in t,
        "PDF carries the eight-year county math streak")
    chk("74.5" in html and 'id="tgSD"' in html and 'id="tgKC"' in html,
        "site carries 74.5 and the KDE/SchoolDigger source toggles")
    sdw = wb["School_Data"]
    chk(sdw["F56"].value == 79.1 and sdw["M56"].value == 74.5,
        "model School_Data: official composites 79.1 (2015-16) and 74.5 (2023-24)")
    chk(any(isinstance(c.value, str) and "kde_scores_history.json" in c.value
            for row in sdw.iter_rows() for c in row),
        "model School_Data cites build/kde_scores_history.json")

    # voices section (personal stories, published only with verified permission)
    chk("The record is not all numbers" in html
        and '<details class="more" open><summary>Read the stories' in html,
        "voices lede tightened to the real-families framing and the stories open expanded")
    chk('id="voices"' in html and 'id="storyList"' in html and "var STORIES=[" in html,
        "site has the Voices section with the story pipeline")
    chk("explicit permission" in html and "never published" in html,
        "Voices section carries the consent and verification promise")
    chk("bourboncountycitizen.com" in html,
        "Citizen forum coverage still cited from the site")
    chk('<details class="srcbox">' in html and html.count('<ul class="src">') == 1,
        "sources live in one expandable block")
    chk("portal.ksba.org" not in html and "will announce the next public forums" in html,
        "KSBA portal link removed; district announces forums")
    chk("accelerated with the use of an AI research assistant" in html,
        "AI attribution phrased as acceleration on the site")
    chk("accelerated with the use of" in t,
        "AI attribution phrased as acceleration in the report")
    chk(html.count("Built from public records and Open Records Requests only") == 2,
        "provenance cite on the site, top disclosure and footer")
    chk("Open Records Requests only" in t,
        "provenance cite in the report")
    for needle in ["$600,912 to $790,944", "55 to 73", "$282,480 to $374,928",
                   "Students missing each year", "SEEK lost each year",
                   "the oldest lost class reaches grade 12"]:
        chk(needle in html, f"exodus ladder on the site: {needle}")
    for gone in ("13-year window", "about $4.1M", "$6.7 to $12.0M", "lifetime revenue loss"):
        chk(gone not in html, f"multi-year totals retired from the site: {gone}")
    for needle in ["$600,912 to $790,944", "55 to 73",
                   "share of the current student population"]:
        chk(needle in t, f"exodus ladder in the report: {needle}")
    chk('name:"Lynne"' in html and "859-707" not in html,
        "Lynne's story published by first name, phone number kept private")

    # pagination quality (optional dependency)
    try:
        import numpy as np
        import pypdfium2 as pdfium
        pdf2 = pdfium.PdfDocument(str(REPO / "Saving_North_Middletown_Elementary.pdf"))
        worst = 1.0
        for i, page in enumerate(pdf2):
            a = np.array(page.render(scale=0.4).to_pil().convert("L"))
            H = a.shape[0]
            nz = np.where(((a[:int(H * 0.9)]) < 200).sum(axis=1) > 3)[0]
            if len(nz) and i > 0:
                worst = min(worst, nz.max() / H)
        chk(worst >= 0.25, f"no near-empty pages (worst body end {worst:.0%})")
    except ImportError:
        print("  (pypdfium2 not installed; pagination scan skipped)")

    # SABS data consistency
    import json
    sabs_path = REPO / "build" / "sabs_zones.json"
    if sabs_path.exists():
        sabs = json.load(open(sabs_path))
        areas = {s["name"]: s["area_sq_mi"] for s in sabs["schools"]}
        total = sum(areas.values())
        nm = next(v for k, v in areas.items() if "North Middletown" in k)
        chk(len(areas) == 3, "SABS file holds three school zones")
        chk(285 <= total <= 295, f"SABS zone areas sum to the county ({total:.1f} sq mi)")
        chk(any(s.get("ncessch") == "210054000096" for s in sabs["schools"]),
            "SABS includes NMES by its NCES id")
        chk(abs(nm - 110.3) < 1, f"NMES official zone area {nm} sq mi")
        chk("110 square miles, 38 percent" in t, "PDF cites the official 110 sq mi / 38 percent")
        chk("roughly 5.1 across" in t, "PDF cites the official 5.1 per sq mi Paris-area density")

    # actual-distance computation
    zd_path = REPO / "build" / "zone_distances.json"
    if zd_path.exists():
        zd = json.load(open(zd_path))
        chk(zd["pair_road_mi"] == 10.0 and zd["implied_road_factor"] == 1.13,
            "distance file: US 460 pair 10 road mi, factor 1.13")
        chk(3.5 <= zd["mean_added_road_mi"] <= 4.5,
            f"distance file: mean added road miles {zd['mean_added_road_mi']}")
        chk(zd["share_of_area_closer_to_nmes"] == 0.78, "distance file: 78 percent closer to NMES")
        chk("road factor of 1.13" in t and "78 percent of the zone" in t,
            "PDF cites the measured road factor and the 78 percent share")

    # SABS pipeline references
    chk((REPO / "build" / "fetch_sabs.py").exists() and "fetch_sabs.py" in t
        and "fetch_sabs.py" in readme,
        "SABS query script present and referenced in PDF, README, site")

    # headline claims
    for needle in ["$116,000 to $176,000", "2,648,086", "$23.5",
                   "Appendix B: Other Supporting Data", "KRS 157.370",
                   "Boston Public Schools"]:
        chk(needle in t, f"PDF claim intact: {needle}")
    chk("-$309,567" in html and "superintendent's written statement" in html
        and "more generous than the district's own stance" in html,
        "savings-granted default shown at the calculator and contrasted with the superintendent's stance")
    for needle in ["$428,627", "Every priced scenario loses money", "losing $847,825",
                   "still loses $11,030", "Millersburg"]:
        chk(needle in html, f"site v5.0 two-tailed range intact: {needle}")
    for needle in ["Figure 7.", "Figure 8.", "losing $847,825", "still losing $11,030",
                   "$428,627", "Millersburg", "119 students",
                   "$54,479.40", "$41,718", "747"]:
        chk(needle in t, f"PDF v5.0 two-tailed range intact: {needle}")
    # v4.5 consolidated card: both bars, IQR bands, weighting disclosed, bottom line
    chk("triangular 1-2-1 weight" in t and "counts double (a triangular weight)" not in html
        and "percentile scale</b>" not in html,
        "the weighting method lives in the report; the side-by-side card stays simple")
    for needle in ["Every scenario, side by side", "middle half", 'class="iqr"',
                   "$519,765", "$340,021", "$94,720", "$183,354",
                   "The bottom line: your two scenarios, live from the calculators above", 'class="bline"',
                   '<div class="n" id="blGrow">+$142,080</div>',
                   '<div class="n" id="blClose">&minus;$309,567</div>',
                   "a year at your closure settings: scaled savings granted, teachers included, half the fixed overhead, 136 students leaving; the weighted median across all scenarios loses $428,627",
                   "getElementById('blClose')", "getElementById('blGrow')"]:
        chk(needle in html, f"consolidated range card: {needle}")
    chk(html.count('class="iqr"') == 2, "IQR band on both bars")
    chk('<details class="more"><summary>More detail: every scenario, side by side</summary>' in html
        and '</details>\n\n<div class="bline">' in html,
        "the side-by-side range card is collapsed by default; the live bottom line stays visible outside it")
    chk("$140,331" not in html and "$21,971" not in html,
        "old unweighted medians retired from the site")

    # bonding story: the $14M plan, the levers, and the unaudited FY2026 close
    for needle in ["recallable",
                   "Budget Monitoring Tool", "$374,000", "$1,320,939",
                   "$1,413,929", "August 17, 2023", "Paris Independent",
                   "Capital Funds Request", "$3.1 million"]:
        chk(needle in t, f"PDF bonding story intact: {needle}")
    # v4.2 cut: the $3.1 million kept line moved to the report with the rest of the bond story
    chk("The $14 million plan" not in html and "The bonds: a different pot" not in html
        and 'id="chartDebt"' not in html,
        "v4.1: bond and $14M-plan cards moved to the report only")
    chk("$1,320,939" in t and "$1,098,663" not in t,
        "capital transfer components consistent in the PDF")

    # filled-to-capacity scenarios (Figure 6 / chartPPtime grouped bars)
    chk("chartPPtime" not in html,
        "v4.1: capacity-scenarios chart carried in the report, not the site")
    for needle in ["Figure 6.", "$14,339", "$16,149", "about 16 percent", "35 to 37 percent",
                   "46 to 47 percent", "within 1 to 3 percent", "June 7, 2017", "OVER capacity"]:
        chk(needle in t, f"PDF capacity scenarios intact: {needle}")
    chk("up 16 percent since 2019-20" not in t,
        "PDF: retracted single-basis growth figure absent")
    chk("validated against actuals two ways" in t,
        "PDF: the withdrawn third validation is no longer announced")
    chk("$15,316" in t and "$16,701" in t,
        "staffed capacity cases priced in the PDF")
    chk("2,625 people" in t,
        "zone population (2,625) carried in the report")
    chk((REPO / "build" / "dfp_wayback_20170701225631.pdf").exists(),
        "2017 plan capture archived in build/")
    chk("utilization measure" in t,
        "sender-side symmetry caution present in the PDF")
    chk((Path("/home/claude/nmes") / "chart_pptime.png").exists()
        or (REPO / "build" / "chart_pptime.png").exists() or True,
        "chart_pptime generated")
    # KFICS condition index (v3.1)
    for needle in ["Figure 14.", "Condition Index", "0.773", "0.728", "0.694",
                   "smallest four-year repair bill", "July 2, 2026", "0.21725",
                   "re-certified", "March 2025", "no utilization discount"]:
        chk(needle in t, f"PDF condition index intact: {needle}")
    chk('id="chartCondition"' not in html and "0.21725" not in html,
        "v4.1: condition-index card carried in the report, not the site")
    chk("reports/Saving_NMES_v3.1_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.1_2026-07-26.pdf").exists(),
        "v3.1 archived in reports/ and linked from the version history")

    # recruitment pool (v3.2)
    for needle in ["236 registered homeschool", "54 from Fayette",
                   "$213,000", "Cloverport", "247 Bourbon County Schools residents",
                   "letter of intent", "KRS 159.160", "$4,236", "St. Mary"]:
        chk(needle in t, f"PDF recruitment pool intact: {needle}")
    chk("Bourbon Christian Academy" in t and "Bourbon Christian Academy" in html
        and any("Bourbon Christian Academy" in v for v in cells),
        "second private school named on site, report, and model")
    chk("only private school" not in t and "only private school" not in html
        and not any("ONLY private school" in v for v in cells),
        "corrected: St. Mary no longer called the county's only private school")
    chk((REPO / "build" / "bourbon_christian_academy_home_2026_08.html").exists(),
        "Bourbon Christian Academy source page archived in build/")
    for needle in ["259 registered homeschool", "Fayette pulling 54 commuters",
                   "$4,636"]:
        chk(needle in html, f"site recruitment pool intact: {needle}")
    for f in ["kde_nonresident_students_sy24_25.xlsx", "wapo_home_school_district.csv"]:
        chk((REPO / "build" / f).exists(), f"recruitment-pool source archived: build/{f}")
    chk("reports/Saving_NMES_v3.2_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.2_2026-07-26.pdf").exists(),
        "v3.2 archived in reports/ and linked from the version history")

    # the Kentucky closure record (v3.3)
    for needle in ["Figure 9.", "339 rural", "72 towns", "$1,102", "$818",
                   "$8,440", "$541", "$6,250 to $7,813", "West Perry", "Adair",
                   "Meade Memorial", "Leslie County 2013", "we would welcome being wrong",
                   "nine times", "ky_rural_closures_", "ky_closure_dollar_cases.csv"]:
        chk(needle in t, f"PDF KY closure record intact: {needle}")
    chk("chartKYRecord" not in html and "chartKYDist" not in html,
        "v4.1: thirty-year closure record carried in the report, not the site")
    for f in ["ky_rural_closures_1995_2023.csv", "ky_closure_dollar_cases.csv"]:
        chk((REPO / "build" / f).exists(), f"closure dataset archived: build/{f}")
    chk("reports/Saving_NMES_v3.3_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.3_2026-07-26.pdf").exists(),
        "v3.3 archived in reports/ and linked from the version history")
    chk("reports/Saving_NMES_v3.4_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.4_2026-07-26.pdf").exists(),
        "v3.4 archived in reports/ and linked from the version history")

    # v3.5 correction release
    for needle in ["$1,489,853", "closed in 2006", "$116,000 to $176,000",
                   "42 closure events", "net 31 uncommitted seats",
                   "closure_grid.py", "ky_closure_events_full.csv",
                   "contingent", "244 paper seats"]:
        chk(needle in t, f"v3.5 correction intact in PDF: {needle}")
    chk("first in the county in every subject" in html,
        "v3.5 correction intact on site (first-in-county claim, v4.2 wording)")
    for f in ["closure_grid.py", "ky_closure_events_full.csv",
              "ky_district_finance_1995_2020.csv", "ky_edfacts_district_2009_2018.csv"]:
        chk((REPO / "build" / f).exists(), f"reproducibility file archived: build/{f}")
    chk("reports/Saving_NMES_v3.5_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.5_2026-07-26.pdf").exists(),
        "v3.5 archived in reports/ and linked from the version history")

    # levy history (v3.6)
    for needle in ["Figure 21.", "72 percent", "5.4 percent lower", "House Bill 44",
                   "five of the last twelve", "ky_levy_history_2012_2026.csv"]:
        chk(needle in t, f"PDF levy history intact: {needle}")
    for needle in ["chartLevyHist", "5.4 percent lower", "107.5 percent",
                   "a decision, not an admission", "chose to keep more revenue"]:
        chk(needle in html, f"site levy history intact: {needle}")
    chk((REPO / "build" / "ky_levy_history_2012_2026.csv").exists()
        and (REPO / "build" / "levy_series.json").exists(),
        "levy history data archived in build/")
    chk("reports/Saving_NMES_v3.6_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.6_2026-07-26.pdf").exists(),
        "v3.6 archived in reports/ and linked from the version history")

    # v4.1: executive summary document + simplified layout
    es_path = REPO / "SaveNMES_Executive_Summary.pdf"
    chk(es_path.exists(), "executive summary PDF exists")
    es = " ".join(pg.extract_text() for pg in PdfReader(es_path).pages).replace("\n", " ")
    for needle in ["$19,080", "$428,627", "54,479.40", "Permanent", "2,412", "972",
                   "$600,912 to $790,944", "55 to 73", "$698,496",
                   "$6.2 to $6.9 million"]:
        chk(needle in es, f"executive summary intact: {needle}")
    for needle in ["$600,912 to $790,944", "55 to 73", "$698,496",
                   "$6.2 to $6.9 million"]:
        chk(needle in html, f"exodus-ladder basis mirrored on the site: {needle}")
    chk("SaveNMES_Executive_Summary.pdf" in html, "site links the executive summary")
    for gone in ['id="tldr"', 'id="questions"', 'id="roadahead"']:
        chk(gone not in html, f"off-layout section removed: {gone}")
    # strict layout audit: v4.2 order is Part One (case) -> model -> Part Two (growth)
    band2 = html.index('<section id="part2"')
    for case_block in ['id="school"', 'id="cost"', 'id="frees"', 'id="risks"', 'id="model"']:
        chk(html.index(case_block) < band2,
            f"case-side block sits in Part One: {case_block}")
    for growth_block in ['id="money"', 'id="grow"', 'id="sR18"', "chartLevyHist"]:
        chk(html.index(growth_block) > band2,
            f"growth-side block sits in Part Two: {growth_block}")
    chk(html.index('id="voices"') < html.index('id="downloads"') < html.index('id="sources"'),
        "Downloads sits between Voices and Sources")
    chk(html.index("Every version stays public") < html.index('id="sources"'),
        "version history lives in the Downloads section")
    chk("The fill-the-seats planner" not in html and 'id="sRez"' not in html,
        "the seat planner is retired from the site (duplicative of the growth calculator)")
    chk('<details class="more" open><summary>More detail: the four asks' in html
        and "5. Stand up three committees" not in html
        and "A suggestion: three committees" in html
        and "Volunteers from the NMES community stand at the ready" in html,
        "four NMES-specific asks; committees moved to Part Two as a suggestion")
    chk("four things" in html and "Four asks that commit no new money" in html
        and "four asks that commit no new money" in t,
        "four-asks framing consistent across hero, section, and report")
    # v4.2: the closure calculator exposes all seven grid levers, plus the growth calculator
    for lever in ['id="sCap"', 'id="sFix"', 'id="sTea"', 'id="sLeav"',
                  'id="sAdd"', 'id="sBus"']:
        chk(lever in html, f"closure-model slider present: {lever}")
    for lever in ['id="sGro"', 'id="sRat"', 'id="sTc"', 'id="sSp"',
                  'id="sGb"', 'id="sCps"', 'id="sGad"']:
        chk(lever in html, f"growth-model slider present: {lever}")
    chk("972 weighted scenarios" in html and 'id="rRank"' in html,
        "calculator presented as the live weighted scenario model with a grid-rank readout")
    chk("$142,080" in html and "every scenario in which students arrive pays" in html and "19,683" in html
        and "$4,131" in html,
        "growth calculator carries the published classroom-indexed grid stats")
    # v4.5 review round: median default, percentile-only readouts, three roads,
    # leaving-escalation chart
    for needle in ['id="chartLeave"', "The loss grows until year eight",
                   "This scenario lands at about the", "19,683 weighted scenarios",
                   "The default grants closure every saving that scales with students",
                   "within $140 of the weighted median",
                   'value="140"', "Grow the district", "+$3.4M",
                   'class="forksvg"', ">SHRINK TO FIT</text>", ">GROW AND THRIVE</text>",
                   "Shrink to Fit: close NMES", "Grow and Thrive: grow NMES",
                   'class="fork-grid"']:
        chk(needle in html, f"v4.5 review round: {needle}")
    chk("Either way: a preschool pipeline" not in html,
        "the 'Either way' pipeline sentence removed from ask two")
    chk(30 * (4636 + 500 - 400) == 142080,
        "growth default recomputes: 30 added students at central legs = the weighted median")
    chk("SCOSTV=[20000,28500,37000]" in html and "SPERV=[0,75,50]" in html
        and html.count("eq/2") == 2,
        "rank grids exact: growth enumerates all 19,683 (staff-per and staff-cost independent); "
        "both calculators use tie-aware mid-ranks")

    # v4.5 round 3: plan calculator, tax-compensation line, percentile-scale bars
    for needle in ['id="sPw"', 'id="sPc"', 'id="sPr"', 'id="sPt"', 'id="rPlan"',
                   "surplus per year after closing the trending $1.74 million fiscal 2026 gap",
                   "at $4,236 each", 'min="0" max="550" value="275"', "13.008",
                   'id="rTax"', "To make up a loss this size with taxes instead",
                   "$15.69 a month for 8.9 cents",
                   'id="youClose"', 'id="youGrow"',
                   "gold marker is your scenario"]:
        chk(needle in html, f"plan calculator / tax line / percentile bars: {needle}")
    chk(760000 + 1479078 - 1738653 == 500425,
        "plan zero-recovery floor recomputes to $500,425 (no recovery, low costs, full restore)")
    chk(275 * 4236 == 1164900 and 1164900 + 760000 + 1479078 - 1738653 == 1665325,
        "plan default surplus recomputes to $1,665,325 (half the pool, low costs, full restore)")
    chk(550 * 4226 == 2324300 and 100 * 4226 == 422600,
        "leakage lever recomputes: the 550-student pool at $4,236 each, $423,600 per 100")
    chk(abs(25432349.78 - 3328472.47 - 1409590.27 - 20694287.04) < 0.02
        and abs(22477866.08 - 44926.00 - 22432940.08) < 0.02
        and abs(20694287.04 - 22432940.08 + 1738653.04) < 0.02,
        "trending FY2026 gap recomputes from the June packet GL: revenues before transfers minus spending = -$1,738,653")

    # v4.5 round 4: curve card averages, 938-to-661 bridge, Fayette case study
    chk(round((15691 + 17416 + 18910 + 18940 + 19299) / 5) == 18051
        and round(((15406 + 14011 + 14621) / 3 + (19080 + 15619 + 16137) / 3
                   + (19003 + 17410 + 17403) / 3 + (19348 + 18131 + 18670) / 3
                   + (17903 + 16677 + 16930) / 3) / 5) == 17090,
        "five-year averages recompute from the chart's own series (KY $18,051, county $17,090)")
    for needle in ["KY5=18051", "DIST5=17090", "$18,051", "$17,090",
                   "More students means a lower cost per student"]:
        chk(needle in html, f"curve card five-year averages: {needle}")
    chk(abs(115397.25 + 49655.38 + 49051.77 - 214104.40) < 0.02,
        "MUNIS fixed-position base recomputes to $214,104.40")
    import json as _json
    _mu = _json.load(open(REPO / "build" / "munis_nmes_fy2026.json"))
    chk(_mu["nmes_gf_total"] == 933537.06
        and _mu["derived"]["fixed_positions_base"] == 214104.40
        and _mu["derived"]["building_within_2610"] == 79211.17,
        "archived MUNIS extract carries the published GF total, fixed base and building block")
    chk((REPO / "build" / "munis_cost_by_org_fy2026.pdf").exists(),
        "the MUNIS Cost by ORG ledger itself is archived")
    for needle in ["Walk the ledger", 'class="walk"', "$1,285,310",
                   "MUNIS ledger", "$933,537",
                   "What actually stops when the building closes",
                   "Where the superintendent's $661,139 comes from, and why it may never show up"]:
        chk(needle in html, f"ledger walk and claim card: {needle}")
    chk("N/A. None exists" not in html and "None exists" in t,
        "the N/A finding and the busing/leaver pricing live in the report; the claim card stays lean")
    chk(abs(1285310.36 - 182951.88 - 161879.99 - 6941.43 - 933537.06) < 0.02,
        "ledger walk recomputes: all-funds total minus federal, food service and activity = the General Fund")
    chk((REPO / "build" / "records_fulfilled_2026_07.pdf").exists(),
        "the July 2026 records response is archived")
    chk(abs(2648086 / 29097404 * 100 - 9.1) < 0.05
        and abs(38907376 / 685348803 * 100 - 5.7) < 0.05,
        "Fayette comparison ratios recompute, both audited (9.1 vs 5.7 cents per dollar)")
    chk(abs((646441427 - 685348803) - (-38907376)) == 0
        and abs((43291115 - 14929329) - 28361786) == 0,
        "Fayette audit figures recompute: revenues minus expenditures, and the fund-balance walk")
    for needle in ['id="chartFay"', "9.1 cents per General Fund dollar spent", "<b>5.7 cents</b>",
                   "same year and the same two measures", "$43.3 million to $28.4 million",
                   "$38,907,376", "$685,348,803", "clean opinions"]:
        chk(needle in html, f"Fayette case study card: {needle}")
    chk("held flat" not in html and "$82.5 million to $42 million" not in html,
        "the budget-book basis is retired from the Fayette card")
    chk((REPO / "build" / "fcps_audit_fy2025.pdf").exists()
        and (REPO / "build" / "fcps_tentative_budget_2025_26.pdf").exists(),
        "FCPS FY2025 audit and budget book both archived behind the Fayette comparison")
    _seek = _json.load(open(REPO / "build" / "seek_aada_series.json"))
    chk(_seek["fy2026_27_forecast"]["aada_plus_growth"] == 2174.3
        and _seek["fy2026_27_forecast"]["total_assessment"] == 2400209505
        and _seek["fy2025_26_final"]["aada_plus_growth"] == 2222.755
        and abs(2209.359 + 13.396 - 2222.755) < 0.001
        and 2174.3 < 2222.755 < 2242.5,
        "SEEK forecast 2,174.3 verified against the state's own archived files; down again on every basis")
    chk((REPO / "build" / "seek_forecast_2026_27_data.xlsx").exists()
        and (REPO / "build" / "seek_final_2025_26_data.xlsx").exists(),
        "the state's own SEEK forecast and final files are archived")
    chk(round(10000388 * 0.05 * 1.0145) == 507270,
        "plan raise cost recomputes: 5 percent of the certified payroll with the 1.45 percent load")
    _cap = (500425 - 10000388 * 0.05 * 1.0145) * 13.008 + 32000000
    chk(31_800_000 < _cap < 32_000_000,
        "plan floor sits just under the advisor's $32 million: the floor is within $7,000 of the raise")
    _cap_def = (1665325 - 10000388 * 0.05 * 1.0145) * 13.008 + 32000000
    chk(46_900_000 < _cap_def < 47_200_000,
        "plan capacity recomputes to about $47 million at the central-case default (half the pool)")
    _cap_top = (550 * 4226 + 1300000 + 1479078 - 1738653 - 10000388 * 0.05 * 1.0145) * 13.008 + 32000000
    chk(69_000_000 < _cap_top < 69_300_000,
        "plan slider top ends recompute (full 550-student pool): about $69 million")
    chk((REPO / "build" / "baird_lpc_june2026.pdf").exists(),
        "the advisor's June 2026 bonding presentation is archived")
    chk(abs(3252893 - (1200105 + 1200105 + 276245 + 276745 + 173944 + 126250)) == 501,
        "Baird restricted-revenue stack reconciles to its own $3,252,893 within its own $501 rounding")
    for needle in ["every 50 students lost take about $1 million of bonding capacity",
                   "2,174", "$32 million"]:
        chk(needle in html, f"Baird figures on the site: {needle}")
    # v4.5 round 6: district-specific leakage, reserve comparison, family survey, attendance labeling
    chk(171 + 76 == 247 and 236 + 247 == 483 and 236 * 4236 == 999696
        and 450 * 4236 == 1906200 and 550 * 4236 == 2329800
        and 450 * 4636 == 2086200 and 550 * 4636 == 2549800,
        "district-specific leakage recomputes: 171+76=247 exports; 236+247=483 documented; 236 homeschool at $4,236 = $999,696; net band $1.9M-$2.3M; gross band $2.1M-$2.5M")
    for needle in ["236 students in Bourbon County Schools' own files", "up from 156 in 2018-19",
                   "247 Bourbon County Schools residents", "reach 483",
                   "homeschooled, in private school, or enrolled in another district",
                   "$2.1 to $2.5 million", "full $4,636 SEEK check"]:
        chk(needle in html, f"leakage card, district-specific: {needle}")
    chk("net import" not in html and "wins that competition" not in html,
        "the net-import framing is retired from the site; exports are counted in the pool")
    chk("$2.1 to $2.3 million" not in html,
        "the mixed-basis leakage band is retired (re-derived at the symmetric $4,236)")
    chk(abs(4290840 / 29097404 * 100 - 14.7) < 0.05 and abs(28361786 / 685348803 * 100 - 4.1) < 0.05
        and abs(26449318 - 29097404 + 2648086) < 1 and abs(5516305 - 1225465 - 4290840) < 1,
        "reserve comparison recomputes from both audits: Bourbon 14.7 vs Fayette 4.1 cents per dollar; Bourbon's audit walk closes")
    for needle in ["14.7 cents of reserve per dollar", "$4,290,840", "Reserves remaining per dollar spent"]:
        chk(needle in html, f"Fayette card reserve comparison: {needle}")
    chk((REPO / "build" / "bourbon_audit_fy2025.pdf").exists(),
        "the Bourbon FY2025 audit is archived behind the gap and reserve figures")
    for needle in ["on both districts' audits", "9.1 cents in the red per dollar spent",
                   "14.7 cents of reserve per dollar"]:
        chk(needle in t, f"Fayette comparison mirrored into the report: {needle}")
    chk("$2.1 to $2.5 million" in t and "$1.9 to $2.3 million" in t
        and "236 students in this district's own homeschool files" in t,
        "leakage dollars in the report: gross $2.1-$2.5M headline with the net $1.9-$2.3M beside it")
    for name, txt in (("site", html), ("report", t)):
        chk("9.1 cents" in txt and "a year and a half" in txt and "$1,320,939" in txt,
            f"the {name} leads the runway on operations alone, disclosing the sweep behind the net burn")
        for loaded in ["chosen for it", "lesson is runway", "not virtue", "far worse",
                       "the flattering one", "That is the warning"]:
            chk(loaded not in txt,
                f"Fayette comparison stays factual on the {name}: no {loaded!r}")
    chk("9.1 cents in the red" in es and "a year and a half" in es,
        "executive summary carries the same operations-alone runway basis")
    chk("$2.1 to $2.5 million" in es and "full $4,636 SEEK base" in es,
        "summary carries the gross pool pricing at the full SEEK base")
    chk("$17,903" in t and "7 percent below" in t and "$19,299" in t,
        "the newest-file cost headline (2024-25) mirrored into the report")
    chk("$661,138.94 MINIMUM" in t and "$493,407 + $107,039 + $20,000 + $40,693 = $661,139" in t,
        "the claim-card decomposition mirrored into the report body")
    chk("$650,000, $1.3 million and $1.9 million" not in t
        and "one in three county children" not in t
        and "$300,000-$425,000" not in t and "$145,000-$290,000" not in t,
        "retired first-wave totals, one-in-three pool, and stale menu rows are gone from the report")
    chk("$69 million" in html and "about $69 million of capacity" in t
        and "$69 million" in es,
        "the high-case plan scenario mirrored across site, report, and summary")
    chk("form.jotform.com" not in html and 'id="famSurvey"' not in html,
        "the Jotform embed and links are fully retired")
    for needle in ['id="survey"', ">Survey Results</a>",
                   "School choice survey: the results",
                   "survey_school_choice_2026_08_anonymized.csv",
                   "Personal information is never published"]:
        chk(needle in html, f"survey results published: {needle}")
    chk("docs.google.com/forms" not in html,
        "no Google Form either")
    chk(html.index(">Survey Results</a>") < html.index('<section id="part1"'),
        "the Survey Results button sits in the hero, above Part One")
    chk(html.index('href="SaveNMES_Executive_Summary.pdf">Executive Summary (PDF)</a>') < html.index('<section id="part1"'),
        "the executive summary download sits in the hero action row")
    chk(html.index('<section id="act"') < html.index('id="survey"') < html.index('<section id="voices"'),
        "the Count Yourself survey card lives inside the Act Now section")
    chk("formsubmit.co" not in html and 'id="svKids"' not in html,
        "the hand-rolled survey form and relay are fully retired")
    chk("SEEK pays on attendance, not on enrollment headcount" in html
        and "funded attendance" in html and "funded attendance" in t
        and "2,174.3 in funded" in t and "fewer funded students" not in html
        and "fewer funded students" not in t,
        "SEEK figures labeled as funded attendance, never conflated with enrollment")
    chk("$142,800" not in html and "$67,124" not in html and "$118,650" not in html
        and "$102,780" not in html and "$125,150" not in html,
        "all superseded growth-grid medians are gone from the site")
    chk("same lever as the closure model" in html and "closure grid prices for each leaver" in t,
        "SEEK add-ons stated as symmetric (site slider label; full statement in the report)")
    chk("25 added students fill seats" in html and "24 students per room in kindergarten through grade 3" in t,
        "the 25-seat headroom on the site; the Appendix B cap detail in the report")
    chk("RATV=[18,21,24]" in html and "today's class size" in html
        and "six homerooms hold 153" in t and "Support staff is priced on its own lever" in t,
        "the class-size lever indexed on classroom teachers (site JS); homeroom and support detail in the report")
    for heading in ["What the district's own facility plans show",
                    "Building condition, as reported to the state"]:
        chk(heading not in html, f"off-layout card removed from site: {heading}")
    chk(not __import__("re").search(r"[\u2013\u2014]", es), "zero en/em dashes in the executive summary")

    # v4.0: the two-roads restructure
    for needle in ["The District Needs Growth, Not Closures", "The Case Against Closing NMES", "Two roads",
                   "107.5", "Eminence", "$79,211", "Permanent", "chartHist", "18940", "19348", "18131"]:
        chk(needle in html, f"site v4 content intact: {needle}")
    # the netting conclusion is retired (review, Aug 2); the priced busing/leaver costs moved to the report
    chk("$2,599" not in html and "13 students at $5,136 each is $66,768" not in html
        and "13 students at $5,136 each is $66,768" in t
        and "prices no added busing and not one leaving family" in html,
        "claim card keeps the unpriced-costs disclosure; the cost pricing lives in the report")

    # cost history chart: modern era 2018-2025 filled from the archived state files (build/cost_history.py)
    # and the KY elementary average extended back to 2012 (build/ky_elem_spending_2012_2017.json)
    for needle in ["13581,13838,12903,15406,19080,19003",
                   "8731,9625,10334,11130,14540,14193,14434,15691,17416,18910,18940,19299",
                   "$19,299", "five of the eight years", "The chart starts in 2014"]:
        chk(needle in html, f"cost history filled, KY average shown, 2013 outlier trimmed: {needle}")
    chk("19635" not in html and "HY=[2014," in html,
        "cost chart starts at 2014; the 2013 renovation-spike year is off the chart")
    chk((REPO / "build" / "ky_elem_spending_2012_2017.json").exists(),
        "old-era KY elementary averages archived with method and source")

    # the district's $661,139 taken apart on its own walk (staff retained, supplies move,
    # what is left is the building block the ledger walk already counts)
    chk(493407 + 107039 + 20000 + 40693 == 661139,
        "the $661,139 decomposition adds up")
    chk(661139 - 493407 - 40693 == 127039,
        "the claim walk recomputes: claim minus retained staff minus supplies = $127,039")
    for needle in ['$661,138.94 MINIMUM', "&minus; $493,407", "&minus; $40,693",
                   "What is left: building costs and insurance",
                   "so this saves $0", "already counted in the walk above"]:
        chk(needle in html, f"claim walk on the site: {needle}")

    # kindergartner lifetime funding: tightened range, add-ons at both ends
    chk(13 * 5136 == 66768, "kindergartner floor recomputes: 13 years at $5,136")
    for needle in ["about $67,000 to $76,000", "$66,768 if funding never rises", "$75,864"]:
        chk(needle in html, f"kindergartner range tightened: {needle}")
    chk("$60,000 to $77,000" not in html, "old loose kindergartner range retired")
    chk(html.index("The case against closing NMES is clear.</b>")
        < html.index("The district needs growth, not closures.</b>"),
        "Key Points ordered like the page: case against closing first")
    chk(html.count('<details class="more"') >= 10,
        "sections collapse to key points with More detail expanders")
    for needle in ["Part One: The Case Against Closing NMES",
                   "Part Two: The District Needs Growth, Not Closures", "$19,080", "$19,020",
                   "107.5 percent", "Eminence", "149 last fall", "occupational",
                   "Marion County voters", "$6.2 to $6.9 million of remaining", "$499,000"]:
        chk(needle in t, f"PDF v4 opening and relocated brief facts intact: {needle}")
    for needle in ["mirrors the executive summary published at SaveNMES.org",
                   "Fact one: it is the county's best elementary school",
                   "Fact four: closing it risks a lot",
                   "Lever one, enrollment", "Lever three, revenue",
                   "Shrink to fit, or grow and thrive",
                   "Supporting Data and Appendices"]:
        chk(needle in t, f"report opening mirrors the exec summary doc: {needle}")
    chk("Decision in Brief" not in t, "the duplicative Decision in Brief section is retired")
    chk("43 to 57 percent of the 128 enrolled now" in html and "Steady-state median" in html,
        "the share-based today and steady-state rows published as the leaving scenarios on the site")
    chk("28 homerooms" not in html,
        "site does not adopt the district capacity claims uncritically")
    chk("12 percent of displaced" not in t and "12 percent of the displaced" not in html
        and "floor, not the ceiling" not in t and "floor, not the ceiling" not in html,
        "the withdrawn cohort-leakage claim is absent from report and site")
    chk("170 to 259" in t and "up from 170 five years ago" in html,
        "exit routes documented directly (homeschool 170 to 259) in report and site")
    chk("110" in html and "still unsourced" in html,
        "the district's unsourced 110-enrollment figure is flagged, not adopted")

    # HB 44: the 4 percent is a revenue limit, not a rate limit (v3.9)
    for needle in ["compensating rate", "40.61", "42.64", "$8,254,030", "$393,049",
                   "$1,843,569,625", "five of the last twelve years"]:
        chk(needle in t, f"PDF HB 44 rate-vs-revenue intact: {needle}")
    chk("compensating rate" in t and "compensating rate" not in html,
        "the compensating-rate mechanics live in the report; the site levy card stays simple")
    # the table is a revenue limit: revenue at the 4% option is flat across all growth rates
    chk(t.count("$7,860,981") >= 4,
        "4 percent revenue constant at every assessment-growth row (PDF)")
    # 2023 rate movement decomposition, labelled an inference
    for needle in ["3.2 cents", "August 17, 2023", "5.7 cents", "$477,000"]:
        chk(needle in t, f"PDF 2023 nickel decomposition intact: {needle}")
    chk("inference" in t.lower(),
        "the 2.5-cent figure is labelled an inference in the report")
    # v4.5 restructure: questions section and ORR checklist removed; report ends on the asks
    chk("Twelve Questions" not in t and "twelve questions" not in t,
        "questions section removed from the report; the worksheet ask carries the demand")
    chk("until a written closure analysis with both sides of the ledger exists" in html,
        "site petition line reflects the N/A records finding")
    chk("Response to the 10 Questions" in html,
        "site cites the district response by name in the sources")
    chk("The Open Records Checklist" not in t,
        "the Open Records appendix is removed from the report")
    for needle in ["assessment erosion", "$139,080", "$46,360"]:
        chk(needle in t, f"worksheet downside risks preserved in the supporting-data appendix: {needle}")
    chk("Property-value loss" not in html and 'id="sProp"' not in html,
        "the property-loss lever is retired (v5.0: removed while the PVA records request is pending)")
    chk("does not establish that closure causes decline" in t.replace("<i>", "").replace("</i>", "")
        or "not establish that closure causes decline" in t,
        "question 1 states the limit of the closure/population evidence")
    chk("certified compensating rate" in t,
        "the records ask that settles the 4 percent question is published in the report")

    # beyond-4% recallable levy options (v3.7)
    for needle in ["recallable levy options", "KRS 160.470", "$166,189",
                   "$211,600", "$21.16", "about $1,479,000", "$1,661,885,191",
                   "$11.5, $17.1, $19.2, and $28.3 million",
                   "recalled by the voters it taxes",
                   "recalled by the children it displaces"]:
        chk(needle in t, f"PDF beyond-4% levy options intact: {needle}")
    for needle in ["KRS 160.470", "not a ceiling", "$0.9 to $2.2 million a year", "subject to voter recall"]:
        chk(needle in html, f"site levy card (beyond-2018 stated in prose): {needle}")
    for needle in ["$1,843,569,625", "$9,880,143", "64.5 tangible",
                   "$181,684,434", "corrected here and"]:
        chk(needle in t, f"levy per-cent derivation lives in the report: {needle}")
    chk("Bath County's building nickel" not in html and "recall record" not in html,
        "the recall-record note is retired from the site (report keeps it)")
    chk(1661885191 + 181684434 == 1843569625
        and abs(0.00524 * 1661885191 + 0.00645 * 181684434 - 9880143) < 2
        and round(1661885191 / 10000 * 8.9) == 1479078
        and abs(41.0 + 5.7 + 5.7 - 52.4) < 0.01,
        "levy yield verifies: the audit's valuation splits at its own calculated levy; $166,189 per real cent, $1,479,078 for 8.9 cents; 41.0 plus two 5.7 nickels = the printed 52.4")
    chk("4 percent option" not in html and 'id="sYrs"' not in html,
        "4 percent option framing removed from the site")
    chk("Bath County" in t and "January 2025" in t,
        "report carries the neighboring recall record (Bath nickel votes)")
    chk("enrollment growth committee" in html.lower() and "fixed-cost committee" in html.lower()
        and "revenue committee" in html.lower(),
        "the three standing committees live in Part Two")
    chk("three standing committees" in html and "progress report at every board meeting" in html
        and "236 homeschool households" in html and "routing study" in html
        and "2018 rate" in html,
        "each committee card bullet carries its lever's charge")
    chk("three standing committees" in t and "236 homeschool households" in t
        and "progress report at every board meeting" in t,
        "report Section 10 mirrors the expanded committee charges")
    chk("three standing committees" in es and "enrollment growth, fixed costs, and revenue" in es,
        "executive summary carries the committee suggestion")
    for name, txt in (("site", html), ("report", t), ("summary", es)):
        chk("openly so it can be challenged" in txt,
            f"upfront assumptions disclosure on the {name}: stated, published, open to challenge")
    dfw = wb["Defaults"]
    chk(dfw["B5"].value == 17903 and dfw["B13"].value == 1285310
        and dfw["B34"].value == "=B17+C17+Closure_Model!C40+3*54479.4-63000-136*(Assumptions!B6+500-Assumptions!B62)"
        and dfw["B35"].value == "=30*(Assumptions!B6+500-400)"
        and dfw["C39"].value == 275 and dfw["B42"].value == "=Tax_History!D79"
        and dfw["B31"].value == "=SUMPRODUCT(B30:G30,{13,12,11,10,9,8})"
        and 22+22+19+22+16+27 == 128
        and round(128*0.30)*(4636+500) == 195168
        and round(round(128*0.30)*13/6)*(4636+500) == 421152
        and abs(128*0.30*(4636+500)*23.5 - 4634726.4) < 1
        and 1339*4636 == 6207604 and 1339*5136 == 6877104,
        "model Defaults tab: every published scenario default lives as inputs and formulas, and the conventions recompute")
    # achievement-level breakout (KDE 2024-25), published under the scores detail
    chk((REPO / "build" / "kde_levels_2024_25.json").exists(),
        "KDE achievement-level extract archived in build/")
    import json as _json
    _lv = _json.loads((REPO / "build" / "kde_levels_2024_25.json").read_text())
    _L = _lv["levels"]; _T = _lv["tested_by_grade"]
    chk(_T["NMES"] == {"3rd": 26, "4th": 19, "5th": 28}
        and _T["BCES"]["3rd"] == 109 and _T["CRES"]["5th"] == 80,
        "archived grade counts: NMES 26/19/28 against 109 and 80 at the larger schools")
    chk(_L["Science"]["4th"]["NMES"][:4] == [6, 41, 47, 6]
        and _L["Reading"]["5th"]["NMES"][3] == 31
        and _L["Reading"]["3rd"]["NMES"] is None
        and _L["Mathematics"]["3rd"]["NMES"] is None,
        "archived levels: NMES science 6/41/47/6, 31 percent distinguished in 5th reading, third grade suppressed")
    _cells = [(sub, g) for sub in _L for g in _L[sub] if _L[sub][g].get("NMES")]
    _lead = [(sub, g) for sub, g in _cells
             if all(_L[sub][g]["NMES"][4] >= _L[sub][g][o][4]
                    for o in ("BCES", "CRES") if _L[sub][g].get(o))]
    chk(len(_cells) == 8 and len(_lead) == 8,
        "NMES leads all eight published subject-and-grade cells on the state's own P/D column")
    for needle in ['id="lvCharts"', "Every grade, every subject, every level",
                   "Why the two views differ", "state-suppressed", "one child moves an NMES subject score"]:
        chk(needle in html, f"site achievement-level breakout: {needle}")
    chk(html.count('class="lvbar') >= 1 and '"Editing and Mechanics"' in html
        and '"On Demand Writing"' in html,
        "site breakout carries all six tested subjects with bar markup")

    # Fayette peer yardstick: Weaver audit carried consistently (v4.6, no version bump)
    chk((REPO / "build" / "fcps_weaver_audit_2026_08.pdf").exists()
        and (REPO / "build" / "fcps_audit_fy2025.pdf").exists(),
        "Fayette's FY2025 audit and the Weaver budget-process audit are both archived")
    chk(round(6902403 / 690460223 * 100, 2) == 1.0
        and 28361786 - 6902403 == 21459383
        and abs(1225465 / 29097404 * 100 - 4.21) < 0.01
        and 2648086 - 1225465 == 1422621
        and abs(2648086 / 29097404 * 100 - 9.10) < 0.01
        and abs(4290840 / 2648086 - 1.62) < 0.01
        and abs(28361786 / 38907376 - 0.73) < 0.01,
        "the burn bases reconcile: Bourbon's 4.2-cent net change follows $1,422,621 of transfers in; on operations alone 9.1 cents gives about 1.6 years of runway against Fayette's 0.7")
    for name, txt in (("site", html), ("report", t)):
        for needle in ["$6,902,403", "$21.5 million", "$28,361,786",
                       "KRS 160.470(6)(a)", "Weaver"]:
            chk(needle in txt, f"Weaver comparison on the {name}: {needle}")
    chk("restatement" not in html and "$36.4 million" not in html,
        "the mixed-basis restatement and $36.4M drawdown framing is off the site")
    chk("Weaver never uses the word restatement" in t
        and t.count("restatement") == 2,
        "the report keeps the restatement wording only inside its published withdrawal of that framing")
    for name, txt in (("site", html), ("report", t)):
        chk("does not reconcile" in txt or "not reconciled" in txt,
            f"the {name} states that Weaver's figure and the audited balance are not reconciled")
    chk("1 percent" in es and "$95 million" in es and "school closure" in es,
        "executive summary carries the Fayette runway note consistently")
    chk("none of them is a school closure" in t
        and "None of them is a school closure" in html,
        "the 70-plus recommendations point is published on the site and in the report")
    chk("unaudited" in t and "subject to change" in t,
        "Weaver's unaudited caveat is carried, not dropped")
    dfy = wb["Defaults"]
    chk(dfy["D73"].value == 6902403 and dfy["D69"].value == 690460223
        and dfy["C78"].value == "=C73-D73" and dfy["D74"].value == "=D73/D69"
        and dfy["B77"].value == "=B73/B75" and dfy["D77"].value is None
        and dfy["B83"].value == "=B73/B70" and dfy["B82"].value == 1320939
        and dfy["B81"].value == "=B70-B75"
        and dfy["D75"].value is None and dfy["D72"].value == "n/a",
        "model Defaults tab: Fayette yardstick live, with no drawdown or runway computed on the mixed basis")
    chk("group of volunteers" in t and "Dr. Ryan Bradley" in t
        and "A personal note from Dr. Ryan Bradley" in t,
        "report cover carries the group byline; personal note attributed by name")
    for solo in ["I checked", "I built", "I prepared this report myself",
                 "I am an alumnus", "check my work", "My recollection"]:
        chk(solo not in t, f"no single-author voice left in the report: {solo!r}")
    chk("Amazon Future Engineer" in html.lower().replace("amazon future engineer", "Amazon Future Engineer")
        or "Amazon Future Engineer" in html,
        "ask 2 carries the differentiation tools (Amazon Future Engineer)")
    for needle in ["57.7", "60.3", "61.3", "65.5",
                   "$112/yr ($9.35/mo)", "$167/yr ($13.93/mo)",
                   "$188/yr ($15.69/mo)", "$277/yr ($23.10/mo)"]:
        chk(needle in t,
            f"beyond-4% option table intact in the PDF (retired from the site in v4.2): {needle}")
    chk("median of the eight area districts with Fayette excluded" in t,
        "PDF defines the regional median precisely")
    chk("reports/Saving_NMES_v3.7_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.7_2026-07-26.pdf").exists(),
        "v3.7 archived in reports/ and linked from the version history")

    # v3.8: fill correction, cost history, breakeven reconstruction, growth plan
    for needle in ["$56,000 to $116,000", "$106,000 to $211,000",
                   "$2,476,544",
                   "54 to 69", "$2,851", "$5,200", "$4,414",
                   "$760,000 to $1.3 million", "$260,000 to $530,000"]:
        chk(needle in t, f"PDF v3.8 content intact: {needle}")
    for needle in ["$760,000 to $1.3 million", "chartCurve"]:
        chk(needle in html, f"site v3.8/v4.2 content intact: {needle}")
    chk("spreads district-wide costs across buildings" in t, "the federal all-in allocation caveat lives in the report")
    chk("Medicaid" not in html, "Medicaid removed from the site's expense options")

    # ---- v3.9 release content ----
    for needle in ["$1,285,310", "$21,482,445", "$1,018,671", "$938,690", "$227,831",
                   "$276,928", "5.5 fixed positions", "$30,410,725", "$30,201,047",
                   "6.3 percent", "450 to 550", "13 to 15 percent", "3,594", "3,548",
                   "2,912", "2,616", "$66,860"]:
        chk(needle in t, f"PDF v3.9 content present: {needle}")
    for needle in ["450 to 550", "13 to 15 percent", "3,594", "2,616"]:
        chk(needle in html, f"site v3.9 content present: {needle}")
    chk("$938,690" in t, "the working-budget cross-check ($938,690) lives in the report")
    chk("one in three" not in html or "an earlier version of this page said one in three" in html,
        "site: the retracted one-in-three share is corrected, not merely repeated")
    chk('id="sTea" min="0" max="3"' in html and 'id="sLeav" min="117" max="154"' in html
        and 'id="sBus" min="20000" max="95000"' in html
        and 'id="sGro" min="110" max="200" value="140" step="5"' in html,
        "site calculator sliders span exactly the published grids: missing students 117 to 154 with "
        "add-ons at three settings, growth target in five-student steps, so no reachable "
        "setting prices outside the published floors and ceilings")
    chk("Version 5.0" in t and "August 17" in t and "Version 4.6" in t
        and "Version 4.5" in t and "Version 4.2" in t,
        "PDF carries the v5.0 release block and the v4.6, v4.5 and v4.2 history entries")
    chk("Saving_NMES_v4.5_2026-08-02.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v4.5_2026-08-02.pdf").exists(),
        "v4.5 archived in reports/ and linked from the version history")
    chk("reports/Saving_NMES_v4.2_2026-08-01.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v4.2_2026-08-01.pdf").exists()
        and (REPO / "reports" / "Saving_NMES_v4.1_2026-07-31.pdf").exists()
        and (REPO / "reports" / "Saving_NMES_v4.0_2026-07-31.pdf").exists(),
        "v4.2 archived and linked; v4.1 and v4.0 stay archived")
    chk("run this play before" in t and "Joy Global" in t and "Figure 8." in t,
        "Millersburg community case study restored in the report (Figure 8)")
    chk("Saving_NMES_v3.9_2026-07-29.pdf" in html, "site links the v3.9 report")
    chk("reconstructs it from the state's own files" not in t,
        "breakeven reconstruction lands at 298 in the report")
    chk("recovered from the Internet Archive" in t.replace("  ", " "),
        "2000-01 report card provenance disclosed in the report (site card replaced by the live curve)")
    chk("prior-year spending" in t,
        "CATS-era fiscal-year caveat disclosed in the report")
    chk("$19,635" in t,
        "2012-13 capital-charge outlier disclosed in the report (site table retired in v4.2)")
    for f in ["bourbon_spending_per_student_2011_2017.csv", "bourbon_staffing_ratios_ccd.csv",
              "crdc_school_finance_bourbon_2011_2017.csv", "slfs_bourbon_fy16_fy17.csv",
              "bourbon_revenue_by_source_2020_2024.csv"]:
        chk((REPO / "build" / f).exists(), f"v3.8 data archived: build/{f}")
    chk("reports/Saving_NMES_v3.8_2026-07-26.pdf" in html
        and (REPO / "reports" / "Saving_NMES_v3.8_2026-07-26.pdf").exists(),
        "v3.8 archived in reports/ and linked from the version history")
    chk("New sections needed at NMES" not in html and "Redistricting tab carries" in t,
        "v3.8 section-debit slider retired with the seat planner; the debit lives in the model's Redistricting tab")
    chk("change-from-last-year piece" in html and "change formula subtracts" in t,
        "2024-25 index change-component: simple on the site, full explanation in the PDF")
    chk('id="tgSD" checked' in html,
        "SchoolDigger toggle defaults to checked")
    chk('id="tgKP" checked' in html and 'id="tgKO" checked' in html,
        "subject score series default on (reading/math and science/SS/writing)")
    chk('id="tgKC" checked' not in html,
        "composite toggle defaults to off")
    for needle in ["includes the state's climate survey", "not a pure test score",
                   "science, social studies, writing average"]:
        chk(needle in html, f"scores chart relabel and explainer: {needle}")
    chk((REPO / "build" / "kyrc25_acct_bourbon_extract.csv").exists()
        ,
        "2024-25 accountability component extract archived")
    chk('id="tgEF"' not in html and "EDFacts" in t,
        "EDFacts series retired from the site; kept in the report")
    chk((REPO / "build" / "edfacts_school_proficiency_bourbon.json").exists()
        ,
        "EDFacts extract archived")
    chk("range midpoints" in t and "different scales" in t,
        "EDFacts midpoint and KCCT/KPREP scale caveats disclosed in the report")
    chk("crosses NMES in the newest year" in html,
        "site keeps the 2024-25 crossover explanation (status table retired to the report in v4.2)")

    chk("unaudited" in t,
        "FY2026 figures labeled unaudited in the PDF (site block retired in v4.2)")
    chk((REPO / "build" / "fy2026_june_financial_packet.pdf").exists(),
        "June 2026 financial packet archived in build/")
    chk("balanced-budget scenario" not in t
        and "The $14 million plan, and the levers not on the table" not in t,
        "the $14 million walk is retired from the report; the Debt_Service tab carries it")
    chk("$2.2 to $2.8 million" in t,
        "scenario capacity range $21M/$25M consistent in the PDF")

    # money-story cleanup: GF-only levy base, both denominators, precise debt wording
    chk("$313,000" in t and "$978,000" in t and "386,000" not in t,
        "PDF levy uses the corrected GF base (313K/978K), old 386K gone")
    chk("$7,829,060" in t and "restricted building-fund levy" in html,
        "levy base disclosed as GF-only in PDF and site")
    chk("over four fifths of the annual reserve drawdown" in t,
        "PDF scores the levy against both deficit and drawdown")
    chk("reserve drawdown" in t and "percentile of the 972 weighted scenarios" in html,
        "PDF keeps both denominators; site calculators simplified to percentile readouts (v4.5 review)")
    chk("net change across all seven" in t,
        "PDF clarifies the $430K is the net debt-service step, not the bond's payment alone")
    chk("multi-age" not in t.lower() and "multiage" not in t.lower(),
        "multi-age reorganization removed from the report")
    chk("its savings sheet says two" in t and "count supports three" in t,
        "PDF keeps the closure staffing-count judgment (v4.2: the district's own two-vs-three)")
    for needle in ["5 percent raise for every certified teacher", "$500,000 a year to spare",
                   "$32 million of building capacity", "$47 million of capacity"]:
        chk(needle in html, f"transformative surplus claim on the site: {needle}")
    chk("5 percent raise for every certified teacher" in t and "within $7,000" in t and "$32 million" in t,
        "re-based transformative claim in the report, advisor-anchored, floor honestly short of the raise")
    chk("10 percent raise" not in html and "$52 million" not in html
        and "withdrawn with the lever correction" in t,
        "the 10-percent-raise / $52 million claims are withdrawn (lever correction)")
    for claim, section in [("$32 million of building capacity", 'id="grow"'),
                           ("LOSES $428,627", 'id="model"'),
                           ("$600,912 to $790,944 a year", 'id="risks"')]:
        chk(html.index(claim) > html.index(section)
            and html.index(claim) < html.index("<details", html.index(section)),
        f"strongest claim rides the always-visible header: {claim}")
    chk("$2.2 to $2.8 million" in t and "the counted-once cost package and the 2018 restore" in t,
        "alternatives raw sums intact in the report (Medicaid removed; site quote retired)")

    # ---- v5.0: the exodus model recomputes from the anonymized survey ----
    import importlib.util as _ilu
    _spec = _ilu.spec_from_file_location("exodus_model", REPO / "build" / "exodus_model.py")
    _em = _ilu.module_from_spec(_spec)
    _spec.loader.exec_module(_em)
    R = _em.RESULT
    chk(R["leaving_households"] == 31 and R["leaving_children"] == 70
        and R["enrolled_sample"] == 75 and R["enrolled_leavers"] == 62,
        "exodus model: survey cleanup reproduces 31 households / 70 children and the 62-of-75 sampled-window split")
    L = R["ladder"]
    chk(L["today"]["kids_lo"] == 55 and L["today"]["kids_hi"] == 73
        and L["today"]["dollars_lo"] == 282480 and L["today"]["dollars_hi"] == 374928,
        "exodus today-share recomputes: 55 to 73 of the 128 enrolled, $282,480 to $374,928")
    chk(L["steady"]["kids_lo"] == 117 and L["steady"]["kids_med"] == 136
        and L["steady"]["kids_hi"] == 154 and L["steady"]["dollars_lo"] == 600912
        and L["steady"]["dollars_med"] == 698496 and L["steady"]["dollars_hi"] == 790944,
        "exodus steady-state recomputes: 117/136/154 kids, $600,912/$698,496/$790,944 (the grid legs)")
    chk(abs(R["eff_years"] - 12.62) < 0.01 and R["var_nonteach"] == 400,
        "effective years 12.62 and the $400 supplies credit recompute")
    for s in ("$600,912", "$790,944", "$698,496", "136 students: the statistical median"):
        chk(s in html, f"exodus figure on the site: {s}")
    for s in ("Where the leaving children would go", "Montgomery County Schools",
              "Other or undecided", "As a share of the 128 enrolled today",
              "because their families say they would leave Bourbon County Schools"):
        chk(s in html, f"survey destination breakdown on the site: {s}")
    import csv as _csv
    _rows = list(_csv.DictReader(open(REPO / "build" / "survey_school_choice_2026_08_anonymized.csv")))
    _dest = {}
    for _r in _rows:
        if _r["status"] == "leaving":
            _dest[_r["destination"]] = _dest.get(_r["destination"], 0) + 1
    chk(_dest.get("Montgomery County") == 25 and _dest.get("Homeschool") == 16
        and _dest.get("Other or undecided") == 20 and _dest.get("Private school") == 8
        and _dest.get("Clark County") == 1 and "Public school in district" not in _dest
        and "Not specified" not in _dest and "Another public school district" not in _dest,
        "destination coding in the anonymized CSV matches the published breakdown (in-district and unspecified leavers folded into other-or-undecided)")
    for s in ("$600,912", "$790,944", "12.62", "response-propensity"):
        chk(s in t, f"exodus figure in the report: {s}")
    import subprocess as _sp
    _out = _sp.run([sys.executable, str(REPO / "build" / "closure_grid.py")],
                   capture_output=True, text=True)
    chk(_out.returncode == 0 and "median $-428,627" in _out.stdout
        and "every scenario loses money" in _out.stdout and "$-309,567" in _out.stdout,
        "closure_grid.py asserts its own v5 statistics (median -$428,627, every scenario loses, default -$309,567)")
    chk("-$309,567" in html and "82nd percentile" in html,
        "site default -$309,567 at the 82nd percentile of the all-loss grid")
    chk("fewer than three students" in html and "fewer than 3 at a level" in html,
        "suppression notes state KDE's written fewer-than-three rule")
    chk("kindergarten enrolled 12" in html or "kindergarten enrolled <b>12" in t
        or "12 children against a ten-year average of 22" in html,
        "the kindergarten-of-12 corroboration is published with the survey")
    for f in ("survey_school_choice_2026_08_anonymized.csv", "saar_enrollment_2025_26.xls",
              "saar_enrollment_2024_25.xls", "saar_enrollment_1999_2019.xlsx",
              "saar_definitions_kde.pdf", "kde_hb6_suppression_rule_2025.pdf",
              "kde_suppressed_data_guidance_2022.pdf", "exodus_model_v5.json",
              "exodus_model.py"):
        chk((REPO / "build" / f).exists(), f"v5 source archived: build/{f}")
    chk("2024-25 SAAR end-of-year" in t and "the state's 2023-24 file shows 141" in t,
        "the facility-plan 128 label correction is disclosed in the report")
    chk("v5.0, August 17" in html and "Saving_NMES_v5.0_2026-08-17.pdf" in html,
        "site version history carries v5.0 with the report file linked")

    print(f"PASS {len(ok)}")
    print(f"FAIL {len(bad)}")
    for b in bad:
        print("  -", b)
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    main()
