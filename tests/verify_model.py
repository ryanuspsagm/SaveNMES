"""Workbook integrity for NMES_Financial_Model.xlsx.

Three passes:
 a. every stored formula tokenizes cleanly and none of them is prose
    parked behind an '=' (the class of defect that once shipped as
    Defaults!F34): a formula containing three or more consecutive
    lowercase words fails;
 b. every reference in every formula resolves: each cross-sheet
    reference names an existing sheet and every referenced single cell
    is non-empty (range references are not emptiness-checked);
 c. note-versus-cell drift: every dollar figure of four or more
    characters quoted inside a plain string cell must trace to a numeric
    cell somewhere in the workbook (within rounding to the dollar), to
    the published site or report (the shared number census), or to the
    explicit whitelist below, each entry carrying its justification.

Run:  python tests/verify_model.py
Needs: pip install openpyxl pypdfium2
Exits nonzero if any check fails.
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer

sys.path.insert(0, str(Path(__file__).resolve().parent))
import number_census

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "NMES_Financial_Model.xlsx"

ok, bad = [], []
def chk(cond, label): (ok if cond else bad).append(label)

# ---- pass a: prose heuristic ------------------------------------------------
PROSE_RE = re.compile(r"\b[a-z]+ [a-z]+ [a-z]+\b")

# ---- pass b: reference grammar ----------------------------------------------
CELL_RE = re.compile(r"^[A-Z]{1,3}\d+$")
RANGE_RE = re.compile(r"^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$")

# Referenced single cells that are legitimately empty. The current
# workbook has none: every single-cell reference in all 557 formulas
# lands on a populated cell, so nothing is whitelisted here. Add
# "Sheet!A1" strings only after confirming the formula tolerates the
# blank.
EMPTY_REF_OK = set()

# ---- pass c: exclusions and whitelist ---------------------------------------
# Historical and changelog-style note cells, excluded by rule:
#  1. KY_Closures column J: per-case notes quoting dollar figures from
#     external closure records (state archives and news accounts), not
#     from this model's own cells.
#  2. Cells carrying an explicit changelog marker: they record what a
#     PRIOR version said and deliberately keep the superseded figure
#     next to the correction (version-history strings, LEGACY-style
#     kept-for-the-record rows).
CHANGELOG_MARKERS = ("withdrawn in v", "corrected in v", "superseded",
                     "through the prior release", "kept for the record",
                     "An earlier version")

# Quoted figures that are real but never live in a cell or on the
# site/report. Keys are normalized (no $ or commas); each value is the
# justification, printed on failure and kept one line per entry.
WHITELIST = {
    # quoted from the district's own documents (budget, ledger, packet, slides)
    "132744": "working-budget function-view school administration line, as printed in the district budget",
    "96107": "General Fund share of MUNIS function 2600 plant operations, as printed on the ledger",
    "119909": "all-funds plant operations from the same working-budget note",
    "49097": "instructional staff support (library and media) line from the working budget",
    "228851": "function-view administration plus plant sum quoted in the same note",
    "71447": "published certified salary schedule Rank I top step (external schedule)",
    "10000388": "General Fund certified payroll behind the 5 percent raise pricing (district payroll records)",
    "209700": "district Response Appendix A.1's own pricing of the four fixed roles",
    "108958.80": "Appendix A.1's printed 'Elementary Teachers: 2' line, quoted verbatim (2 x 54,479.40)",
    "8902321": "Cane Ridge total site spending on the 2023-24 state filing (external filing)",
    "82866": "KDE nickel equalization phase-in for FY2025, from the state SEEK schedules",
    "55515": "KDE nickel equalization phase-in for FY2026, from the state SEEK schedules",
    "153900": "months-1-to-11 miscellaneous revenue run rate from the district's own packet",
    "1098663": "Building Fund (320) component of the $1,320,939 sweep, per the GF receipts ledger",
    "222276": "Capital Outlay (310) component of the same sweep, per the GF receipts ledger",
    "1098633": "the packet's fund 320 page prints this figure; quoted to document the district's own $30 internal discrepancy",
    "1120203": "the Building Fund transfer as budgeted in September 2025 (district budget document)",
    "4079193": "Budget Monitoring Tool projected balance, from the district's own tool",
    "18600946": "the 2026 draft DFP's printed high school renovation cost (district slides)",
    "9421": "the printed component-versus-total gap on the district's own draft-plan slides",
    "8305": "the state revenue definition nearest the response's implied per-student revenue",
    "8255": "the district response's implied revenue per student (300-student breakeven reconstruction)",
    "8855": "local plus federal revenue per student from the state files",
    "166315": "Millersburg cost per student in FY2001 dollars, from the state archive",
    "155776": "the matching modern per-student figure in the same note",
    # derived side figures the notes compute in passing (not stored as cells)
    "118000": "rounded grid outcome for 30 students at the central busing and supplies legs",
    "409446": "the 972-scenario closure grid's central-case loss, produced by build/closure_grid.py",
    "28500": "growth grid support-staff mid cost leg; lives on the site as comma-free JS (SCOSTV)",
    "4726": "net SEEK value per added student at the default legs (4,626 + 500 - 400)",
    "64000": "rounded corner-case output of the redistricting levers, noted for the record",
    "3670": "rounded class-cap marginal cost per absorbed student",
    "4400": "rounded per-student cost margin between the three schools after the move",
    "18474": "derived per-student bound at the draft plan's 154 rating",
    "259888": "NMES year-over-year total cost change from the state filings (extreme bound)",
    "16243": "per-student form of the same year-over-year change",
    "18527": "derived extreme-bound per-student cost at the approved 174 rating",
    "3328472": "June 2026 GL packet fund balance before transfers (cent-level version verified in validate_all)",
    "2954484": "June 2026 GL packet fund balance after transfers",
    "1409590": "June 2026 GL packet transfers in (cent-level version verified in validate_all)",
    "313162": "levy compounding year 1 on the levy tab's own base",
    "977568": "levy compounding cumulative year 3 on the same base",
    "7558635": "computed GF tax yield on the certified assessment (derivation note beside the actual collections)",
}

FIG_RE = re.compile(r"\$[\d,]{4,}(?:\.\d\d)?")


def main():
    wb = load_workbook(XLSX, data_only=False)
    wbv = load_workbook(XLSX, data_only=True)
    sheetnames = set(wb.sheetnames)
    defined = set(wb.defined_names)

    formulas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.data_type == "f" or (isinstance(c.value, str)
                                          and str(c.value).startswith("=")):
                    formulas.append((ws.title, c.coordinate, str(c.value)))
    chk(len(formulas) >= 500, f"workbook holds its formulas ({len(formulas)})")

    # ---- a: every formula tokenizes; none is prose behind '=' ----
    tok_fail, prose = [], []
    for sheet, coord, f in formulas:
        try:
            Tokenizer(f)
        except Exception as exc:
            tok_fail.append(f"{sheet}!{coord}: {exc}")
            continue
        if PROSE_RE.search(f):
            prose.append(f"{sheet}!{coord}: {f[:80]!r}")
    chk(not tok_fail,
        f"all {len(formulas)} formulas tokenize"
        if not tok_fail else f"formulas fail to tokenize: {tok_fail}")
    chk(not prose,
        "no formula is prose behind '=' (the Defaults!F34 class)"
        if not prose else f"prose stored as formulas: {prose}")

    # ---- b: every reference resolves ----
    bad_sheet, empty_ref, odd = [], [], []
    checked_refs = 0
    for sheet, coord, f in formulas:
        try:
            items = Tokenizer(f).items
        except Exception:
            continue  # already reported above
        for tok in items:
            if tok.type != "OPERAND" or tok.subtype != "RANGE":
                continue
            v = tok.value
            if "!" in v:
                sn, rng = v.rsplit("!", 1)
                sn = sn.strip("'")
            else:
                sn, rng = sheet, v
            if sn not in sheetnames:
                bad_sheet.append(f"{sheet}!{coord} -> {sn}")
                continue
            rng = rng.replace("$", "")
            if RANGE_RE.match(rng):
                continue  # multi-cell ranges: existence checked, emptiness not
            if not CELL_RE.match(rng):
                if v in defined or rng in defined:
                    continue  # defined name, resolves by definition
                odd.append(f"{sheet}!{coord} -> {v}")
                continue
            checked_refs += 1
            if wb[sn][rng].value is None and f"{sn}!{rng}" not in EMPTY_REF_OK:
                empty_ref.append(f"{sheet}!{coord} -> {sn}!{rng}")
    chk(not bad_sheet,
        "every cross-sheet reference names an existing sheet"
        if not bad_sheet else f"references to missing sheets: {bad_sheet}")
    chk(not odd,
        "every range operand parses as a cell, range, or defined name"
        if not odd else f"unresolvable operands: {odd}")
    chk(not empty_ref,
        f"every referenced single cell is non-empty ({checked_refs} checked)"
        if not empty_ref else f"formulas referencing empty cells: {empty_ref}")

    # ---- c: note-versus-cell drift ----
    numeric = set()
    for ws in wbv.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    numeric.add(round(float(c.value)))
    census = number_census.site_numbers() | number_census.report_numbers()

    drift, figures = [], 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                if ws.title == "KY_Closures" and c.column_letter == "J":
                    continue  # exclusion rule 1: external-record case notes
                if any(m in v for m in CHANGELOG_MARKERS):
                    continue  # exclusion rule 2: changelog-style notes
                for m in FIG_RE.finditer(v):
                    plain = m.group().lstrip("$").replace(",", "")
                    if not plain.strip("."):
                        continue
                    figures += 1
                    if round(float(plain)) in numeric:
                        continue  # (i) lives in a cell, to the dollar
                    if plain in census:
                        continue  # (ii) published on the site or in the report
                    if plain in WHITELIST:
                        continue  # (iii) whitelisted with a justification
                    drift.append(f"{ws.title}!{c.coordinate} quotes ${plain}"
                                 f" :: {v[:70]!r}")
    chk(not drift,
        f"note-versus-cell drift: all {figures} quoted dollar figures trace"
        if not drift else "untraceable quoted figures:\n    " + "\n    ".join(drift))

    print(f"PASS {len(ok)}")
    print(f"FAIL {len(bad)}")
    for b in bad:
        print("  -", b)
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    main()
