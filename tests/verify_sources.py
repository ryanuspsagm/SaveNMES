"""Primary-source verification: every load-bearing constant in
build/sources.json is re-extracted from its archived source and compared
to the manifest, then compared to where the constant lives in code.

This layer exists to kill the internally-consistent-but-wrong failure
class: a wrong SEEK base ($4,636) once passed 741 consistency checks
because nothing compared constants to primary sources. The enacted value
is $4,626 (build/ky_acts_2026_ch168_hb500.pdf, page 20).

Run:  python tests/verify_sources.py
Needs: pip install pypdfium2 openpyxl xlrd
Entries marked verify:false in the manifest are reported as SKIP with
their reason, never silently. Exits nonzero if any check fails.
"""
import csv
import json
import re
import sys
from pathlib import Path

import xlrd
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]

ok, bad, skip = [], [], []
def chk(cond, label): (ok if cond else bad).append(label)


_pdf_cache = {}
def pdf_page_text(path, page):
    """Text of a 1-based PDF page, whitespace collapsed to single spaces."""
    key = (str(path), page)
    if key not in _pdf_cache:
        import pypdfium2 as pdfium
        doc = pdfium.PdfDocument(str(path))
        raw = doc[page - 1].get_textpage().get_text_range()
        _pdf_cache[key] = re.sub(r"\s+", " ", raw)
    return _pdf_cache[key]


def numeric(s):
    return float(str(s).replace(",", "").replace("_", ""))


def close(a, b):
    if isinstance(a, list) or isinstance(b, list):
        return (isinstance(a, list) and isinstance(b, list) and len(a) == len(b)
                and all(close(x, y) for x, y in zip(a, b)))
    return abs(float(a) - float(b)) < 0.005


def json_get(obj, dotted):
    for part in dotted.split("."):
        obj = obj[part]
    return obj


def xls_row_lookup(path, sheet, row_contains, column_header):
    sh = xlrd.open_workbook(str(path)).sheet_by_name(sheet)
    header = None
    for r in range(sh.nrows):
        row = [sh.cell_value(r, c) for c in range(sh.ncols)]
        if header is None and str(row[0]).strip() == "School":
            header = [str(v).strip() for v in row]
            continue
        if header and row_contains in str(row[0]):
            return numeric(row[header.index(column_header)])
    raise LookupError(f"{row_contains} not found in {path.name}:{sheet}")


def xlsx_first_numeric_after_name(path, sheet, name):
    """Row holding `name` in any cell: the first numeric value after that
    cell is the first grade column (EL / Grade 00) in the 1999-2019 file."""
    ws = load_workbook(str(path), read_only=True)[sheet]
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row):
            if isinstance(v, str) and name in v:
                for w in row[i + 1:]:
                    if isinstance(w, (int, float)):
                        return float(w)
    raise LookupError(f"{name} not found in {path.name}:{sheet}")


def extract(entry):
    """Pull the value out of the archived source per the manifest spec."""
    spec = entry["extraction"]
    kind = spec["type"]
    src = REPO / entry["source_file"] if entry["source_file"] != "external" else None

    if kind == "pdf_regex":
        txt = pdf_page_text(src, spec["page"])
        m = re.search(spec["regex"], txt)
        if not m:
            raise LookupError(f"regex not found on page {spec['page']} of {src.name}")
        return numeric(m.group(1))

    if kind == "json_key":
        return json_get(json.load(open(src)), spec["key"])

    if kind == "levy_split":
        txt = pdf_page_text(src, spec["page"])
        total = numeric(re.search(
            r"total real estate and personal property valuation of \$([\d,]+)", txt).group(1))
        yield_ = numeric(re.search(
            r"calculated collection should yield \$([\d,]+)", txt).group(1))
        rate_re = float(re.search(r"real estate rate of ([\d.]+)", txt).group(1)) / 10000
        rate_tp = float(re.search(r"personal property rate of ([\d.]+) cents", txt).group(1)) / 10000
        tangible = round((yield_ - rate_re * total) / (rate_tp - rate_re))
        real = round(total - tangible)
        # the split must reproduce the audit's own printed yield to the dollar
        if abs(rate_re * real + rate_tp * tangible - yield_) >= 2:
            raise LookupError("levy split does not reproduce the printed yield")
        return real if spec["component"] == "real_estate" else tangible

    if kind == "xls_row_lookup":
        return xls_row_lookup(src, spec["sheet"], spec["row_contains"], spec["column_header"])

    if kind == "saar_k_series":
        out = []
        for pt in spec["points"]:
            f = REPO / pt["file"]
            if f.suffix == ".xlsx":
                out.append(xlsx_first_numeric_after_name(
                    f, pt["sheet"], "North Middletown Elementary"))
            else:
                out.append(xls_row_lookup(f, pt["sheet"], "North Middletown", "Grade 00"))
        return out

    if kind == "xlsx_cell":
        ws = load_workbook(str(src), read_only=True, data_only=True)[spec["sheet"]]
        label = ws[spec["row_label_cell"]].value
        if spec["row_label_contains"] not in str(label):
            raise LookupError(f"{spec['row_label_cell']} is {label!r}, "
                              f"expected it to contain {spec['row_label_contains']!r}")
        return ws[spec["cell"]].value

    if kind == "xlsx_filter_sum":
        ws = load_workbook(str(src), read_only=True, data_only=True)[spec["sheet"]]
        mi, si = spec["match_column"] - 1, spec["sum_column"] - 1
        total = 0
        for row in ws.iter_rows(values_only=True):
            if len(row) > si and row[mi] == spec["match_value"] \
                    and isinstance(row[si], (int, float)):
                total += row[si]
        return total

    if kind == "csv_lookup":
        for r in csv.DictReader(open(src)):
            if all(r[k] == v for k, v in spec["where"].items()):
                return numeric(r[spec["column"]])
        raise LookupError(f"no row matches {spec['where']} in {src.name}")

    if kind == "csv_row_count":
        return len(list(csv.DictReader(open(src))))

    if kind == "csv_distinct_count":
        return len({r[spec["column"]] for r in csv.DictReader(open(src))})

    if kind == "csv_count_where":
        return sum(1 for r in csv.DictReader(open(src))
                   if all(r[k] == v for k, v in spec["where"].items()))

    raise LookupError(f"unknown extraction type {kind!r}")


def check_code_ref(entry, ref):
    """The constant as assigned in code must match the manifest value.
    SEEK may migrate into build/constants.py; if the named file no longer
    carries the assignment but constants.py does, the check follows it."""
    path = REPO / ref["file"]
    text = path.read_text()
    m = re.search(ref["regex"], text, re.MULTILINE)
    where = ref["file"]
    if not m and entry["name"].startswith("seek_base") and (REPO / "build" / "constants.py").exists():
        text = (REPO / "build" / "constants.py").read_text()
        m = re.search(r"^SEEK\s*=\s*(\d+)\b", text, re.MULTILINE)
        where = "build/constants.py (assignment moved)"
    if not m:
        chk(False, f"{entry['name']}: assignment found in {where}")
        return
    got = numeric(m.group(ref.get("group", 1)))
    if ref.get("transform") == "half":
        got = got / 2
    chk(close(got, entry["value"]),
        f"{entry['name']}: code constant in {where} = {entry['value']}")


def main():
    manifest = json.load(open(REPO / "build" / "sources.json"))
    entries = manifest["entries"]
    by_name = {e["name"]: e for e in entries}

    # phase 1: every machine-checkable value re-extracted from its source
    for e in entries:
        if not e["verify"]:
            skip.append(f"{e['name']} = {e['value']}: {e.get('note', 'no machine-checkable source')}")
        else:
            try:
                got = extract(e)
                chk(close(got, e["value"]),
                    f"{e['name']}: source {e['source_file']} yields {e['value']} (got {got})")
            except Exception as exc:
                chk(False, f"{e['name']}: extraction failed ({exc})")

        # phase 2: the constant as it lives in code, even for skipped sources
        for ref in e.get("code_refs", []):
            check_code_ref(e, ref)

    # cross-entry relations the models and validate_all.py rely on
    chk(close(by_name["teacher_loaded_cost_pair"]["value"],
              2 * by_name["teacher_loaded_cost"]["value"]),
        "teacher pair is exactly twice the single loaded cost")
    chk(close(by_name["fixed_positions_base"]["value"],
              by_name["fixed_pos_school_administration"]["value"]
              + by_name["fixed_pos_custodial"]["value"]
              + by_name["fixed_pos_library"]["value"]),
        "fixed-position base equals the sum of its three components")
    chk(close(by_name["audit_total_valuation"]["value"],
              by_name["audit_real_estate_valuation"]["value"]
              + by_name["audit_tangible_valuation"]["value"]),
        "valuation split sums to the audit's printed total")
    k = by_name["saar_kindergarten_series"]["value"]
    chk(close(round(sum(k) / len(k), 2), 22.25),
        "kindergarten series averages 22.25, the ten-year average of 22 cited on the site")

    # if the constants ever consolidate, the consolidated file must agree too
    cpath = REPO / "build" / "constants.py"
    if cpath.exists():
        m = re.search(r"^SEEK\s*=\s*(\d+)\b", cpath.read_text(), re.MULTILINE)
        chk(m is not None and numeric(m.group(1)) == by_name["seek_base_fy2027"]["value"],
            "build/constants.py SEEK matches the enacted FY2027 base")

    print(f"PASS {len(ok)}")
    print(f"SKIP {len(skip)}")
    for s in skip:
        print("  ~", s)
    print(f"FAIL {len(bad)}")
    for b in bad:
        print("  -", b)
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    main()
