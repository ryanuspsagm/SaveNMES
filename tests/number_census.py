"""Number census across the four public artifacts.

Extracts every dollar figure and every large comma-separated number from
the site (index.html, tags stripped, aria-labels and inline JS included),
the text layer of the full report and the executive summary (pypdfium2),
and every string cell of the financial model workbook. Numbers are
normalized (no $ or commas) and deduplicated.

Asserts:
  a. every number in the executive summary also appears in the report,
     the site, or the workbook (the summary invents nothing);
  b. the full census (sorted unique numbers with a per-artifact presence
     map) equals the committed tests/census_baseline.json, so any number
     added to or removed from any artifact fails with a readable diff
     until the baseline is deliberately regenerated.

Run:    python tests/number_census.py
Update: python tests/number_census.py --update   (rewrites the baseline)
Needs:  pip install openpyxl pypdfium2
Exits nonzero if any check fails.
"""
import json
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
BASELINE = Path(__file__).resolve().parent / "census_baseline.json"

# Dollar figures like $4,626 / $54,479.40 and bare numbers with thousands
# separators like 1,479,078 or 2,174.3. The lookarounds keep a match from
# starting or ending inside a larger number.
DOLLAR_RE = re.compile(r"\$[\d,]+(?:\.\d\d)?")
BARE_RE = re.compile(r"(?<![\d,.$])\d{1,3}(?:,\d{3})+(?:\.\d+)?(?![\d,])")

ok, bad = [], []
def chk(cond, label): (ok if cond else bad).append(label)


def normalize(token):
    """$1,479,078 -> 1479078; 54,479.40 -> 54479.40."""
    return token.lstrip("$").replace(",", "")


def extract(text):
    out = set()
    for rx in (DOLLAR_RE, BARE_RE):
        for m in rx.finditer(text):
            v = normalize(m.group())
            if v and v.strip("."):
                out.add(v)
    return out


def site_text():
    html = (REPO / "index.html").read_text()
    # aria-label values sit inside tags and would be lost to the strip
    aria = " ".join(re.findall(r'aria-label="([^"]*)"', html))
    # stripping tags keeps script bodies, so the inline JS is included
    return re.sub(r"<[^>]+>", " ", html) + " " + aria


def pdf_text(path):
    import pypdfium2 as pdfium
    doc = pdfium.PdfDocument(str(path))
    return "\n".join(pg.get_textpage().get_text_range() for pg in doc)


def workbook_text():
    from openpyxl import load_workbook
    wb = load_workbook(REPO / "NMES_Financial_Model.xlsx", data_only=False)
    return "\n".join(c.value for ws in wb.worksheets for row in ws.iter_rows()
                     for c in row if isinstance(c.value, str))


def site_numbers():
    return extract(site_text())


def report_numbers():
    return extract(pdf_text(REPO / "Saving_North_Middletown_Elementary.pdf"))


def summary_numbers():
    return extract(pdf_text(REPO / "SaveNMES_Executive_Summary.pdf"))


def workbook_numbers():
    return extract(workbook_text())


def census():
    """Presence map: normalized number -> sorted list of artifacts."""
    arts = {"site": site_numbers(), "report": report_numbers(),
            "summary": summary_numbers(), "workbook": workbook_numbers()}
    every = sorted(set().union(*arts.values()), key=lambda v: (float(v), v))
    return {n: sorted(a for a in arts if n in arts[a]) for n in every}, arts


def main():
    update = "--update" in sys.argv[1:]
    table, arts = census()

    # (a) the executive summary invents no numbers
    grounded = arts["report"] | arts["site"] | arts["workbook"]
    loose = sorted(arts["summary"] - grounded, key=lambda v: (float(v), v))
    chk(not loose,
        f"every summary number appears in report, site, or workbook "
        f"({len(arts['summary'])} checked)" if not loose else
        f"summary numbers missing everywhere else: {loose}")

    # (b) the census matches the committed baseline exactly
    if update:
        BASELINE.write_text(json.dumps({"numbers": table}, indent=1) + "\n")
        print(f"baseline rewritten: {len(table)} numbers -> {BASELINE}")
    if not BASELINE.exists():
        chk(False, "census baseline missing; run with --update to create it")
    else:
        base = json.loads(BASELINE.read_text())["numbers"]
        added = sorted(set(table) - set(base), key=lambda v: (float(v), v))
        removed = sorted(set(base) - set(table), key=lambda v: (float(v), v))
        moved = sorted(n for n in set(table) & set(base) if table[n] != base[n])
        if added or removed or moved:
            print("census drift against tests/census_baseline.json:")
            for n in added:
                print(f"  + {n} (new, in {','.join(table[n])})")
            for n in removed:
                print(f"  - {n} (gone, was in {','.join(base[n])})")
            for n in moved:
                print(f"  ~ {n} (was {','.join(base[n])}, now {','.join(table[n])})")
            print("  deliberate change? rerun with --update to accept it")
        chk(not (added or removed or moved),
            f"census matches baseline ({len(table)} numbers)" if not
            (added or removed or moved) else
            f"census drifted: +{len(added)} -{len(removed)} ~{len(moved)}")

    print(f"PASS {len(ok)}")
    print(f"FAIL {len(bad)}")
    for b in bad:
        print("  -", b)
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    main()
